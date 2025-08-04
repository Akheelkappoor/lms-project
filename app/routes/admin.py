from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, current_app
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
import os
from app.utils.helper import upload_file_to_s3
import json
from flask_wtf.csrf import generate_csrf
from app import db
from app.models.user import User
from app.models.department import Department  
from app.models.tutor import Tutor
from app.models.student import Student
from app.models.class_model import Class
from app.models.attendance import Attendance
from app.forms.user import CreateUserForm, EditUserForm, TutorRegistrationForm, StudentRegistrationForm
from app.utils.tutor_matching import TutorMatchingEngine, SearchQueryProcessor, AvailabilityChecker, monitor_search_performance
from functools import wraps
from app.utils.email import send_password_reset_email, send_onboarding_email
from app.forms.user import EditStudentForm
from sqlalchemy import text
from app.utils.enhanced_email_subjects import create_better_email_subject, get_enhanced_subject_options
import urllib.parse
from flask import make_response
from app.utils.allocation_helper import allocation_helper
from sqlalchemy import or_, and_, func, text

from app.utils.advanced_permissions import (
    require_permission, 
    require_any_permission, 
    require_all_permissions,
    require_role,
    PermissionRegistry,
    PermissionUtils
)

bp = Blueprint('admin', __name__)

matching_engine = TutorMatchingEngine()

def admin_required(f):
    """Decorator to require admin access"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role not in ['superadmin', 'admin', 'coordinator']:
            flash('Access denied. Insufficient permissions.', 'error')
            return redirect(url_for('dashboard.index'))
        return f(*args, **kwargs)
    return decorated_function

def save_uploaded_file(file, subfolder):
    """Save uploaded file and return filename"""
    if file and file.filename:
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
        filename = timestamp + filename
        
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], subfolder, filename)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        file.save(file_path)
        return filename
    return None


@bp.route('/users')
@login_required
@require_permission('user_management')
def users():
    """User management page"""
    page = request.args.get('page', 1, type=int)
    role_filter = request.args.get('role', '')
    dept_filter = request.args.get('department', '', type=int)
    search = request.args.get('search', '')
    
    query = User.query
    
    if role_filter:
        query = query.filter_by(role=role_filter)
    
    if dept_filter:
        query = query.filter_by(department_id=dept_filter)
    
    if search:
        query = query.filter(
            (User.full_name.contains(search)) |
            (User.username.contains(search)) |
            (User.email.contains(search))
        )
    
    users = query.order_by(User.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    departments = Department.query.filter_by(is_active=True).all()

    filtered_args = request.args.to_dict()
    filtered_args.pop('page', None)

    return render_template('admin/users.html', users=users, departments=departments, filtered_args=filtered_args)

@bp.route('/users/create', methods=['GET', 'POST'])
@login_required
@require_permission('user_management')
def create_user():
    """Create new user - DEBUG VERSION WITH ONBOARDING EMAIL"""
    form = CreateUserForm()
    
    if request.method == 'POST':
        print("=== POST REQUEST RECEIVED ===")
        print(f"Form data: {dict(request.form)}")
        print(f"Files: {dict(request.files)}")
        
        print(f"Form validates: {form.validate()}")
        print(f"Form errors: {form.errors}")
        
        if form.validate_on_submit():
            print("=== FORM VALIDATION PASSED ===")
            try:
                print("Creating user object...")
                
                # Store password before hashing for email
                plain_password = form.password.data
                print(f"Plain password stored for email: {plain_password}")
                
                # Create user object
                user = User(
                    username=form.username.data,
                    email=form.email.data,
                    full_name=form.full_name.data,
                    phone=form.phone.data,
                    role=form.role.data,
                    department_id=form.department_id.data if form.department_id.data else None,
                    address=form.address.data,
                    working_hours=form.working_hours.data,
                    joining_date=form.joining_date.data or datetime.utcnow().date(),
                    is_active=form.is_active.data,
                    is_verified=True,
                    created_at=datetime.utcnow()
                )
                
                print(f"User object created: {user.username}")
                
                # Set password
                user.set_password(plain_password)
                print("Password set")
                
                # Handle profile picture upload
                if form.profile_picture.data:
                    print("Processing profile picture...")
                    s3_url = upload_file_to_s3(form.profile_picture.data, folder=f"{current_app.config['UPLOAD_FOLDER']}/profiles")
                    user.profile_picture = s3_url
                    print(f"Profile picture saved: {s3_url}")
                
                # Handle emergency contact
                emergency_contact = {
                    'name': request.form.get('emergency_name', ''),
                    'phone': request.form.get('emergency_phone', ''),
                    'relation': request.form.get('emergency_relation', '')
                }
                if emergency_contact['name']:
                    user.set_emergency_contact(emergency_contact)
                    print(f"Emergency contact set: {emergency_contact}")
                
                print("About to add to database...")
                # Save to database
                db.session.add(user)
                print("User added to session")
                
                db.session.commit()
                print("DATABASE COMMIT SUCCESSFUL!")
                
                # Send onboarding email
                try:
                    print("=== SENDING ONBOARDING EMAIL ===")
                    print(f"Email: {user.email}")
                    print(f"User: {user.full_name}")
                    print(f"Role: {user.role}")
                    
                    send_onboarding_email(user, plain_password)
                    print("=== ONBOARDING EMAIL SENT SUCCESSFULLY ===")
                    
                    flash(f'User {user.full_name} created successfully! Onboarding email sent to {user.email}.', 'success')
                    
                except Exception as email_error:
                    print(f"=== EMAIL SENDING FAILED ===")
                    print(f"Email error type: {type(email_error)}")
                    print(f"Email error message: {str(email_error)}")
                    import traceback
                    print(f"Email traceback: {traceback.format_exc()}")
                    
                    flash(f'User {user.full_name} created successfully, but failed to send onboarding email. Please send credentials manually to {user.email}.', 'warning')
                
                print("=== SUCCESS: REDIRECTING ===")
                return redirect(url_for('admin.users'))
                
            except Exception as e:
                print(f"=== EXCEPTION OCCURRED ===")
                print(f"Exception type: {type(e)}")
                print(f"Exception message: {str(e)}")
                print(f"Exception details: {repr(e)}")
                import traceback
                print(f"Traceback: {traceback.format_exc()}")
                
                db.session.rollback()
                flash(f'Error creating user: {str(e)}', 'error')
        else:
            print("=== FORM VALIDATION FAILED ===")
            print(f"Validation errors: {form.errors}")
    
    print("=== RETURNING TEMPLATE ===")
    return render_template('admin/create_user.html', form=form)

@bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@require_permission('user_management')
def edit_user(user_id):
    """Edit user"""
    user = User.query.get_or_404(user_id)
    form = EditUserForm(user_id=user.id, obj=user)
    
    if form.validate_on_submit():
        try:
            # Update user fields
            user.full_name = form.full_name.data
            user.email = form.email.data
            user.phone = form.phone.data
            user.role = form.role.data
            user.department_id = form.department_id.data if form.department_id.data else None
            user.address = form.address.data
            user.working_hours = form.working_hours.data
            user.joining_date = form.joining_date.data
            user.is_active = form.is_active.data
            
            # Handle profile picture upload
            if form.profile_picture.data:
                s3_url = upload_file_to_s3(form.profile_picture.data, folder=f"{current_app.config['UPLOAD_FOLDER']}/profiles")
                user.profile_picture = s3_url
            
            db.session.commit()
            flash(f'User {user.full_name} updated successfully!', 'success')
            return redirect(url_for('admin.users'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating user: {str(e)}', 'error')
    
    return render_template('admin/edit_user.html', form=form, user=user)

@bp.route('/users/<int:user_id>/toggle-status', methods=['POST'])
@login_required
@require_permission('user_management')
def toggle_user_status(user_id):
    """Toggle user active status"""
    user = User.query.get_or_404(user_id)
    
    if user.role == 'superadmin' and current_user.role != 'superadmin':
        return jsonify({'error': 'Cannot modify superadmin account'}), 403
    
    try:
        user.is_active = not user.is_active
        db.session.commit()
        
        status = 'activated' if user.is_active else 'deactivated'
        flash(f'User {user.full_name} has been {status}.', 'success')
        
        return jsonify({'success': True, 'status': user.is_active})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Error updating user status'}), 500

@bp.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
@require_permission('user_management')
def delete_user(user_id):
    """Delete user (superadmin only)"""
    if current_user.role != 'superadmin':
        return jsonify({'error': 'Only superadmin can delete users'}), 403
    
    user = User.query.get_or_404(user_id)
    
    if user.role == 'superadmin':
        return jsonify({'error': 'Cannot delete superadmin account'}), 403
    
    try:
        # Check dependencies before deleting
        if user.role == 'tutor' and user.tutor_profile:
            if Class.query.filter_by(tutor_id=user.tutor_profile.id).first():
                return jsonify({'error': 'Cannot delete tutor with existing classes'}), 400
        
        db.session.delete(user)
        db.session.commit()
        
        flash(f'User {user.full_name} deleted successfully.', 'success')
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Error deleting user'}), 500

# ============ DEPARTMENT MANAGEMENT ROUTES ============

@bp.route('/departments')
@login_required
@require_role('superadmin', 'admin')
def departments():
    """Department management page"""
    departments = Department.query.order_by(Department.created_at.desc()).all()
    return render_template('admin/departments.html', departments=departments)

@bp.route('/departments/create', methods=['POST'])
@login_required
@require_role('superadmin', 'admin')
def create_department():
    """Create new department"""
    if current_user.role not in ['superadmin', 'admin']:
        return jsonify({'error': 'Insufficient permissions'}), 403
    
    data = request.get_json()
    
    try:
        department = Department(
            name=data['name'],
            code=data['code'],
            description=data.get('description', ''),
            created_by=current_user.id
        )
        
        db.session.add(department)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Department created successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Error creating department'}), 500

@bp.route('/departments/<int:dept_id>/permissions', methods=['GET', 'POST'])
@login_required
@require_role('superadmin', 'admin')
def department_permissions(dept_id):
    """Manage department permissions"""
    if current_user.role not in ['superadmin', 'admin']:
        flash('Insufficient permissions.', 'error')
        return redirect(url_for('admin.departments'))
    
    department = Department.query.get_or_404(dept_id)
    
    if request.method == 'POST':
        permissions = request.form.getlist('permissions')
        department.set_permissions(permissions)
        db.session.commit()
        flash(f'Permissions updated for {department.name}', 'success')
        return redirect(url_for('admin.departments'))
    
    all_permissions = [
        'user_management', 'tutor_management', 'student_management',
        'class_management', 'attendance_management', 'schedule_management',
        'demo_management', 'finance_management', 'report_generation',
        'communication', 'profile_management', 'quality_checking'
    ]
    
    return render_template('admin/department_permissions.html', 
                         department=department, all_permissions=all_permissions)

@bp.route('/departments/<int:dept_id>/data')
@login_required
@require_role('superadmin', 'admin')
def get_department_data(dept_id):
    """Get department data for editing"""
    department = Department.query.get_or_404(dept_id)
    return jsonify({
        'success': True,
        'department': {
            'id': department.id,
            'name': department.name,
            'code': department.code,
            'description': department.description
        }
    })

@bp.route('/departments/<int:dept_id>/update', methods=['POST'])
@login_required
@require_role('superadmin', 'admin')
def update_department(dept_id):
    """Update department"""
    department = Department.query.get_or_404(dept_id)
    data = request.get_json()
    
    try:
        department.name = data['name']
        department.code = data['code']
        department.description = data.get('description', '')
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Department updated successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Error updating department'}), 500

@bp.route('/departments/<int:dept_id>/toggle-status', methods=['POST'])
@login_required
@require_role('superadmin', 'admin')
def toggle_department_status(dept_id):
    """Toggle department active status"""
    department = Department.query.get_or_404(dept_id)
    
    try:
        department.is_active = not department.is_active
        db.session.commit()
        
        status = 'activated' if department.is_active else 'deactivated'
        return jsonify({'success': True, 'message': f'Department {status} successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Error updating department status'}), 500

# ============ TUTOR MANAGEMENT ROUTES ============

@bp.route('/tutors')
@login_required
@require_permission('tutor_management')
def tutors():
    """Tutor management page"""
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', '')
    dept_filter = request.args.get('department', '', type=int)
    search = request.args.get('search', '')
    
    query = Tutor.query.join(User)
    
    if status_filter:
        query = query.filter(Tutor.status == status_filter)
    
    if dept_filter:
        query = query.filter(User.department_id == dept_filter)
    
    if search:
        query = query.filter(User.full_name.contains(search))
    
    tutors = query.order_by(Tutor.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    departments = Department.query.filter_by(is_active=True).all()

    filtered_args = request.args.to_dict()
    filtered_args.pop('page', None)  
    
    return render_template('admin/tutors.html', tutors=tutors, departments=departments, filtered_args=filtered_args )


@bp.route('/tutors/register', methods=['GET', 'POST'])
@login_required
@require_permission('tutor_management')
def register_tutor():
    """Register new tutor with test score"""
    form = TutorRegistrationForm()
    
    if form.validate_on_submit():
        try:
            # Validate department selection
            if form.department_id.data == 0:
                raise ValueError("Please select a valid department")
            
            # Validate test score
            if form.test_score.data is None or form.test_score.data < 0 or form.test_score.data > 100:
                raise ValueError("Test score must be between 0 and 100")
            
            # Store password before hashing for email
            plain_password = form.password.data
            
            # Create user account
            user = User(
                username=form.username.data,
                email=form.email.data,
                full_name=form.full_name.data,
                phone=form.phone.data,
                role='tutor',
                department_id=form.department_id.data,
                address=form.address.data,
                is_active=True,
                is_verified=False,
                joining_date=datetime.now().date()
            )
            user.set_password(plain_password)
            
            db.session.add(user)
            db.session.flush()  # Get user ID
            
            # Create tutor profile with test score
            tutor = Tutor(
                user_id=user.id,
                qualification=form.qualification.data,
                experience=form.experience.data,
                salary_type=form.salary_type.data,
                monthly_salary=form.monthly_salary.data,
                hourly_rate=form.hourly_rate.data,
                # Test Score Information - NEW FIELDS
                test_score=form.test_score.data,
                test_date=form.test_date.data,
                test_notes=form.test_notes.data,
                status='pending',
                verification_status='pending',
                date_of_birth=form.date_of_birth.data,
                state=form.state.data,
                pin_code=form.pin_code.data
            )
            
            # Set subjects, grades, and boards
            tutor.set_subjects([s.strip() for s in form.subjects.data.split(',')])
            tutor.set_grades([g.strip() for g in form.grades.data.split(',')])
            tutor.set_boards([b.strip() for b in form.boards.data.split(',')])
            
            # Handle document uploads
            documents = {}
            required_docs = ['aadhaar_card', 'pan_card', 'resume', 'degree_certificate']
            
            for doc_name in required_docs:
                file_field = getattr(form, doc_name).data

                if has_file_content(file_field):
                    s3_url = upload_file_to_s3(file_field, folder=f"{current_app.config['UPLOAD_FOLDER']}/documents")
                    if s3_url:
                        documents[doc_name.replace('_card', '').replace('_certificate', '')] = s3_url
                    else:
                        raise ValueError(f"{doc_name.replace('_', ' ').title()} upload failed.")
                else:
                    raise ValueError(f"{doc_name.replace('_', ' ').title()} is required")
            
            tutor.set_documents(documents)
            
            # Handle video uploads
            required_videos = ['demo_video', 'interview_video']
            for video_name in required_videos:
                file_field = getattr(form, video_name).data
                if has_file_content(file_field):
                    s3_url = upload_file_to_s3(file_field, folder=f"{current_app.config['UPLOAD_FOLDER']}/videos")
                    if s3_url:
                        setattr(tutor, video_name, s3_url)
                    else:
                        raise ValueError(f"{video_name.replace('_', ' ').title()} upload failed.")
            
            # Set bank details
            bank_details = {
                'account_holder_name': form.account_holder_name.data,
                'bank_name': form.bank_name.data,
                'branch_name': form.branch_name.data,
                'account_number': form.account_number.data,
                'ifsc_code': form.ifsc_code.data
            }
            tutor.set_bank_details(bank_details)
            
            db.session.add(tutor)
            db.session.commit()
            
            # Send onboarding email
            try:
                send_onboarding_email(user, plain_password)
                
                # Success message with test score info
                test_grade = tutor.get_test_score_grade()
                success_message = f'Tutor {user.full_name} registered successfully with test score {tutor.test_score}/100 ({test_grade}). Onboarding email sent to {user.email}.'
                flash(success_message, 'success')
                
            except Exception as email_error:
                # Warning message with test score info
                test_grade = tutor.get_test_score_grade()
                warning_message = f'Tutor {user.full_name} registered successfully with test score {tutor.test_score}/100 ({test_grade}), but failed to send onboarding email. Please send credentials manually to {user.email}.'
                flash(warning_message, 'warning')
            
            return redirect(url_for('admin.tutors'))
            
        except ValueError as ve:
            db.session.rollback()
            flash(str(ve), 'error')
        except Exception as e:
            db.session.rollback()
            flash(f'Error registering tutor: {str(e)}', 'error')
    
    return render_template('admin/register_tutor.html', form=form)

def has_file_content(file_field):
    """Check if file field actually has a file"""
    return (file_field and 
            hasattr(file_field, 'filename') and 
            file_field.filename and 
            file_field.filename.strip() != '')

@bp.route('/tutors/<int:tutor_id>')
@login_required
@require_permission('tutor_management')
def tutor_details(tutor_id):
    """View tutor details"""
    tutor = Tutor.query.get_or_404(tutor_id)
    
    # Get tutor's classes
    classes = Class.query.filter_by(tutor_id=tutor.id).order_by(Class.scheduled_date.desc()).limit(10).all()
    
    # Get attendance summary
    attendance_summary = Attendance.get_attendance_summary(tutor_id=tutor.id)
    
    return render_template('admin/tutor_details.html', 
                         tutor=tutor, classes=classes, 
                         attendance_summary=attendance_summary)

@bp.route('/tutors/<int:tutor_id>/verify', methods=['POST'])
@login_required
@require_permission('tutor_management')
def verify_tutor(tutor_id):
    """Verify tutor profile"""
    tutor = Tutor.query.get_or_404(tutor_id)
    
    # Validate request data
    if not request.is_json:
        return jsonify({'success': False, 'error': 'Request must be JSON'}), 400
    
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'No data provided'}), 400
    
    action = data.get('action')
    if not action or action not in ['approve', 'reject']:
        return jsonify({'success': False, 'error': 'Invalid or missing action'}), 400
    
    # Check if tutor can be verified
    if tutor.verification_status == 'verified' and action == 'approve':
        return jsonify({'success': False, 'error': 'Tutor is already verified'}), 400
    
    try:
        if action == 'approve':
            tutor.verification_status = 'verified'
            tutor.status = 'active'
            tutor.user.is_verified = True
            message = f'Tutor {tutor.user.full_name} has been verified and activated.'
            
        elif action == 'reject':
            tutor.verification_status = 'rejected'
            tutor.status = 'inactive'
            tutor.user.is_verified = False
            message = f'Tutor {tutor.user.full_name} verification has been rejected.'
        
        # Commit changes to database
        db.session.commit()
        
        # Log the action for audit trail
        current_app.logger.info(
            f'Admin {current_user.username} (ID: {current_user.id}) '
            f'{action}d tutor verification for {tutor.user.full_name} (ID: {tutor_id})'
        )
        
        return jsonify({
            'success': True,
            'message': message,
            'data': {
                'tutor_id': tutor_id,
                'verification_status': tutor.verification_status,
                'tutor_status': tutor.status,
                'user_verified': tutor.user.is_verified
            }
        })
        
    except Exception as e:
        db.session.rollback()
        error_msg = f'Database error during tutor verification: {str(e)}'
        current_app.logger.error(error_msg)
        return jsonify({
            'success': False, 
            'error': 'Failed to update tutor verification status. Please try again.'
        }), 500

@bp.route('/tutors/<int:tutor_id>/toggle-status', methods=['POST'])
@login_required
@require_permission('tutor_management')
def toggle_tutor_status(tutor_id):
    """Toggle tutor active/inactive status"""
    tutor = Tutor.query.get_or_404(tutor_id)
    
    try:
        # Determine new status
        if tutor.status == 'active':
            new_tutor_status = 'inactive'
            new_user_status = False
            action_text = 'deactivated'
        else:
            new_tutor_status = 'active'
            new_user_status = True
            action_text = 'activated'
        
        # Update both tutor and user status
        tutor.status = new_tutor_status
        tutor.user.is_active = new_user_status
        
        # Commit changes
        db.session.commit()
        
        message = f'Tutor {tutor.user.full_name} has been {action_text}.'
        
        # Log the action
        current_app.logger.info(
            f'Admin {current_user.username} (ID: {current_user.id}) '
            f'{action_text} tutor {tutor.user.full_name} (ID: {tutor_id})'
        )
        
        return jsonify({
            'success': True,
            'message': message,
            'data': {
                'tutor_id': tutor_id,
                'tutor_status': tutor.status,
                'user_active': tutor.user.is_active,
                'action': action_text
            }
        })
        
    except Exception as e:
        db.session.rollback()
        error_msg = f'Database error during status toggle: {str(e)}'
        current_app.logger.error(error_msg)
        return jsonify({
            'success': False,
            'error': 'Failed to update tutor status. Please try again.'
        }), 500

# ============ STUDENT MANAGEMENT ROUTES ============

# Replace the students route in app/routes/admin.py

@bp.route('/students')
@login_required
@require_permission('student_management')
def students():
    """Student management page"""
    from sqlalchemy import or_
    
    page = request.args.get('page', 1, type=int)
    grade_filter = request.args.get('grade', '').strip()
    dept_filter_raw = request.args.get('department', '').strip()
    search = request.args.get('search', '').strip()
    
    # Convert department filter to int only if it's a valid number
    dept_filter = None
    if dept_filter_raw and dept_filter_raw.isdigit():
        dept_filter = int(dept_filter_raw)
    
    # Build base query
    query = Student.query
    
    # Apply department access check for coordinators FIRST
    if current_user.role == 'coordinator':
        query = query.filter_by(department_id=current_user.department_id)
    
    # Apply filters only if they have valid values
    if grade_filter:
        query = query.filter_by(grade=grade_filter)
    
    # Only apply department filter if user selected a specific department AND user is not a coordinator
    if dept_filter and dept_filter > 0 and current_user.role != 'coordinator':
        query = query.filter_by(department_id=dept_filter)
    
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Student.full_name.ilike(search_term),
                Student.email.ilike(search_term)
            )
        )
    
    # Get paginated results
    students = query.order_by(Student.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    # Calculate stats from the filtered query
    stats = {
        'total_students': query.count(),
        'active_students': query.filter_by(enrollment_status='active').count(),
        'paused_students': query.filter_by(enrollment_status='paused').count(),
        'completed_students': query.filter_by(enrollment_status='completed').count(),
        'dropped_students': query.filter_by(enrollment_status='dropped').count()
    }
    
    # Get departments (respect coordinator permissions)
    if current_user.role == 'coordinator':
        departments = Department.query.filter_by(
            id=current_user.department_id, 
            is_active=True
        ).all()
    else:
        departments = Department.query.filter_by(is_active=True).all()
    
    return render_template('admin/students.html', 
                         students=students, 
                         departments=departments,
                         stats=stats)

# Replace the register_student route in app/routes/admin.py

@bp.route('/students/register', methods=['GET', 'POST'])
@login_required
@require_permission('student_management')
def register_student():
    """Register new student"""
    from app.forms.user import StudentRegistrationForm
    
    form = StudentRegistrationForm()
    
    # Get all unique subjects for dropdown (for the subject enhancement)
    all_students = Student.query.all()
    all_subjects = set()
    for s in all_students:
        subjects = s.get_subjects_enrolled()
        all_subjects.update(subjects)
    
    # Add some common subjects if none exist yet
    if not all_subjects:
        all_subjects = {
            'Mathematics', 'Physics', 'Chemistry', 'Biology', 'English', 
            'Hindi', 'History', 'Geography', 'Economics', 'Computer Science',
            'Accountancy', 'Business Studies', 'Political Science', 'Sociology'
        }
    
    if form.validate_on_submit():
        try:
            # Create student
            student = Student(
                full_name=form.full_name.data,
                email=form.email.data,
                phone=form.phone.data,
                date_of_birth=form.date_of_birth.data,
                address=form.address.data,
                state=form.state.data,
                pin_code=form.pin_code.data,
                grade=form.grade.data,
                board=form.board.data,
                school_name=form.school_name.data,
                academic_year=form.academic_year.data,
                course_start_date=form.course_start_date.data,
                department_id=form.department_id.data,
                relationship_manager=form.relationship_manager.data,
                created_at=datetime.utcnow()
            )
            
            # Parent details
            parent_details = {
                'father': {
                    'name': form.father_name.data or '',
                    'phone': form.father_phone.data or '',
                    'email': form.father_email.data or '',
                    'profession': form.father_profession.data or '',
                    'workplace': getattr(form, 'father_workplace', None) and form.father_workplace.data or ''
                },
                'mother': {
                    'name': form.mother_name.data or '',
                    'phone': form.mother_phone.data or '',
                    'email': form.mother_email.data or '',
                    'profession': form.mother_profession.data or '',
                    'workplace': getattr(form, 'mother_workplace', None) and form.mother_workplace.data or ''
                }
            }
            student.set_parent_details(parent_details)
            
            # Academic profile
            academic_profile = {
                'siblings': getattr(form, 'siblings', None) and form.siblings.data or '',
                'hobbies': [h.strip() for h in (getattr(form, 'hobbies', None) and form.hobbies.data or '').split(',') if h.strip()],
                'learning_styles': [l.strip() for l in (getattr(form, 'learning_styles', None) and form.learning_styles.data or '').split(',') if l.strip()],
                'learning_patterns': [p.strip() for p in (getattr(form, 'learning_patterns', None) and form.learning_patterns.data or '').split(',') if p.strip()],
                'parent_feedback': getattr(form, 'parent_feedback', None) and form.parent_feedback.data or ''
            }
            student.set_academic_profile(academic_profile)
            
            # Subjects
            if form.subjects_enrolled.data:
                subjects = [s.strip() for s in form.subjects_enrolled.data.split(',') if s.strip()]
                student.set_subjects_enrolled(subjects)
            
            # Fee structure
            fee_structure = {
                'total_fee': float(getattr(form, 'total_fee', None) and form.total_fee.data or 0),
                'amount_paid': float(getattr(form, 'amount_paid', None) and form.amount_paid.data or 0),
                'payment_mode': getattr(form, 'payment_mode', None) and form.payment_mode.data or '',
                'payment_schedule': getattr(form, 'payment_schedule', None) and form.payment_schedule.data or ''
            }
            fee_structure['balance_amount'] = fee_structure['total_fee'] - fee_structure['amount_paid']
            student.set_fee_structure(fee_structure)
            
            # Handle document uploads
            documents = {}
            
            # Process file uploads if they exist
            for field_name in ['marksheet', 'student_aadhaar', 'school_id']:
                if hasattr(form, field_name):
                    file_field = getattr(form, field_name)
                    if file_field.data:
                        try:
                            # Upload to S3 or your file storage system
                            file_url = upload_file_to_s3(file_field.data, folder=f"{current_app.config['UPLOAD_FOLDER']}/students")
                            if file_url:
                                documents[field_name] = file_url
                        except Exception as e:
                            print(f"Error uploading {field_name}: {e}")
                            flash(f'Error uploading {field_name.replace("_", " ").title()}', 'warning')
            
            if documents:
                student.set_documents(documents)
            
            db.session.add(student)
            db.session.commit()
            
            flash(f'Student {student.full_name} registered successfully!', 'success')
            return redirect(url_for('admin.students'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error registering student: {str(e)}', 'error')
            print(f"Error registering student: {e}")  # For debugging
    else:
        # Print form errors for debugging
        if form.errors:
            print(f"Form validation errors: {form.errors}")
    
    return render_template('admin/register_student.html', 
                         form=form, 
                         all_subjects=sorted(list(all_subjects)))

@bp.route('/students/<int:student_id>')
@login_required
@require_permission('student_management')
def student_details(student_id):
    """View student details"""
    student = Student.query.get_or_404(student_id)
    
    # Check department access
    if current_user.role == 'coordinator' and current_user.department_id != student.department_id:
        flash('Access denied. You can only view students from your department.', 'error')
        return redirect(url_for('admin.students'))
    
    # Get student's classes
    student_classes = []
    all_classes = Class.query.order_by(Class.scheduled_date.desc()).all()
    
    for cls in all_classes:
        if student.id in cls.get_students():
            student_classes.append(cls)
    
    student_classes = student_classes[:20]  # Limit to recent 20 classes
    
    # Get attendance summary
    attendance_summary = Attendance.get_attendance_summary(student_id=student.id)
    
    # Get upcoming classes
    upcoming_classes = []
    for cls in Class.query.filter(
        Class.scheduled_date >= date.today(),
        Class.status == 'scheduled'
    ).order_by(Class.scheduled_date, Class.scheduled_time).all():
        if student.id in cls.get_students():
            upcoming_classes.append(cls)
    
    upcoming_classes = upcoming_classes[:5]  # Next 5 classes
    
    return render_template('admin/student_details.html',
                         student=student,
                         classes=student_classes,
                         upcoming_classes=upcoming_classes,
                         attendance_summary=attendance_summary)

# COMPLETE REPLACEMENT for edit_student route in app/routes/admin.py

@bp.route('/students/<int:student_id>/edit', methods=['GET', 'POST'])
@login_required
@require_permission('student_management')
def edit_student(student_id):
    """Edit student information"""
    from datetime import datetime
    from app.forms.user import EditStudentForm
    
    student = Student.query.get_or_404(student_id)
    
    # Check department access
    if current_user.role == 'coordinator' and current_user.department_id != student.department_id:
        flash('Access denied. You can only edit students from your department.', 'error')
        return redirect(url_for('admin.students'))
    
    # Initialize form with student object (this handles basic fields)
    form = EditStudentForm(student_id=student.id, obj=student)
    
    # Pre-populate form with existing data on GET request
    if request.method == 'GET':
        # Get parent details and populate form
        parent_details = student.get_parent_details()
        if parent_details:
            father = parent_details.get('father', {})
            mother = parent_details.get('mother', {})
            
            # Populate father fields
            form.father_name.data = father.get('name', '')
            form.father_phone.data = father.get('phone', '')
            form.father_email.data = father.get('email', '')
            form.father_profession.data = father.get('profession', '')
            
            # Populate mother fields  
            form.mother_name.data = mother.get('name', '')
            form.mother_phone.data = mother.get('phone', '')
            form.mother_email.data = mother.get('email', '')
            form.mother_profession.data = mother.get('profession', '')
        
        # Get academic profile and populate form
        academic_profile = student.get_academic_profile()
        if academic_profile:
            form.siblings.data = str(academic_profile.get('siblings', ''))
            form.hobbies.data = ', '.join(academic_profile.get('hobbies', []))
            form.learning_styles.data = ', '.join(academic_profile.get('learning_styles', []))
            form.parent_feedback.data = academic_profile.get('parent_feedback', '')
        
        # Get subjects and populate form
        subjects = student.get_subjects_enrolled()
        if subjects:
            form.subjects_enrolled.data = ', '.join(subjects)
        
        # Get fee structure and populate form
        fee_structure = student.get_fee_structure()
        if fee_structure:
            form.total_fee.data = fee_structure.get('total_fee', 0)
            form.amount_paid.data = fee_structure.get('amount_paid', 0)
            form.payment_mode.data = fee_structure.get('payment_mode', '')
            form.payment_schedule.data = fee_structure.get('payment_schedule', '')
    
    if form.validate_on_submit():
        try:
            # Update basic information
            student.full_name = form.full_name.data
            student.email = form.email.data
            student.phone = form.phone.data
            student.date_of_birth = form.date_of_birth.data
            student.address = form.address.data
            student.state = form.state.data
            student.pin_code = form.pin_code.data
            student.grade = form.grade.data
            student.board = form.board.data
            student.school_name = form.school_name.data
            student.academic_year = form.academic_year.data
            student.course_start_date = form.course_start_date.data
            student.relationship_manager = form.relationship_manager.data
            
            # Update parent details
            parent_details = {
                'father': {
                    'name': form.father_name.data or '',
                    'phone': form.father_phone.data or '',
                    'email': form.father_email.data or '',
                    'profession': form.father_profession.data or ''
                },
                'mother': {
                    'name': form.mother_name.data or '',
                    'phone': form.mother_phone.data or '',
                    'email': form.mother_email.data or '',
                    'profession': form.mother_profession.data or ''
                }
            }
            student.set_parent_details(parent_details)
            
            # Update academic profile
            academic_profile = {
                'siblings': form.siblings.data or '',
                'hobbies': [h.strip() for h in (form.hobbies.data or '').split(',') if h.strip()],
                'learning_styles': [l.strip() for l in (form.learning_styles.data or '').split(',') if l.strip()],
                'parent_feedback': form.parent_feedback.data or ''
            }
            student.set_academic_profile(academic_profile)
            
            # Update subjects enrolled
            if form.subjects_enrolled.data:
                subjects = [s.strip() for s in form.subjects_enrolled.data.split(',') if s.strip()]
                student.set_subjects_enrolled(subjects)
            
            # Update fee structure
            fee_structure = student.get_fee_structure()
            fee_structure.update({
                'total_fee': float(form.total_fee.data or 0),
                'amount_paid': float(form.amount_paid.data or 0),
                'payment_mode': form.payment_mode.data or '',
                'payment_schedule': form.payment_schedule.data or ''
            })
            fee_structure['balance_amount'] = fee_structure['total_fee'] - fee_structure['amount_paid']
            student.set_fee_structure(fee_structure)
            
            # Set department if provided
            if form.department_id.data:
                student.department_id = form.department_id.data
            
            # Set updated timestamp
            student.updated_at = datetime.utcnow()
            
            db.session.commit()
            flash(f'Student {student.full_name} updated successfully!', 'success')
            return redirect(url_for('admin.students'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating student: {str(e)}', 'error')
            print(f"Error updating student: {e}")
    else:
        # Print form errors for debugging
        if form.errors:
            print(f"Form validation errors: {form.errors}")
    
    # Get all unique subjects for dropdown (for the subject enhancement)
    all_students = Student.query.all()
    all_subjects = set()
    for s in all_students:
        subjects = s.get_subjects_enrolled()
        all_subjects.update(subjects)
    
    return render_template('admin/edit_student.html', 
                         form=form, 
                         student=student,
                         all_subjects=sorted(list(all_subjects)))

@bp.route('/students/<int:student_id>/deactivate', methods=['POST'])
@login_required
@require_permission('student_management')
def deactivate_student(student_id):
    """Deactivate a student"""
    student = Student.query.get_or_404(student_id)
    
    # Check department access
    if current_user.role == 'coordinator' and current_user.department_id != student.department_id:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        student.is_active = False
        student.enrollment_status = 'dropped'
        db.session.commit()
        
        flash(f'Student {student.full_name} has been deactivated.', 'success')
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Error deactivating student'}), 500

@bp.route('/students/<int:student_id>/activate', methods=['POST'])
@login_required
@require_permission('student_management')
def activate_student(student_id):
    """Activate a student"""
    student = Student.query.get_or_404(student_id)
    
    # Check department access
    if current_user.role == 'coordinator' and current_user.department_id != student.department_id:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        student.is_active = True
        student.enrollment_status = 'active'
        db.session.commit()
        
        flash(f'Student {student.full_name} has been activated.', 'success')
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Error activating student'}), 500


@bp.route('/students/<int:student_id>/toggle-status', methods=['POST'])
@login_required
@require_permission('student_management')
def toggle_student_status(student_id):
    """Toggle student active status"""
    student = Student.query.get_or_404(student_id)
    
    # Check department access
    if current_user.role == 'coordinator' and current_user.department_id != student.department_id:
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        # Toggle status
        student.is_active = not student.is_active
        
        # Update enrollment status based on active status
        if student.is_active:
            student.enrollment_status = 'active'
        else:
            student.enrollment_status = 'paused'
        
        # Set updated timestamp
        student.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        status_text = 'activated' if student.is_active else 'deactivated'
        return jsonify({
            'success': True, 
            'message': f'Student {student.full_name} {status_text} successfully',
            'new_status': student.is_active
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error updating student status: {str(e)}'}), 500

@bp.route('/students/<int:student_id>/delete', methods=['DELETE'])
@login_required
@require_permission('student_management')
def delete_student(student_id):
    """Delete student (superadmin only)"""
    if current_user.role != 'superadmin':
        return jsonify({'error': 'Only superadmin can delete students'}), 403
    
    student = Student.query.get_or_404(student_id)
    
    try:
        student_name = student.full_name
        
        # You might want to soft delete instead of hard delete
        # For now, let's mark as deleted instead of actual deletion
        student.is_active = False
        student.enrollment_status = 'deleted'
        student.updated_at = datetime.utcnow()
        
        # Or for hard delete (uncomment below and comment above):
        # db.session.delete(student)
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'Student {student_name} deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error deleting student: {str(e)}'}), 500

# ============ CLASS MANAGEMENT ROUTES ============

@bp.route('/classes')
@login_required
@require_permission('class_management')
def classes():
    """Class management page"""
    page = request.args.get('page', 1, type=int)
    date_filter = request.args.get('date', '')
    tutor_filter = request.args.get('tutor', '', type=int)
    status_filter = request.args.get('status', '')
    class_type_filter = request.args.get('class_type', '')
    
    query = Class.query
    
    # Apply filters
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            query = query.filter_by(scheduled_date=filter_date)
        except ValueError:
            pass
    
    if tutor_filter:
        query = query.filter_by(tutor_id=tutor_filter)
    
    if status_filter:
        query = query.filter_by(status=status_filter)
        
    if class_type_filter:
        query = query.filter_by(class_type=class_type_filter)
    
    classes = query.order_by(Class.scheduled_date, Class.scheduled_time).paginate(
        page=page, per_page=20, error_out=False
    )
    
    # Get tutors - only those with availability set AND with user relationship
    available_tutors = []
    all_tutors = []
    
    try:
        # Get all tutors with user relationship
        tutors_query = Tutor.query.join(User).filter(User.is_active==True).all()
        
        for tutor in tutors_query:
            all_tutors.append(tutor)
            availability = tutor.get_availability()
            if availability and tutor.status == 'active':
                available_tutors.append(tutor)
    except Exception as e:
        print(f"Error loading tutors: {e}")
        tutors_query = Tutor.query.all()
        for tutor in tutors_query:
            if hasattr(tutor, 'user') and tutor.user:
                all_tutors.append(tutor)
                availability = tutor.get_availability()
                if availability and tutor.status == 'active':
                    available_tutors.append(tutor)
    
    # Get students - Student model has fields directly
    students = []
    try:
        students = Student.query.filter(Student.is_active==True).all()
    except Exception as e:
        print(f"Error loading students: {e}")
        students = []
    
    # Create students dictionary for quick lookup
    students_dict = {s.id: s for s in students}
    
    # Check if there are tutors without availability
    tutors_without_availability = [t for t in all_tutors if not t.get_availability()]
    
    # Convert to dictionaries for JavaScript
    tutors_dict_data = []
    for tutor in available_tutors:
        try:
            tutor_dict = {
                'id': tutor.id,
                'user_name': tutor.user.full_name if tutor.user else 'Unknown',
                'email': tutor.user.email if tutor.user else '',
                'subjects': tutor.get_subjects(),
                'grades': tutor.get_grades(),
                'boards': tutor.get_boards(),
                'has_availability': bool(tutor.get_availability()),
                'status': tutor.status,
                'rating': tutor.rating or 0
            }
            tutors_dict_data.append(tutor_dict)
        except Exception as e:
            print(f"Error processing tutor {tutor.id}: {e}")
    
    students_dict_data = []
    for student in students:
        try:
            student_dict = {
                'id': student.id,
                'full_name': student.full_name,
                'grade': student.grade,
                'board': student.board,
                'subjects_enrolled': student.get_subjects_enrolled() if hasattr(student, 'get_subjects_enrolled') else []
            }
            students_dict_data.append(student_dict)
        except Exception as e:
            print(f"Error processing student {student.id}: {e}")
    
    return render_template('admin/classes.html', 
                         classes=classes, 
                         tutors=available_tutors,  # Only tutors with availability
                         all_tutors=all_tutors,    # All tutors for reference
                         tutors_without_availability=tutors_without_availability,
                         students=students,
                         students_dict=students_dict,
                         tutors_dict_data=tutors_dict_data,  # For JavaScript
                         students_dict_data=students_dict_data,  # For JavaScript
                         today=date.today(),
                         csrf_token=generate_csrf)


@bp.route('/classes/create', methods=['POST'])
@login_required
@require_permission('class_management')
def create_class():
    """Create a new class with availability validation"""
    try:
        # Get tutor and validate availability
        tutor_id = int(request.form['tutor_id'])
        tutor = Tutor.query.get_or_404(tutor_id)
        
        # Check if tutor has set availability
        availability = tutor.get_availability()
        if not availability:
            flash(f'Cannot create class: {tutor.user.full_name} has not set their availability yet. Please ask the tutor to set their schedule first.', 'error')
            return redirect(url_for('admin.classes'))
        
        # Check if tutor is active
        if tutor.status != 'active':
            flash(f'Cannot create class: {tutor.user.full_name} is not in active status.', 'error')
            return redirect(url_for('admin.classes'))
        
        # Get class details
        scheduled_date = datetime.strptime(request.form['scheduled_date'], '%Y-%m-%d').date()
        scheduled_time = datetime.strptime(request.form['scheduled_time'], '%H:%M').time()
        
        # Check if tutor is available at the requested time
        day_of_week = scheduled_date.strftime('%A').lower()  # monday, tuesday, etc.
        time_str = scheduled_time.strftime('%H:%M')
        
        if not tutor.is_available_at(day_of_week, time_str):
            flash(f'Cannot create class: {tutor.user.full_name} is not available on {day_of_week.title()} at {time_str}. Please choose a different time or tutor.', 'error')
            return redirect(url_for('admin.classes'))
        
        # Check for scheduling conflicts (tutor already has a class at this time)
        existing_class = Class.query.filter_by(
            tutor_id=tutor_id,
            scheduled_date=scheduled_date,
            scheduled_time=scheduled_time,
            status='scheduled'
        ).first()
        
        if existing_class:
            flash(f'Cannot create class: {tutor.user.full_name} already has a class scheduled at this time.', 'error')
            return redirect(url_for('admin.classes'))
        
        # Create class if all validations pass
        class_data = {
            'subject': request.form['subject'],
            'class_type': request.form['class_type'],
            'scheduled_date': scheduled_date,
            'scheduled_time': scheduled_time,
            'duration': int(request.form['duration']),
            'tutor_id': tutor_id,
            'grade': request.form.get('grade', ''),
            'meeting_link': request.form.get('meeting_link', ''),
            'class_notes': request.form.get('class_notes', ''),
            'status': 'scheduled',
            'created_by': current_user.id
        }
        
        new_class = Class(**class_data)
        db.session.add(new_class)
        db.session.flush()  # Get class ID
        
        # Handle student assignment based on class type
        students = []
        if class_data['class_type'] == 'one_on_one' or class_data['class_type'] == 'demo':
            primary_student_id = request.form.get('primary_student_id')
            if primary_student_id:
                students = [int(primary_student_id)]
        elif class_data['class_type'] == 'group':
            students = [int(s) for s in request.form.getlist('students')]
        
        if students:
            new_class.set_students(students)
        
        db.session.commit()
        flash(f'Class created successfully for {class_data["scheduled_date"]} at {time_str}!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating class: {str(e)}', 'error')
    
    return redirect(url_for('admin.classes'))

# Replace the bulk_create_classes route in app/routes/admin.py

@bp.route('/classes/bulk-create', methods=['POST'])
@login_required
@require_permission('class_management')
def bulk_create_classes():
    """Create multiple classes in bulk with availability validation"""
    try:
        from datetime import timedelta
        
        # Get form data with validation
        subject = request.form.get('subject', '').strip()
        grade = request.form.get('grade', '').strip()
        duration = request.form.get('duration', '')
        tutor_id = request.form.get('tutor_id', '')
        class_type = request.form.get('class_type', '')
        
        # Validate required fields
        if not all([subject, grade, duration, tutor_id, class_type]):
            flash('All basic fields (subject, grade, duration, tutor, class type) are required.', 'error')
            return redirect(url_for('admin.classes'))
        
        # Convert to proper types
        try:
            duration = int(duration)
            tutor_id = int(tutor_id)
        except ValueError:
            flash('Invalid duration or tutor selection.', 'error')
            return redirect(url_for('admin.classes'))
        
        # Get students list
        students = request.form.getlist('students')
        if not students:
            flash('At least one student must be selected.', 'error')
            return redirect(url_for('admin.classes'))
        
        try:
            students = [int(s) for s in students]
        except ValueError:
            flash('Invalid student selection.', 'error')
            return redirect(url_for('admin.classes'))
        
        # Get schedule data
        start_date_str = request.form.get('start_date', '')
        end_date_str = request.form.get('end_date', '')
        start_time_str = request.form.get('start_time', '')
        days_of_week = request.form.getlist('days_of_week')
        
        if not all([start_date_str, end_date_str, start_time_str, days_of_week]):
            flash('All schedule fields (start date, end date, start time, days of week) are required.', 'error')
            return redirect(url_for('admin.classes'))
        
        # Parse dates and time
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            start_time = datetime.strptime(start_time_str, '%H:%M').time()
            days_of_week = [int(d) for d in days_of_week]
        except ValueError as e:
            flash('Invalid date or time format.', 'error')
            return redirect(url_for('admin.classes'))
        
        # Validate date range
        if end_date <= start_date:
            flash('End date must be after start date.', 'error')
            return redirect(url_for('admin.classes'))
        
        # Validate tutor
        tutor = Tutor.query.get_or_404(tutor_id)
        availability = tutor.get_availability()
        
        if not availability:
            flash(f'Cannot create classes: {tutor.user.full_name} has not set their availability yet.', 'error')
            return redirect(url_for('admin.classes'))
        
        if tutor.status != 'active':
            flash(f'Cannot create classes: {tutor.user.full_name} is not in active status.', 'error')
            return redirect(url_for('admin.classes'))
        
        # Validate students exist
        student_objects = Student.query.filter(Student.id.in_(students)).all()
        if len(student_objects) != len(students):
            flash('One or more selected students do not exist.', 'error')
            return redirect(url_for('admin.classes'))
        
        # Create classes
        created_count = 0
        skipped_count = 0
        current_date = start_date
        
        while current_date <= end_date:
            # Check if current day is in selected days of week
            if current_date.weekday() in days_of_week:
                day_of_week = current_date.strftime('%A').lower()
                time_str = start_time.strftime('%H:%M')
                
                # Check tutor availability for this day/time
                if tutor.is_available_at(day_of_week, time_str):
                    # Check for existing classes at this time
                    existing_class = Class.query.filter_by(
                        tutor_id=tutor_id,
                        scheduled_date=current_date,
                        scheduled_time=start_time,
                        status='scheduled'
                    ).first()
                    
                    if not existing_class:
                        # Create the class
                        class_data = {
                            'subject': subject,
                            'class_type': class_type,
                            'scheduled_date': current_date,
                            'scheduled_time': start_time,
                            'duration': duration,
                            'tutor_id': tutor_id,
                            'grade': grade,
                            'meeting_link': request.form.get('meeting_link', ''),
                            'class_notes': request.form.get('class_notes', ''),
                            'status': 'scheduled',
                            'created_by': current_user.id
                        }
                        
                        new_class = Class(**class_data)
                        db.session.add(new_class)
                        db.session.flush()  # Get the ID
                        
                        # Set students for the class
                        if hasattr(new_class, 'set_students'):
                            new_class.set_students(students)
                        else:
                            # Fallback if set_students method doesn't exist
                            if students:
                                new_class.primary_student_id = students[0]
                        
                        created_count += 1
                    else:
                        skipped_count += 1
                else:
                    skipped_count += 1
            
            current_date += timedelta(days=1)
        
        db.session.commit()
        
        # Show results
        if created_count > 0:
            message = f'{created_count} classes created successfully!'
            if skipped_count > 0:
                message += f' {skipped_count} classes were skipped due to availability conflicts or existing bookings.'
            flash(message, 'success')
        else:
            flash('No classes were created. Please check tutor availability and existing schedules.', 'warning')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating bulk classes: {str(e)}', 'error')
        print(f"Bulk create error: {e}")  # For debugging
    
    return redirect(url_for('admin.classes'))

@bp.route('/api/v1/tutor/<int:tutor_id>/availability')
@login_required
@require_permission('class_management')
def api_tutor_availability(tutor_id):
    """Get tutor availability for AJAX requests"""
    try:
        tutor = Tutor.query.get_or_404(tutor_id)
        availability = tutor.get_availability()
        
        return jsonify({
            'success': True,
            'has_availability': bool(availability),
            'availability': availability,
            'status': tutor.status,
            'can_teach': tutor.status == 'active' and bool(availability),
            'tutor_name': tutor.user.full_name if tutor.user else 'Unknown',
            'subjects': tutor.get_subjects(),
            'grades': tutor.get_grades(),
            'boards': tutor.get_boards()
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/tutors/<int:tutor_id>/edit', methods=['GET', 'POST'])
@login_required
@require_permission('tutor_management')
def edit_tutor(tutor_id):
    """Edit tutor profile (admin route)"""
    tutor = Tutor.query.get_or_404(tutor_id)
    
    # Check department access for coordinators
    if current_user.role == 'coordinator':
        if current_user.department_id != tutor.user.department_id:
            flash('Access denied. You can only edit tutors from your department.', 'error')
            return redirect(url_for('admin.tutors'))
    
    # Redirect to the profile edit route with tutor_user_id
    return redirect(url_for('profile.edit_tutor_profile', tutor_user_id=tutor.user_id))

# ADD these routes to your app/routes/admin.py file

@bp.route('/tutors/<int:tutor_id>/documents/upload', methods=['POST'])
@login_required
@require_permission('tutor_management')
def upload_tutor_document(tutor_id):
    """Upload document for a tutor (admin only)"""
    try:
        tutor = Tutor.query.get_or_404(tutor_id)
        
        # Check permissions
        if current_user.role not in ['superadmin', 'admin']:
            return jsonify({'error': 'Access denied'}), 403
        
        document_type = request.form.get('document_type')
        file = request.files.get('document')
        
        if not file or not document_type:
            return jsonify({'error': 'File and document type are required'}), 400
        
        # Validate file type
        allowed_extensions = {'pdf', 'jpg', 'jpeg', 'png'}
        if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            return jsonify({'error': 'Invalid file type. Only PDF, JPG, and PNG files are allowed'}), 400
        
        # Validate file size (10MB limit)
        if request.content_length > 10 * 1024 * 1024:
            return jsonify({'error': 'File size exceeds 10MB limit'}), 400
        
        # Upload to S3 (using your existing upload_file_to_s3 function)
        s3_url = upload_file_to_s3(file, folder=f"{current_app.config['UPLOAD_FOLDER']}/documents")
        if not s3_url:
            return jsonify({'error': 'Failed to upload file to S3'}), 500
        
        # Update tutor documents
        documents = tutor.get_documents()
        documents[document_type] = {
            'filename': s3_url,  # Store S3 URL directly
            'uploaded_at': datetime.now().isoformat(),
            'uploaded_by': current_user.full_name
        }
        tutor.set_documents(documents)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{document_type.replace("_", " ").title()} uploaded successfully',
            'filename': s3_url
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error uploading tutor document: {str(e)}')
        return jsonify({'error': f'Error uploading document: {str(e)}'}), 500


@bp.route('/tutors/<int:tutor_id>/videos/upload', methods=['POST'])
@login_required
@require_permission('tutor_management')
def upload_tutor_video(tutor_id):
    """Upload video for a tutor (admin only)"""
    try:
        tutor = Tutor.query.get_or_404(tutor_id)
        
        # Check permissions
        if current_user.role not in ['superadmin', 'admin']:
            return jsonify({'error': 'Access denied'}), 403
        
        video_type = request.form.get('video_type')
        file = request.files.get('video')
        
        if not file or not video_type:
            return jsonify({'error': 'File and video type are required'}), 400
        
        # Validate file type
        allowed_extensions = {'mp4', 'avi', 'mov', 'wmv'}
        if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            return jsonify({'error': 'Invalid file type. Only MP4, AVI, MOV, and WMV files are allowed'}), 400
        
        # Validate file size (100MB limit)
        if request.content_length > 100 * 1024 * 1024:
            return jsonify({'error': 'File size exceeds 100MB limit'}), 400
        
        # Upload to S3 (using your existing upload_file_to_s3 function)
        s3_url = upload_file_to_s3(file, folder=f"{current_app.config['UPLOAD_FOLDER']}/videos")
        if not s3_url:
            return jsonify({'error': 'Failed to upload video to S3'}), 500
        
        # Update tutor video
        if video_type == 'demo':
            tutor.demo_video = s3_url
        elif video_type == 'interview':
            tutor.interview_video = s3_url
        else:
            return jsonify({'error': 'Invalid video type'}), 400
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{video_type.title()} video uploaded successfully',
            'filename': s3_url
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error uploading tutor video: {str(e)}')
        return jsonify({'error': f'Error uploading video: {str(e)}'}), 500


@bp.route('/tutors/<int:tutor_id>/documents/<document_type>/delete', methods=['POST'])
@login_required
@require_permission('tutor_management')
def delete_tutor_document(tutor_id, document_type):
    """Delete document for a tutor (admin only)"""
    try:
        tutor = Tutor.query.get_or_404(tutor_id)
        
        # Check permissions
        if current_user.role not in ['superadmin', 'admin']:
            return jsonify({'error': 'Access denied'}), 403
        
        documents = tutor.get_documents()
        s3_url = None
        
        if document_type in documents:
            if isinstance(documents[document_type], dict):
                s3_url = documents[document_type].get('filename')
            else:
                s3_url = documents[document_type]
            
            # Remove from documents
            del documents[document_type]
            tutor.set_documents(documents)
        
        # Note: For S3 files, you might want to implement S3 deletion
        # This would require additional S3 delete functionality
        # For now, we just remove the reference from the database
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{document_type.replace("_", " ").title()} deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error deleting tutor document: {str(e)}')
        return jsonify({'error': f'Error deleting document: {str(e)}'}), 500

@bp.route('/tutors/<int:tutor_id>/videos/<video_type>/delete', methods=['POST'])
@login_required
@require_permission('tutor_management')
def delete_tutor_video(tutor_id, video_type):
    """Delete video for a tutor (admin only)"""
    try:
        tutor = Tutor.query.get_or_404(tutor_id)
        
        # Check permissions
        if current_user.role not in ['superadmin', 'admin']:
            return jsonify({'error': 'Access denied'}), 403
        
        s3_url = None
        
        # Get current video URL and remove
        if video_type == 'demo' and tutor.demo_video:
            s3_url = tutor.demo_video
            tutor.demo_video = None
        elif video_type == 'interview' and tutor.interview_video:
            s3_url = tutor.interview_video
            tutor.interview_video = None
        else:
            return jsonify({'error': 'Video not found or invalid type'}), 404
        
        # Note: For S3 files, you might want to implement S3 deletion
        # This would require additional S3 delete functionality
        # For now, we just remove the reference from the database
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{video_type.title()} video deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error deleting tutor video: {str(e)}')
        return jsonify({'error': f'Error deleting video: {str(e)}'}), 500


# Helper function to validate file content (add this if it doesn't exist)
def has_file_content(file_field):
    """Check if file field has actual content"""
    return file_field and hasattr(file_field, 'filename') and file_field.filename != ''


@bp.route('/classes/<int:class_id>')
@login_required
@require_permission('class_management')
def class_details(class_id):
    """View class details"""
    class_obj = Class.query.get_or_404(class_id)
    
    # Get students for this class
    try:
        student_ids = class_obj.get_students() if hasattr(class_obj, 'get_students') else []
        students = Student.query.filter(Student.id.in_(student_ids)).all() if student_ids else []
    except:
        students = []
    
    return render_template('admin/class_details.html', class_item=class_obj, students=students)

    
@bp.route('/api/v1/compatible-tutors')
@login_required
@admin_required
def api_compatible_tutors():
    """Get tutors compatible with student criteria"""
    try:
        student_id = request.args.get('student_id', type=int)
        subject = request.args.get('subject', '').lower()
        grade = request.args.get('grade', '')
        board = request.args.get('board', '')
        
        # Get all active tutors with availability
        tutors = Tutor.query.filter_by(status='active').all()
        compatible_tutors = []
        
        for tutor in tutors:
            # Must have availability
            if not tutor.get_availability():
                continue
            
            # Check subject compatibility
            if subject:
                tutor_subjects = [s.lower() for s in tutor.get_subjects()]
                if not any(subject in ts or ts in subject for ts in tutor_subjects):
                    continue
            
            # Check grade compatibility
            if grade:
                tutor_grades = [str(g) for g in tutor.get_grades()]
                if tutor_grades and str(grade) not in tutor_grades:
                    continue
            
            # Check board compatibility
            if board:
                tutor_boards = [b.lower() for b in tutor.get_boards()]
                if tutor_boards and board.lower() not in tutor_boards:
                    continue
            
            compatible_tutors.append({
                'id': tutor.id,
                'name': tutor.user.full_name if tutor.user else 'Unknown',
                'email': tutor.user.email if tutor.user else '',
                'subjects': tutor.get_subjects(),
                'grades': tutor.get_grades(),
                'boards': tutor.get_boards(),
                'rating': tutor.rating or 0,
                'completion_rate': tutor.get_completion_rate()
            })
        
        return jsonify({
            'success': True,
            'tutors': compatible_tutors,
            'total': len(compatible_tutors)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
@bp.route('/api/v1/student/<int:student_id>/details')
@login_required
@require_permission('student_management')
def api_student_details(student_id):
    """Get student details for tutor matching"""
    try:
        student = Student.query.get_or_404(student_id)
        
        return jsonify({
            'success': True,
            'student': {
                'id': student.id,
                'full_name': student.full_name,
                'grade': student.grade,
                'board': student.board,
                'subjects_enrolled': student.get_subjects_enrolled(),
                'favorite_subjects': student.get_favorite_subjects() if hasattr(student, 'get_favorite_subjects') else [],
                'difficult_subjects': student.get_difficult_subjects() if hasattr(student, 'get_difficult_subjects') else [],
                'learning_style': student.get_academic_profile().get('learning_styles', []) if hasattr(student, 'get_academic_profile') else []
            }
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============ TIMETABLE MANAGEMENT ============

@bp.route('/timetable')
@login_required
@require_permission('class_management')
def timetable():
    """Timetable management page"""
    try:
        # Get departments, tutors, and students for the dropdowns
        departments = Department.query.filter_by(is_active=True).all()
        
        # Get tutors with their user relationships
        tutors = Tutor.query.join(User).filter(
            Tutor.status == 'active',
            User.is_active == True
        ).all()
        
        # Get active students (limit to prevent slow loading)
        students = Student.query.filter(
            Student.is_active == True
        ).limit(100).all()
        
        return render_template('admin/timetable.html', 
                             departments=departments,
                             tutors=tutors,
                             students=students)
                             
    except Exception as e:
        print(f"Error loading timetable page: {str(e)}")
        flash('Error loading timetable page', 'error')
        return redirect(url_for('dashboard.index'))

# ============ API ROUTES FOR DASHBOARD/TIMETABLE ============

@bp.route('/api/v1/timetable/today')
@login_required
@admin_required
def api_timetable_today():
    """Get today's timetable data with WORKING FILTERS"""
    try:
        # Get date parameter
        date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        target_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        
        # Get search filters - FIXED VERSION
        search = request.args.get('search', '').strip()
        tutor_id = request.args.get('tutor_id', type=int)
        student_id = request.args.get('student_id', type=int)
        department_id = request.args.get('department_id', type=int)
        
        # Base query for the specific date
        query = Class.query.filter(Class.scheduled_date == target_date)
        
        # Apply tutor filter - WORKING VERSION
        if tutor_id:
            query = query.filter(Class.tutor_id == tutor_id)
        
        # Apply student filter - FIXED VERSION
        if student_id:
            query = query.filter(
                db.or_(
                    Class.primary_student_id == student_id,
                    Class.demo_student_id == student_id,
                    Class.students.like(f'%{student_id}%')
                )
            )
        
        # Apply department filter - WORKING VERSION
        if department_id:
            query = query.join(Tutor, Class.tutor_id == Tutor.id)\
                         .join(User, Tutor.user_id == User.id)\
                         .filter(User.department_id == department_id)
        
        # Apply search filter - FIXED VERSION
        if search:
            query = query.join(Tutor, Class.tutor_id == Tutor.id, isouter=True)\
                         .join(User, Tutor.user_id == User.id, isouter=True)\
                         .filter(
                db.or_(
                    Class.subject.ilike(f'%{search}%'),
                    User.full_name.ilike(f'%{search}%'),
                    Class.grade.ilike(f'%{search}%'),
                    Class.board.ilike(f'%{search}%')
                )
            )
        
        # Get all classes for the day
        classes = query.order_by(Class.scheduled_time).all()
        
        # Build response data
        classes_data = []
        for cls in classes:
            try:
                # Get tutor name safely
                tutor_name = 'No Tutor Assigned'
                tutor_email = ''
                department_name = ''
                if cls.tutor and cls.tutor.user:
                    tutor_name = cls.tutor.user.full_name
                    tutor_email = cls.tutor.user.email or ''
                    if cls.tutor.user.department:
                        department_name = cls.tutor.user.department.name
                
                # Get student details - ENHANCED VERSION
                student_count = 0
                student_names = []
                student_emails = []
                
                if cls.class_type == 'demo' and cls.demo_student_id:
                    from app.models.demo_student import DemoStudent
                    demo_student = DemoStudent.query.get(cls.demo_student_id)
                    if demo_student:
                        student_count = 1
                        student_names = [demo_student.full_name]
                        student_emails = [demo_student.email or '']
                elif cls.primary_student_id:
                    student = Student.query.get(cls.primary_student_id)
                    if student:
                        student_count = 1
                        student_names = [student.full_name]
                        student_emails = [student.email or '']
                elif cls.students:
                    try:
                        import json
                        student_ids = json.loads(cls.students)
                        students = Student.query.filter(Student.id.in_(student_ids)).all()
                        student_count = len(students)
                        student_names = [s.full_name for s in students]
                        student_emails = [s.email or '' for s in students]
                    except (json.JSONDecodeError, TypeError):
                        student_count = 0
                        student_names = []
                        student_emails = []
                
                class_data = {
                    'id': cls.id,
                    'subject': cls.subject or 'No Subject',
                    'class_type': cls.class_type or 'regular',
                    'scheduled_date': cls.scheduled_date.strftime('%Y-%m-%d'),
                    'scheduled_time': cls.scheduled_time.strftime('%H:%M') if cls.scheduled_time else '00:00',
                    'duration': cls.duration or 60,
                    'status': cls.status or 'scheduled',
                    'tutor_name': tutor_name,
                    'tutor_email': tutor_email,
                    'tutor_id': cls.tutor_id,
                    'department_name': department_name,
                    'student_count': student_count,
                    'student_names': student_names,
                    'student_emails': student_emails,
                    'grade': cls.grade or '',
                    'board': cls.board or '',
                    'meeting_link': cls.meeting_link or '',
                    'platform': cls.platform or '',
                    'class_notes': cls.class_notes or '',
                    'can_reschedule': cls.can_be_rescheduled() if hasattr(cls, 'can_be_rescheduled') else True,
                    'can_cancel': cls.can_be_cancelled() if hasattr(cls, 'can_be_cancelled') else True
                }
                classes_data.append(class_data)
                
            except Exception as e:
                print(f"Error processing class {cls.id}: {str(e)}")
                continue
        
        # Calculate statistics
        total_classes = len(classes_data)
        scheduled_count = len([c for c in classes_data if c['status'] == 'scheduled'])
        ongoing_count = len([c for c in classes_data if c['status'] == 'ongoing'])
        completed_count = len([c for c in classes_data if c['status'] == 'completed'])
        cancelled_count = len([c for c in classes_data if c['status'] == 'cancelled'])
        
        stats = {
            'total_classes': total_classes,
            'today': {
                'scheduled': scheduled_count,
                'ongoing': ongoing_count,
                'completed': completed_count,
                'cancelled': cancelled_count
            }
        }
        
        return jsonify({
            'success': True,
            'classes': classes_data,
            'stats': stats,
            'date': target_date.strftime('%Y-%m-%d'),
            'filters_applied': {
                'search': search,
                'tutor_id': tutor_id,
                'student_id': student_id,
                'department_id': department_id
            }
        })
        
    except Exception as e:
        print(f"Error in api_timetable_today: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'classes': [],
            'stats': {'total_classes': 0, 'today': {'scheduled': 0, 'ongoing': 0, 'completed': 0, 'cancelled': 0}}
        }), 500

@bp.route('/api/v1/timetable/week')
@login_required
@admin_required
def api_timetable_week():
    """Get weekly timetable data with working filters"""
    try:
        date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        target_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        
        # Get start of week (Monday)
        start_of_week = target_date - timedelta(days=target_date.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        
        # Apply same filters as daily view
        search = request.args.get('search', '').strip()
        tutor_id = request.args.get('tutor_id', type=int)
        student_id = request.args.get('student_id', type=int)
        department_id = request.args.get('department_id', type=int)
        
        query = Class.query.filter(
            Class.scheduled_date >= start_of_week,
            Class.scheduled_date <= end_of_week
        )
        
        # Apply filters (same logic as daily)
        if tutor_id:
            query = query.filter(Class.tutor_id == tutor_id)
        
        if student_id:
            query = query.filter(
                db.or_(
                    Class.primary_student_id == student_id,
                    Class.demo_student_id == student_id,
                    Class.students.like(f'%{student_id}%')
                )
            )
        
        if department_id:
            query = query.join(Tutor, Class.tutor_id == Tutor.id)\
                         .join(User, Tutor.user_id == User.id)\
                         .filter(User.department_id == department_id)
        
        if search:
            query = query.join(Tutor, Class.tutor_id == Tutor.id, isouter=True)\
                         .join(User, Tutor.user_id == User.id, isouter=True)\
                         .filter(
                db.or_(
                    Class.subject.ilike(f'%{search}%'),
                    User.full_name.ilike(f'%{search}%'),
                    Class.grade.ilike(f'%{search}%')
                )
            )
        
        classes = query.order_by(Class.scheduled_date, Class.scheduled_time).all()
        
        # Group classes by date
        classes_by_date = {}
        for cls in classes:
            date_str = cls.scheduled_date.strftime('%Y-%m-%d')
            if date_str not in classes_by_date:
                classes_by_date[date_str] = []
            
            # Build class data
            tutor_name = cls.tutor.user.full_name if cls.tutor and cls.tutor.user else 'No Tutor'
            student_count = 0
            if cls.primary_student_id:
                student_count = 1
            elif cls.students:
                try:
                    import json
                    student_ids = json.loads(cls.students)
                    student_count = len(student_ids)
                except:
                    pass
            
            classes_by_date[date_str].append({
                'id': cls.id,
                'subject': cls.subject,
                'time': cls.scheduled_time.strftime('%H:%M') if cls.scheduled_time else '00:00',
                'tutor_name': tutor_name,
                'student_count': student_count,
                'status': cls.status,
                'duration': cls.duration
            })
        
        return jsonify({
            'success': True,
            'classes_by_date': classes_by_date,
            'week_start': start_of_week.strftime('%Y-%m-%d'),
            'week_end': end_of_week.strftime('%Y-%m-%d'),
            'total_classes': len(classes)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

        
@bp.route('/api/v1/timetable/monthly-stats')
@login_required
@admin_required
def api_monthly_stats():
    """Get monthly statistics with filters"""
    try:
        year = request.args.get('year', datetime.now().year, type=int)
        
        # Apply filters for monthly stats too
        tutor_id = request.args.get('tutor_id', type=int)
        student_id = request.args.get('student_id', type=int)
        department_id = request.args.get('department_id', type=int)
        
        monthly_stats = {}
        
        # Get data for each month
        for month in range(1, 13):
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
            
            # Count classes for this month with filters
            query = Class.query.filter(
                Class.scheduled_date >= start_date,
                Class.scheduled_date <= end_date
            )
            
            # Apply same filters
            if tutor_id:
                query = query.filter(Class.tutor_id == tutor_id)
            
            if student_id:
                query = query.filter(
                    db.or_(
                        Class.primary_student_id == student_id,
                        Class.demo_student_id == student_id,
                        Class.students.like(f'%{student_id}%')
                    )
                )
            
            if department_id:
                query = query.join(Tutor, Class.tutor_id == Tutor.id)\
                             .join(User, Tutor.user_id == User.id)\
                             .filter(User.department_id == department_id)
            
            month_classes = query.count()
            
            month_names = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                          'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
            
            monthly_stats[month_names[month-1]] = month_classes
        
        return jsonify({
            'success': True,
            'stats': monthly_stats,
            'year': year
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
        
        
@bp.route('/api/v1/timetable/class/<int:class_id>')
@login_required
@admin_required
def api_get_class_details(class_id):
    """Get detailed information about a specific class"""
    try:
        cls = Class.query.get_or_404(class_id)
        
        # Get tutor details
        tutor_info = {
            'id': None,
            'name': 'No Tutor Assigned',
            'email': '',
            'phone': '',
            'subjects': []
        }
        
        if cls.tutor and cls.tutor.user:
            tutor_info = {
                'id': cls.tutor.id,
                'name': cls.tutor.user.full_name,
                'email': cls.tutor.user.email or '',
                'phone': cls.tutor.user.phone or '',
                'subjects': cls.tutor.get_subjects() if hasattr(cls.tutor, 'get_subjects') else []
            }
        
        # Get student details
        students_info = []
        
        if cls.class_type == 'demo' and cls.demo_student_id:
            from app.models.demo_student import DemoStudent
            demo_student = DemoStudent.query.get(cls.demo_student_id)
            if demo_student:
                students_info.append({
                    'id': demo_student.id,
                    'name': demo_student.full_name,
                    'email': demo_student.email,
                    'phone': demo_student.phone,
                    'type': 'demo'
                })
        
        elif cls.primary_student_id:
            student = Student.query.get(cls.primary_student_id)
            if student:
                students_info.append({
                    'id': student.id,
                    'name': student.full_name,
                    'email': student.email,
                    'phone': student.phone,
                    'type': 'regular'
                })
        
        elif cls.students:
            try:
                import json
                student_ids = json.loads(cls.students)
                for student_id in student_ids:
                    student = Student.query.get(student_id)
                    if student:
                        students_info.append({
                            'id': student.id,
                            'name': student.full_name,
                            'email': student.email,
                            'phone': student.phone,
                            'type': 'regular'
                        })
            except (json.JSONDecodeError, TypeError):
                pass
        
        class_details = {
            'id': cls.id,
            'subject': cls.subject,
            'class_type': cls.class_type,
            'grade': cls.grade,
            'board': cls.board,
            'scheduled_date': cls.scheduled_date.strftime('%Y-%m-%d'),
            'scheduled_time': cls.scheduled_time.strftime('%H:%M') if cls.scheduled_time else '00:00',
            'duration': cls.duration,
            'status': cls.status,
            'platform': cls.platform,
            'meeting_link': cls.meeting_link,
            'meeting_id': cls.meeting_id,
            'class_notes': cls.class_notes,
            'topics_covered': cls.topics_covered,
            'homework_assigned': cls.homework_assigned,
            'tutor': tutor_info,
            'students': students_info,
            'created_at': cls.created_at.strftime('%Y-%m-%d %H:%M') if cls.created_at else '',
            'actual_start_time': cls.actual_start_time.strftime('%Y-%m-%d %H:%M') if cls.actual_start_time else None,
            'actual_end_time': cls.actual_end_time.strftime('%Y-%m-%d %H:%M') if cls.actual_end_time else None
        }
        
        return jsonify({
            'success': True,
            'class': class_details
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

        
@bp.route('/api/v1/timetable/month-details/<int:month>')
@login_required
@admin_required
def api_month_details(month):
    """Get detailed classes for a specific month"""
    try:
        year = request.args.get('year', datetime.now().year, type=int)
        
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        classes = Class.query.filter(
            Class.scheduled_date >= start_date,
            Class.scheduled_date <= end_date
        ).order_by(Class.scheduled_date, Class.scheduled_time).all()
        
        # Group classes by date
        classes_by_date = {}
        for cls in classes:
            date_str = cls.scheduled_date.strftime('%Y-%m-%d')
            if date_str not in classes_by_date:
                classes_by_date[date_str] = []
            
            tutor_name = 'No Tutor'
            if cls.tutor and cls.tutor.user:
                tutor_name = cls.tutor.user.full_name
            
            classes_by_date[date_str].append({
                'id': cls.id,
                'subject': cls.subject,
                'time': cls.scheduled_time.strftime('%H:%M') if cls.scheduled_time else '00:00',
                'tutor_name': tutor_name,
                'status': cls.status
            })
        
        return jsonify({
            'success': True,
            'classes_by_date': classes_by_date,
            'month': month,
            'year': year,
            'total_classes': len(classes)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
        
@bp.route('/api/v1/timetable/year')
@login_required
@admin_required
def api_timetable_year():
    """Get yearly timetable data with detailed calendar structure"""
    try:
        year = request.args.get('year', datetime.now().year, type=int)
        search = request.args.get('search', '').strip()
        tutor_id = request.args.get('tutor_id', type=int)
        student_id = request.args.get('student_id', type=int)
        department_id = request.args.get('department_id', type=int)
        
        # Get all classes for the year
        start_date = datetime(year, 1, 1).date()
        end_date = datetime(year, 12, 31).date()
        
        # Build query with filters
        query = Class.query.filter(
            Class.scheduled_date >= start_date,
            Class.scheduled_date <= end_date
        )
        
        # Apply search filters
        if tutor_id:
            query = query.filter(Class.tutor_id == tutor_id)
        
        if student_id:
            query = query.filter(
                db.or_(
                    Class.primary_student_id == student_id,
                    Class.demo_student_id == student_id,
                    Class.students.like(f'%{student_id}%')
                )
            )
        
        if department_id:
            query = query.join(Tutor, Class.tutor_id == Tutor.id)\
                         .join(User, Tutor.user_id == User.id)\
                         .filter(User.department_id == department_id)
        
        if search:
            query = query.join(Tutor, Class.tutor_id == Tutor.id, isouter=True)\
                         .join(User, Tutor.user_id == User.id, isouter=True)\
                         .filter(
                db.or_(
                    Class.subject.ilike(f'%{search}%'),
                    User.full_name.ilike(f'%{search}%'),
                    Class.grade.ilike(f'%{search}%')
                )
            )
        
        # Get all classes
        classes = query.order_by(Class.scheduled_date, Class.scheduled_time).all()
        
        # Group classes by month and date
        year_data = {}
        monthly_stats = {}
        
        # Initialize 12 months
        for month in range(1, 13):
            month_key = datetime(year, month, 1).strftime('%B').lower()
            year_data[month] = {}
            monthly_stats[month_key] = 0
        
        # Process each class
        for cls in classes:
            try:
                month = cls.scheduled_date.month
                date_key = cls.scheduled_date.strftime('%Y-%m-%d')
                
                # Initialize date if not exists
                if date_key not in year_data[month]:
                    year_data[month][date_key] = []
                
                # Get tutor name safely
                tutor_name = 'No Tutor Assigned'
                tutor_email = ''
                department_name = ''
                if cls.tutor and cls.tutor.user:
                    tutor_name = cls.tutor.user.full_name
                    tutor_email = cls.tutor.user.email or ''
                    if cls.tutor.user.department:
                        department_name = cls.tutor.user.department.name
                
                # Get student details
                student_count = 0
                student_names = []
                student_emails = []
                
                if cls.class_type == 'demo' and cls.demo_student_id:
                    from app.models.demo_student import DemoStudent
                    demo_student = DemoStudent.query.get(cls.demo_student_id)
                    if demo_student:
                        student_count = 1
                        student_names = [demo_student.full_name]
                        student_emails = [demo_student.email or '']
                elif cls.primary_student_id:
                    student = Student.query.get(cls.primary_student_id)
                    if student:
                        student_count = 1
                        student_names = [student.full_name]
                        student_emails = [student.email or '']
                elif cls.students:
                    try:
                        import json
                        student_ids = json.loads(cls.students)
                        students = Student.query.filter(Student.id.in_(student_ids)).all()
                        student_count = len(students)
                        student_names = [s.full_name for s in students]
                        student_emails = [s.email or '' for s in students]
                    except (json.JSONDecodeError, TypeError):
                        student_count = 0
                        student_names = []
                        student_emails = []
                
                # Create class data
                class_data = {
                    'id': cls.id,
                    'subject': cls.subject or 'No Subject',
                    'class_type': cls.class_type or 'regular',
                    'scheduled_date': cls.scheduled_date.strftime('%Y-%m-%d'),
                    'scheduled_time': cls.scheduled_time.strftime('%H:%M') if cls.scheduled_time else '00:00',
                    'duration': cls.duration or 60,
                    'status': cls.status or 'scheduled',
                    'tutor_name': tutor_name,
                    'tutor_email': tutor_email,
                    'tutor_id': cls.tutor_id,
                    'department_name': department_name,
                    'student_count': student_count,
                    'student_names': student_names,
                    'student_emails': student_emails,
                    'grade': cls.grade or '',
                    'board': cls.board or '',
                    'meeting_link': cls.meeting_link or '',
                    'platform': cls.platform or '',
                    'class_notes': cls.class_notes or ''
                }
                
                # Add to year data
                year_data[month][date_key].append(class_data)
                
                # Update monthly stats
                month_key = cls.scheduled_date.strftime('%B').lower()
                monthly_stats[month_key] = monthly_stats.get(month_key, 0) + 1
                
            except Exception as e:
                print(f"Error processing class {cls.id}: {str(e)}")
                continue
        
        # Calculate total statistics
        total_classes = len(classes)
        scheduled_count = len([c for c in classes if c.status == 'scheduled'])
        completed_count = len([c for c in classes if c.status == 'completed'])
        cancelled_count = len([c for c in classes if c.status == 'cancelled'])
        ongoing_count = len([c for c in classes if c.status == 'ongoing'])
        
        # Create month details for calendar rendering
        months_detail = []
        month_names = [
            'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December'
        ]
        
        for month_num in range(1, 13):
            month_name = month_names[month_num - 1]
            month_key = month_name.lower()
            month_classes_count = monthly_stats.get(month_key, 0)
            
            months_detail.append({
                'month_number': month_num,
                'month_name': month_name,
                'month_key': month_key,
                'class_count': month_classes_count,
                'classes_by_date': year_data.get(month_num, {})
            })
        
        return jsonify({
            'success': True,
            'year': year,
            'total_classes': total_classes,
            'months_detail': months_detail,
            'monthly_stats': monthly_stats,
            'year_classes_by_month': year_data,
            'stats': {
                'total_classes': total_classes,
                'scheduled': scheduled_count,
                'completed': completed_count,
                'cancelled': cancelled_count,
                'ongoing': ongoing_count
            },
            'filters_applied': {
                'search': search,
                'tutor_id': tutor_id,
                'student_id': student_id,
                'department_id': department_id
            }
        })
        
    except Exception as e:
        print(f"Error in api_timetable_year: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'year': year,
            'months_detail': [],
            'monthly_stats': {},
            'stats': {'total_classes': 0, 'scheduled': 0, 'completed': 0, 'cancelled': 0, 'ongoing': 0}
        }), 500
       
@bp.route('/api/v1/tutor/<int:tutor_id>/details')
@login_required
@admin_required
def api_tutor_details(tutor_id):
    """Get tutor detailed information"""
    try:
        tutor = Tutor.query.get_or_404(tutor_id)
        
        details = {
            'name': tutor.user.full_name if tutor.user else 'Unknown',
            'email': tutor.user.email if tutor.user else 'No email',
            'phone': tutor.user.phone if tutor.user else None,
            'experience': tutor.experience,
            'subjects': tutor.get_subjects(),
            'status': tutor.status,
            'qualification': tutor.qualification,
            'hourly_rate': tutor.hourly_rate,
            'monthly_salary': tutor.monthly_salary
        }
        
        return jsonify({
            'success': True,
            'details': details
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
               
        
# ============ QUICK CLASS CREATION ============

@bp.route('/api/v1/timetable/quick-class', methods=['POST'])
@login_required
@admin_required
def api_create_quick_class():
    """Create a quick class - OPTIONAL FEATURE"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['subject', 'tutor_id', 'scheduled_date', 'scheduled_time', 'duration']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
        
        # Parse date and time
        scheduled_date = datetime.strptime(data['scheduled_date'], '%Y-%m-%d').date()
        scheduled_time = datetime.strptime(data['scheduled_time'], '%H:%M').time()
        
        # Check for conflicts (if you have this method)
        if hasattr(Class, 'check_time_conflict'):
            conflict_exists, conflicting_class = Class.check_time_conflict(
                data['tutor_id'],
                scheduled_date,
                scheduled_time,
                data['duration']
            )
            
            if conflict_exists:
                return jsonify({
                    'success': False,
                    'error': f'Time conflict with existing class: {conflicting_class.subject} at {conflicting_class.scheduled_time}'
                }), 400
        
        # Create new class
        new_class = Class(
            subject=data['subject'],
            class_type=data.get('class_type', 'one_on_one'),
            grade=data.get('grade', ''),
            board=data.get('board', ''),
            scheduled_date=scheduled_date,
            scheduled_time=scheduled_time,
            duration=data['duration'],
            tutor_id=data['tutor_id'],
            primary_student_id=data.get('student_id') if data.get('student_id') else None,
            meeting_link=data.get('meeting_link', ''),
            platform=data.get('platform', 'zoom'),
            class_notes=data.get('notes', ''),
            status='scheduled',
            created_by=current_user.id
        )
        
        # Calculate end time (if you have this method)
        if hasattr(new_class, 'calculate_end_time'):
            new_class.calculate_end_time()
        
        db.session.add(new_class)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Class created successfully',
            'class_id': new_class.id,
            'class_data': {
                'id': new_class.id,
                'subject': new_class.subject,
                'scheduled_date': new_class.scheduled_date.strftime('%Y-%m-%d'),
                'scheduled_time': new_class.scheduled_time.strftime('%H:%M'),
                'tutor_name': new_class.tutor.user.full_name if new_class.tutor and new_class.tutor.user else 'Unknown'
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    

# ============ EXPORT FUNCTIONALITY ============
import csv
import io
from datetime import timedelta
try:
    from icalendar import Calendar, Event
    from pytz import timezone
    ICALENDAR_AVAILABLE = True
except ImportError:
    ICALENDAR_AVAILABLE = False
    print("⚠️ icalendar not available - calendar export will use simple format")


@bp.route('/timetable/export')
@login_required
@admin_required
def timetable_export():
    """Timetable export management page"""
    tutors = Tutor.query.all()
    departments = Department.query.filter_by(is_active=True).all()
    
    return render_template('admin/timetable_export.html', 
                         tutors=tutors, 
                         departments=departments)


# REPLACE the generate_timetable_excel() function in admin.py with this FIXED version:

def generate_timetable_excel(classes, period_name):
    """Generate Excel export for timetable - FIXED VERSION"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"
        
        # Title
        ws['A1'] = f"Class Timetable - {period_name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')  # Changed to H1 to match 8 columns
        
        # Headers (row 3 to leave space after title)
        headers = ['Date', 'Time', 'End Time', 'Subject', 'Tutor', 'Students', 'Duration', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows (starting from row 4)
        for row, cls in enumerate(classes, 4):
            tutor_name = cls.tutor.user.full_name if cls.tutor and cls.tutor.user else 'No Tutor'
            
            # Get student names
            students = 'No Students'
            if cls.class_type == 'demo':
                students = 'Demo Student'
            else:
                try:
                    student_objects = cls.get_student_objects()
                    if student_objects:
                        students = ', '.join([s.full_name for s in student_objects])
                except:
                    students = 'No Students'
            
            # Calculate end time
            end_time = ''
            if cls.scheduled_time and cls.duration:
                try:
                    start_datetime = datetime.combine(datetime.today(), cls.scheduled_time)
                    end_datetime = start_datetime + timedelta(minutes=cls.duration)
                    end_time = end_datetime.time().strftime('%H:%M')
                except:
                    end_time = 'N/A'
            
            data = [
                cls.scheduled_date.strftime('%Y-%m-%d') if cls.scheduled_date else '',
                cls.scheduled_time.strftime('%H:%M') if cls.scheduled_time else '',
                end_time,
                cls.subject or '',
                tutor_name,
                students,
                f"{cls.duration} min" if cls.duration else '',
                cls.status.title() if cls.status else ''
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                # Add border and alignment for better readability
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # FIXED: Auto-adjust column widths (avoiding MergedCell error)
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        default_widths = [12, 8, 8, 30, 20, 25, 10, 12]  # Default widths for each column
        
        for i, col_letter in enumerate(column_letters):
            max_length = default_widths[i]
            
            # Check data rows only (skip merged cells)
            for row in range(3, ws.max_row + 1):
                cell = ws[f'{col_letter}{row}']
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = min(cell_length + 2, 50)  # Cap at 50
            
            ws.column_dimensions[col_letter].width = max_length
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
        
    except ImportError:
        raise Exception("openpyxl not installed. Please install: pip install openpyxl")
    except Exception as e:
        print(f"Excel generation error: {str(e)}")
        raise e

@bp.route('/api/v1/timetable/export-preview', methods=['POST'])
@login_required
@admin_required
def api_export_preview():
    """Generate detailed export preview with real data - FIXED VERSION"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data received'}), 400
        
        export_type = data.get('export_type', 'pdf')
        period = data.get('period', 'today')
        date_param = data.get('date', datetime.now().strftime('%Y-%m-%d'))
        scope = data.get('scope', 'all')
        tutor_id = data.get('tutor_id')
        department_id = data.get('department_id')
        subject_filter = data.get('subject_filter', '')
        
        print(f"📊 Preview request: {export_type}, {period}, {date_param}")
        
        # Parse date
        try:
            target_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        except ValueError as e:
            return jsonify({'success': False, 'error': f'Invalid date format: {date_param}'}), 400
        
        # Build query based on period
        if period == 'today':
            query = Class.query.filter(Class.scheduled_date == target_date)
            period_name = target_date.strftime('%B %d, %Y')
        elif period == 'week':
            start_of_week = target_date - timedelta(days=target_date.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            query = Class.query.filter(
                Class.scheduled_date >= start_of_week,
                Class.scheduled_date <= end_of_week
            )
            period_name = f"Week of {start_of_week.strftime('%B %d, %Y')}"
        elif period == 'month':
            first_day = target_date.replace(day=1)
            if target_date.month == 12:
                last_day = target_date.replace(year=target_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                last_day = target_date.replace(month=target_date.month + 1, day=1) - timedelta(days=1)
            query = Class.query.filter(
                Class.scheduled_date >= first_day,
                Class.scheduled_date <= last_day
            )
            period_name = target_date.strftime('%B %Y')
        elif period == 'quarter':
            quarter = (target_date.month - 1) // 3 + 1
            quarter_start = target_date.replace(month=(quarter-1)*3 + 1, day=1)
            if quarter == 4:
                quarter_end = target_date.replace(year=target_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                quarter_end = target_date.replace(month=quarter*3 + 1, day=1) - timedelta(days=1)
            query = Class.query.filter(
                Class.scheduled_date >= quarter_start,
                Class.scheduled_date <= quarter_end
            )
            period_name = f"Q{quarter} {target_date.year}"
        elif period == 'year':
            year_start = target_date.replace(month=1, day=1)
            year_end = target_date.replace(month=12, day=31)
            query = Class.query.filter(
                Class.scheduled_date >= year_start,
                Class.scheduled_date <= year_end
            )
            period_name = str(target_date.year)
        else:
            query = Class.query.filter(Class.scheduled_date == target_date)
            period_name = target_date.strftime('%B %d, %Y')
        
        # Apply filters
        if scope != 'all':
            query = query.filter(Class.status == scope)
        
        if tutor_id:
            query = query.filter(Class.tutor_id == tutor_id)
        
        if department_id:
            # Join with tutor and user to filter by department
            query = query.join(Tutor).join(User).filter(User.department_id == department_id)
        
        if subject_filter:
            query = query.filter(Class.subject.contains(subject_filter))
        
        # Department access check for coordinators
        if current_user.role == 'coordinator':
            query = query.join(Tutor).join(User).filter(User.department_id == current_user.department_id)
        
        # Get classes
        classes = query.order_by(Class.scheduled_date, Class.scheduled_time).all()
        
        print(f"📊 Found {len(classes)} classes")
        
        # Calculate statistics
        total_records = len(classes)
        total_hours = sum([cls.duration / 60 for cls in classes if cls.duration]) if classes else 0
        
        # Get unique tutors and students
        unique_tutors = len(set([cls.tutor_id for cls in classes if cls.tutor_id]))
        
        unique_students = set()
        for cls in classes:
            if cls.primary_student_id:
                unique_students.add(cls.primary_student_id)
            if cls.students:
                try:
                    import json
                    student_ids = json.loads(cls.students) if isinstance(cls.students, str) else cls.students
                    if isinstance(student_ids, list):
                        unique_students.update(student_ids)
                except:
                    pass
        unique_students = len(unique_students)
        
        # Status breakdown
        status_counts = {}
        for cls in classes:
            status = cls.status or 'unknown'
            status_counts[status] = status_counts.get(status, 0) + 1
        
        # Estimate file size and pages
        if export_type == 'pdf':
            estimated_pages = max(1, (total_records // 20) + 1)
            estimated_size = f"{estimated_pages * 50}KB"
        elif export_type == 'csv':
            estimated_size = f"{max(1, total_records * 0.5):.1f}KB"
        elif export_type == 'calendar':
            estimated_size = f"{max(1, total_records * 1.2):.1f}KB"
        else:
            estimated_size = "Unknown"
        
        # Generate preview HTML
        preview_html = f"""
        <div style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;">
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 1.5rem; border-radius: 8px; margin-bottom: 1.5rem;">
                <h3 style="margin: 0 0 0.5rem 0; font-size: 1.25rem;">📊 {export_type.upper()} Export Preview</h3>
                <p style="margin: 0; opacity: 0.9;">Period: {period_name} | {total_records} classes found</p>
            </div>
        """
        
        if classes:
            preview_html += """
            <div style="background: white; border: 1px solid #e5e7eb; border-radius: 8px; overflow: hidden; margin-bottom: 1.5rem;">
                <table style="width: 100%; border-collapse: collapse;">
                    <thead>
                        <tr style="background: #f9fafb;">
                            <th style="padding: 0.75rem; text-align: left; border-bottom: 1px solid #e5e7eb; font-weight: 600;">Date</th>
                            <th style="padding: 0.75rem; text-align: left; border-bottom: 1px solid #e5e7eb; font-weight: 600;">Time</th>
                            <th style="padding: 0.75rem; text-align: left; border-bottom: 1px solid #e5e7eb; font-weight: 600;">Subject</th>
                            <th style="padding: 0.75rem; text-align: left; border-bottom: 1px solid #e5e7eb; font-weight: 600;">Tutor</th>
                            <th style="padding: 0.75rem; text-align: left; border-bottom: 1px solid #e5e7eb; font-weight: 600;">Status</th>
                        </tr>
                    </thead>
                    <tbody>
            """
            
            # Show first 10 classes in preview
            for i, cls in enumerate(classes[:10]):
                tutor_name = cls.tutor.user.full_name if cls.tutor and cls.tutor.user else 'No Tutor'
                status_color = '#10b981' if cls.status == 'completed' else '#3b82f6' if cls.status == 'scheduled' else '#ef4444'
                
                preview_html += f"""
                        <tr style="{'background: #f9fafb;' if i % 2 == 0 else ''}">
                            <td style="padding: 0.75rem; border-bottom: 1px solid #f3f4f6;">{cls.scheduled_date.strftime('%Y-%m-%d') if cls.scheduled_date else 'N/A'}</td>
                            <td style="padding: 0.75rem; border-bottom: 1px solid #f3f4f6;">{cls.scheduled_time.strftime('%H:%M') if cls.scheduled_time else 'N/A'}</td>
                            <td style="padding: 0.75rem; border-bottom: 1px solid #f3f4f6;">{cls.subject or 'N/A'}</td>
                            <td style="padding: 0.75rem; border-bottom: 1px solid #f3f4f6;">{tutor_name}</td>
                            <td style="padding: 0.75rem; border-bottom: 1px solid #f3f4f6;">
                                <span style="background: {status_color}; color: white; padding: 0.25rem 0.5rem; border-radius: 4px; font-size: 0.75rem; font-weight: 500;">
                                    {(cls.status or 'unknown').title()}
                                </span>
                            </td>
                        </tr>
                """
            
            if len(classes) > 10:
                preview_html += f"""
                        <tr>
                            <td colspan="5" style="padding: 0.75rem; text-align: center; color: #6b7280; font-style: italic;">
                                ... and {len(classes) - 10} more classes
                            </td>
                        </tr>
                """
        
            preview_html += """
                    </tbody>
                </table>
            </div>
            """
            
            # Status breakdown
            preview_html += """
            <div style="background: #f8f9fa; padding: 1rem; border-radius: 6px;">
                <h4 style="margin: 0 0 1rem 0;">Status Breakdown</h4>
            """
            
            for status, count in status_counts.items():
                percentage = (count / total_records * 100) if total_records > 0 else 0
                color = '#10b981' if status == 'completed' else '#3b82f6' if status == 'scheduled' else '#ef4444'
                preview_html += f"""
                    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 0.5rem;">
                        <span style="text-transform: capitalize;">{status}</span>
                        <div style="display: flex; align-items: center; gap: 0.5rem;">
                            <div style="width: 100px; height: 8px; background: #e5e7eb; border-radius: 4px; overflow: hidden;">
                                <div style="width: {percentage}%; height: 100%; background: {color};"></div>
                            </div>
                            <span style="font-weight: bold; min-width: 60px;">{count} ({percentage:.1f}%)</span>
                        </div>
                    </div>
                """
            
            preview_html += """
            </div>
            """
        else:
            preview_html += """
            <div style="text-align: center; padding: 3rem; color: #6b7280;">
                <div style="font-size: 3rem; margin-bottom: 1rem;">📅</div>
                <h3 style="margin: 0 0 0.5rem 0;">No Classes Found</h3>
                <p style="margin: 0;">No classes match your selected criteria for this period.</p>
            </div>
            """
        
        preview_html += """
        </div>
        """
        
        return jsonify({
            'success': True,
            'total_records': total_records,
            'estimated_size': estimated_size,
            'page_count': estimated_pages if export_type == 'pdf' else 1,
            'generation_time': "0.1",
            'preview_html': preview_html,
            'statistics': {
                'total_classes': total_records,
                'total_hours': round(total_hours, 1),
                'unique_tutors': unique_tutors,
                'unique_students': unique_students,
                'status_breakdown': status_counts,
                'period_name': period_name
            }
        })
        
    except Exception as e:
        print(f"📊 Export preview error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    
    
    
@bp.route('/api/v1/timetable/export-pdf')
@login_required
@admin_required
def api_export_timetable_pdf():
    """Fixed PDF export endpoint"""
    try:
        date_param = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        view = request.args.get('view', 'today')
        
        # Use your existing PDF generation logic
        target_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        
        if view == 'today':
            classes = Class.query.filter(Class.scheduled_date == target_date).all()
        elif view == 'week':
            start_of_week = target_date - timedelta(days=target_date.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            classes = Class.query.filter(
                Class.scheduled_date >= start_of_week,
                Class.scheduled_date <= end_of_week
            ).all()
        else:  # month
            first_day = target_date.replace(day=1)
            if target_date.month == 12:
                last_day = target_date.replace(year=target_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                last_day = target_date.replace(month=target_date.month + 1, day=1) - timedelta(days=1)
            classes = Class.query.filter(
                Class.scheduled_date >= first_day,
                Class.scheduled_date <= last_day
            ).all()
        
        # Generate PDF using your existing function
        pdf_data = generate_timetable_pdf(classes, f"{view.title()} View", view)
        
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=timetable_{date_param}.pdf'
        
        return response
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
@bp.route('/api/v1/timetable/export', methods=['POST'])
@login_required
@admin_required
def api_timetable_export():
    """Complete export system for all formats"""
    try:
        print("🚀 Export request received")
        
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data received'}), 400
        
        export_type = data.get('export_type', 'pdf')
        period = data.get('period', 'today')
        date_param = data.get('date', datetime.now().strftime('%Y-%m-%d'))
        
        print(f"📊 Export: {export_type}, Period: {period}, Date: {date_param}")
        
        # Parse date
        try:
            target_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'error': f'Invalid date: {date_param}'}), 400
        
        # Get classes based on period
        if period == 'today':
            classes = Class.query.filter(Class.scheduled_date == target_date).all()
            period_name = target_date.strftime('%B %d, %Y')
        elif period == 'week':
            start_of_week = target_date - timedelta(days=target_date.weekday())
            end_of_week = start_of_week + timedelta(days=6)
            classes = Class.query.filter(
                Class.scheduled_date >= start_of_week,
                Class.scheduled_date <= end_of_week
            ).all()
            period_name = f"Week of {start_of_week.strftime('%B %d, %Y')}"
        elif period == 'month':
            first_day = target_date.replace(day=1)
            if target_date.month == 12:
                last_day = target_date.replace(year=target_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                last_day = target_date.replace(month=target_date.month + 1, day=1) - timedelta(days=1)
            classes = Class.query.filter(
                Class.scheduled_date >= first_day,
                Class.scheduled_date <= last_day
            ).all()
            period_name = target_date.strftime('%B %Y')
        else:
            classes = Class.query.filter(Class.scheduled_date == target_date).all()
            period_name = target_date.strftime('%B %d, %Y')
        
        print(f"📊 Found {len(classes)} classes")
        
        # Generate export based on type
        if export_type == 'pdf':
            print("📄 Generating PDF")
            pdf_data = generate_timetable_pdf(classes, period_name, period)
            response = make_response(pdf_data)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename=timetable_{date_param}.pdf'
            return response
            
        elif export_type == 'csv':
            print("📊 Generating CSV")
            csv_data = generate_timetable_csv(classes, period_name)
            response = make_response(csv_data)
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['Content-Disposition'] = f'attachment; filename=timetable_{date_param}.csv'
            return response
            
        elif export_type == 'excel':
            print("📊 Generating Excel")
            excel_data = generate_timetable_excel(classes, period_name)
            response = make_response(excel_data)
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename=timetable_{date_param}.xlsx'
            return response
            
        elif export_type == 'calendar':
            print("📅 Generating Calendar")
            calendar_data = generate_timetable_calendar(classes, period_name)
            response = make_response(calendar_data)
            response.headers['Content-Type'] = 'text/calendar; charset=utf-8'
            response.headers['Content-Disposition'] = f'attachment; filename=timetable_{date_param}.ics'
            return response
            
        else:
            return jsonify({'success': False, 'error': f'Unsupported export type: {export_type}'}), 400
        
    except Exception as e:
        print(f"❌ Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    
def generate_timetable_pdf(classes, period_name, view_type):
    """Generate PDF content for timetable - FIXED VERSION"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from flask import make_response
        import io
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, margin=0.5*inch)
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#F1A150'),
            alignment=1,  # Center
            spaceAfter=20
        )
        
        # Build content
        content = []
        
        # Title
        content.append(Paragraph(f"Class Timetable - {period_name}", title_style))
        content.append(Spacer(1, 20))
        
        if classes:
            # Create table data
            table_data = [['Date', 'Time', 'Subject', 'Tutor', 'Duration', 'Status']]
            
            for cls in classes:
                tutor_name = cls.tutor.user.full_name if cls.tutor and cls.tutor.user else 'No Tutor'
                
                table_data.append([
                    cls.scheduled_date.strftime('%m/%d/%Y'),
                    cls.scheduled_time.strftime('%H:%M') if cls.scheduled_time else '00:00',
                    cls.subject[:30] + '...' if len(cls.subject) > 30 else cls.subject,
                    tutor_name[:20] + '...' if len(tutor_name) > 20 else tutor_name,
                    f"{cls.duration} min",
                    cls.status.title()
                ])
            
            # Create table
            table = Table(table_data, colWidths=[1*inch, 0.8*inch, 2*inch, 1.5*inch, 0.8*inch, 0.8*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F1A150')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            content.append(table)
            
            # Add summary
            content.append(Spacer(1, 30))
            content.append(Paragraph(f"<b>Total Classes:</b> {len(classes)}", styles['Normal']))
            content.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
            
        else:
            content.append(Paragraph("No classes scheduled for this period", styles['Normal']))
        
        # Build PDF
        doc.build(content)
        buffer.seek(0)
        
        return buffer.read()
        
    except Exception as e:
        print(f"PDF generation error: {str(e)}")
        raise e
    
def generate_timetable_csv(classes, period_name):
    """Generate CSV export for timetable"""
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Date', 'Time', 'End Time', 'Subject', 'Tutor', 'Students', 
            'Duration (mins)', 'Status', 'Class Type', 'Grade', 'Board',
            'Platform', 'Meeting Link', 'Notes'
        ])
        
        # Write data rows
        for cls in classes:
            tutor_name = cls.tutor.user.full_name if cls.tutor and cls.tutor.user else 'No Tutor'
            
            # Get student names
            students = 'No Students'
            if cls.class_type == 'demo':
                students = 'Demo Student'
            else:
                student_objects = cls.get_student_objects()
                if student_objects:
                    students = ', '.join([s.full_name for s in student_objects])
            
            # Calculate end time
            end_time = ''
            if cls.scheduled_time and cls.duration:
                start_datetime = datetime.combine(datetime.today(), cls.scheduled_time)
                end_datetime = start_datetime + timedelta(minutes=cls.duration)
                end_time = end_datetime.time().strftime('%H:%M')
            
            writer.writerow([
                cls.scheduled_date.strftime('%Y-%m-%d') if cls.scheduled_date else '',
                cls.scheduled_time.strftime('%H:%M') if cls.scheduled_time else '',
                end_time,
                cls.subject or '',
                tutor_name,
                students,
                cls.duration or '',
                cls.status.title() if cls.status else '',
                cls.class_type.replace('_', ' ').title() if cls.class_type else '',
                cls.grade or '',
                cls.board or '',
                cls.platform or '',
                cls.meeting_link or '',
                cls.class_notes or ''
            ])
        
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        print(f"CSV generation error: {str(e)}")
        raise e
    
def generate_timetable_calendar(classes, period_name):
    """Generate Calendar export for timetable"""
    try:
        if ICALENDAR_AVAILABLE:
            # Use advanced icalendar library
            cal = Calendar()
            cal.add('prodid', '-//I2Global LMS//Timetable Export//EN')
            cal.add('version', '2.0')
            cal.add('calscale', 'GREGORIAN')
            cal.add('method', 'PUBLISH')
            cal.add('x-wr-calname', f'I2Global LMS - {period_name}')
            
            tz = timezone('Asia/Kolkata')
            
            for cls in classes:
                if not cls.scheduled_date or not cls.scheduled_time:
                    continue
                    
                event = Event()
                tutor_name = cls.tutor.user.full_name if cls.tutor and cls.tutor.user else 'No Tutor'
                
                # Event details
                event.add('summary', f'{cls.subject} - {cls.class_type.replace("_", " ").title()}')
                
                description = f"""Subject: {cls.subject}
Tutor: {tutor_name}
Type: {cls.class_type.replace('_', ' ').title()}
Duration: {cls.duration} minutes
Status: {cls.status.title() if cls.status else 'N/A'}"""
                
                if cls.meeting_link:
                    description += f"\nMeeting Link: {cls.meeting_link}"
                    
                event.add('description', description)
                
                # Timing
                start_datetime = datetime.combine(cls.scheduled_date, cls.scheduled_time)
                start_datetime = tz.localize(start_datetime)
                event.add('dtstart', start_datetime)
                
                end_datetime = start_datetime + timedelta(minutes=cls.duration or 60)
                event.add('dtend', end_datetime)
                
                # Location
                event.add('location', cls.platform or 'Online Class')
                event.add('uid', f'class-{cls.id}@i2global-lms.com')
                event.add('dtstamp', datetime.now(tz))
                
                cal.add_component(event)
            
            return cal.to_ical().decode('utf-8')
        
        else:
            # Use simple iCal format
            ical_lines = [
                "BEGIN:VCALENDAR",
                "VERSION:2.0",
                "PRODID:-//I2Global LMS//Timetable Export//EN",
                "CALSCALE:GREGORIAN",
                f"X-WR-CALNAME:I2Global LMS - {period_name}"
            ]
            
            for cls in classes:
                if not cls.scheduled_date or not cls.scheduled_time:
                    continue
                    
                tutor_name = cls.tutor.user.full_name if cls.tutor and cls.tutor.user else 'No Tutor'
                start_datetime = datetime.combine(cls.scheduled_date, cls.scheduled_time)
                end_datetime = start_datetime + timedelta(minutes=cls.duration or 60)
                
                ical_lines.extend([
                    "BEGIN:VEVENT",
                    f"UID:class-{cls.id}@i2global-lms.com",
                    f"DTSTART:{start_datetime.strftime('%Y%m%dT%H%M%S')}",
                    f"DTEND:{end_datetime.strftime('%Y%m%dT%H%M%S')}",
                    f"SUMMARY:{cls.subject} - {cls.class_type}",
                    f"DESCRIPTION:Tutor: {tutor_name}\\nDuration: {cls.duration} min",
                    "LOCATION:Online Class",
                    "END:VEVENT"
                ])
            
            ical_lines.append("END:VCALENDAR")
            return '\n'.join(ical_lines)
        
    except Exception as e:
        print(f"Calendar generation error: {str(e)}")
        raise e
    
def generate_timetable_ical(classes, period_name):
    """Generate iCal export for timetable"""
    try:
        # Create calendar
        cal = Calendar()
        cal.add('prodid', '-//I2Global LMS//Timetable Export//EN')
        cal.add('version', '2.0')
        cal.add('calscale', 'GREGORIAN')
        cal.add('method', 'PUBLISH')
        cal.add('x-wr-calname', f'I2Global LMS - {period_name}')
        cal.add('x-wr-caldesc', f'Class schedule for {period_name}')
        
        # Set timezone (adjust as needed)
        tz = timezone('Asia/Kolkata')  # Change to your timezone
        
        # Add events
        for cls in classes:
            if not cls.scheduled_date or not cls.scheduled_time:
                continue
                
            event = Event()
            
            # Event details
            tutor_name = cls.tutor.user.full_name if cls.tutor and cls.tutor.user else 'No Tutor'
            
            # Get student names
            if cls.class_type == 'demo':
                students = 'Demo Student'
            else:
                student_objects = cls.get_student_objects()
                students = ', '.join([s.full_name for s in student_objects]) if student_objects else 'No Students'
            
            # Event title
            event.add('summary', f'{cls.subject} - {cls.class_type.replace("_", " ").title()}')
            
            # Event description
            description = f"""
Class Details:
Subject: {cls.subject}
Tutor: {tutor_name}
Students: {students}
Type: {cls.class_type.replace('_', ' ').title()}
Grade: {cls.grade or 'N/A'}
Board: {cls.board or 'N/A'}
Duration: {cls.duration} minutes
Status: {cls.status.title() if cls.status else 'N/A'}
Platform: {cls.platform or 'N/A'}
Meeting Link: {cls.meeting_link or 'N/A'}
Notes: {cls.class_notes or 'N/A'}
            """.strip()
            
            event.add('description', description)
            
            # Event timing
            start_datetime = datetime.combine(cls.scheduled_date, cls.scheduled_time)
            start_datetime = tz.localize(start_datetime)
            event.add('dtstart', start_datetime)
            
            # End time
            if cls.duration:
                end_datetime = start_datetime + timedelta(minutes=cls.duration)
                event.add('dtend', end_datetime)
            else:
                # Default 1 hour if no duration
                end_datetime = start_datetime + timedelta(hours=1)
                event.add('dtend', end_datetime)
            
            # Event location
            if cls.platform and cls.meeting_link:
                event.add('location', f'{cls.platform}: {cls.meeting_link}')
            elif cls.platform:
                event.add('location', cls.platform)
            else:
                event.add('location', 'Online Class')
            
            # Event categories
            event.add('categories', [cls.subject, cls.class_type, 'I2Global LMS'])
            
            # Event status based on class status
            if cls.status == 'completed':
                event.add('status', 'CONFIRMED')
            elif cls.status == 'cancelled':
                event.add('status', 'CANCELLED')
            else:
                event.add('status', 'TENTATIVE')
            
            # Unique ID
            event.add('uid', f'class-{cls.id}@i2global-lms.com')
            
            # Created timestamp
            event.add('dtstamp', datetime.now(tz))
            
            # Add alarm/reminder (15 minutes before)
            from icalendar import Alarm
            alarm = Alarm()
            alarm.add('action', 'DISPLAY')
            alarm.add('description', f'Reminder: {cls.subject} class starting in 15 minutes')
            alarm.add('trigger', timedelta(minutes=-15))
            event.add_component(alarm)
            
            cal.add_component(event)
        
        return cal.to_ical().decode('utf-8')
        
    except Exception as e:
        print(f"iCal generation error: {str(e)}")
        raise e    
    
    
@bp.route('/api/v1/timetable/enhanced-email-preview', methods=['POST'])
@login_required
@admin_required
def api_enhanced_email_preview():
    """Enhanced email preview with better recipient handling"""
    try:
        data = request.get_json()
        date_param = data.get('date', datetime.now().strftime('%Y-%m-%d'))
        period = data.get('period', 'single')
        recipients_type = data.get('recipients', 'all')
        specific_ids = data.get('specific_ids', [])
        subject_style = data.get('subject_style', 'standard')
        
        target_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        
        # Get classes based on period
        classes = get_classes_for_period(target_date, period)
        
        # Get recipients with enhanced details
        recipients_list = get_enhanced_recipients(recipients_type, specific_ids, classes)
        
        # Generate enhanced preview
        preview_html = generate_enhanced_email_preview(classes, target_date, period, recipients_list, subject_style)
        
        return jsonify({
            'success': True,
            'preview_html': preview_html,
            'recipient_count': len(recipients_list),
            'class_count': len(classes),
            'recipients_breakdown': get_recipients_breakdown(recipients_list),
            'estimated_send_time': calculate_send_time(len(recipients_list))
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ============ EMAIL PREVIEW AND SENDING ============

@bp.route('/timetable/email')
@login_required
@require_permission('class_management')
def timetable_email():
    """Timetable email management page"""
    tutors = Tutor.query.all()
    students = Student.query.all()
    departments = Department.query.filter_by(is_active=True).all()
    
    return render_template('admin/timetable_email.html', 
                         tutors=tutors, 
                         students=students, 
                         departments=departments)

    

@bp.route('/api/v1/timetable/subject-preview', methods=['POST'])
@login_required
@admin_required
def api_subject_preview():
    """Generate smart email subject preview"""
    try:
        data = request.get_json()
        date = data.get('date')
        period = data.get('period', 'single')
        style = data.get('style', 'standard')
        recipient_type = data.get('recipient_type', 'tutor')
        recipient_name = data.get('recipient_name', 'User')
        
        # Parse date for better formatting
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
        formatted_date = date_obj.strftime('%B %d, %Y')
        day_name = date_obj.strftime('%A')
        
        # Generate contextual subjects
        subjects = {}
        
        if style == 'standard':
            if period == 'single':
                subjects['standard'] = f"📋 Class Schedule for {formatted_date}"
            elif period == 'week':
                subjects['standard'] = f"📋 Weekly Schedule - Week of {formatted_date}"
            elif period == 'month':
                subjects['standard'] = f"📋 Monthly Schedule - {date_obj.strftime('%B %Y')}"
            else:
                subjects['standard'] = f"📋 Yearly Schedule - {date_obj.year}"
                
        elif style == 'personal':
            if recipient_type == 'tutor':
                subjects['personal'] = f"👋 Hi {recipient_name}! Your classes for {day_name}, {formatted_date}"
            else:
                subjects['personal'] = f"👋 Hi {recipient_name}! Your schedule for {day_name}"
                
        elif style == 'contextual':
            # Get current hour for smart timing
            current_hour = datetime.now().hour
            if current_hour < 12:
                greeting = "Good morning"
            elif current_hour < 17:
                greeting = "Good afternoon"
            else:
                greeting = "Good evening"
                
            subjects['contextual'] = f"🧠 {greeting}! Smart schedule update for {formatted_date}"
        
        return jsonify({
            'success': True,
            'subject_options': subjects,
            'formatted_date': formatted_date,
            'day_name': day_name
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============ HELPER FUNCTIONS ============

def get_classes_for_period(target_date, period):
    """Get classes for specified period"""
    if period == 'week':
        start_of_week = target_date - timedelta(days=target_date.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        return Class.query.filter(
            Class.scheduled_date >= start_of_week,
            Class.scheduled_date <= end_of_week
        ).order_by(Class.scheduled_date, Class.scheduled_time).all()
    elif period == 'month':
        first_day = target_date.replace(day=1)
        if target_date.month == 12:
            last_day = target_date.replace(year=target_date.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last_day = target_date.replace(month=target_date.month + 1, day=1) - timedelta(days=1)
        return Class.query.filter(
            Class.scheduled_date >= first_day,
            Class.scheduled_date <= last_day
        ).order_by(Class.scheduled_date, Class.scheduled_time).all()
    elif period == 'year':
        start_date = target_date.replace(month=1, day=1)
        end_date = target_date.replace(month=12, day=31)
        return Class.query.filter(
            Class.scheduled_date >= start_date,
            Class.scheduled_date <= end_date
        ).order_by(Class.scheduled_date, Class.scheduled_time).all()
    else:  # single day
        return Class.query.filter(Class.scheduled_date == target_date).order_by(Class.scheduled_time).all()

def get_enhanced_recipients(recipients_type, specific_ids, classes):
    """Get enhanced recipient list with detailed information"""
    recipients_list = []
    
    if recipients_type == 'all' or recipients_type == 'tutors':
        tutors = Tutor.query.join(User).filter(User.email.isnot(None)).all()
        for tutor in tutors:
            if tutor.user and tutor.user.email:
                # Count classes for this tutor
                tutor_classes = [cls for cls in classes if cls.tutor_id == tutor.id]
                recipients_list.append({
                    'type': 'tutor',
                    'id': tutor.id,
                    'name': tutor.user.full_name,
                    'email': tutor.user.email,
                    'department': tutor.user.department.name if tutor.user.department else 'N/A',
                    'class_count': len(tutor_classes),
                    'subjects': list(set(cls.subject for cls in tutor_classes))
                })
    
    if recipients_type == 'all' or recipients_type == 'students':
        students = Student.query.filter(Student.email.isnot(None)).all()
        for student in students:
            if student.email:
                # Count classes for this student
                student_classes = []
                for cls in classes:
                    if cls.primary_student_id == student.id:
                        student_classes.append(cls)
                    elif cls.students:
                        try:
                            import json
                            student_ids = json.loads(cls.students)
                            if student.id in student_ids:
                                student_classes.append(cls)
                        except:
                            pass
                
                recipients_list.append({
                    'type': 'student',
                    'id': student.id,
                    'name': student.full_name,
                    'email': student.email,
                    'department': student.department.name if student.department else 'N/A',
                    'class_count': len(student_classes),
                    'subjects': list(set(cls.subject for cls in student_classes))
                })
    
    if recipients_type == 'specific' and specific_ids:
        # Handle specific recipients
        for recipient_id in specific_ids:
            if recipient_id.startswith('tutor_'):
                tutor_id = int(recipient_id.replace('tutor_', ''))
                tutor = Tutor.query.get(tutor_id)
                if tutor and tutor.user and tutor.user.email:
                    tutor_classes = [cls for cls in classes if cls.tutor_id == tutor.id]
                    recipients_list.append({
                        'type': 'tutor',
                        'id': tutor.id,
                        'name': tutor.user.full_name,
                        'email': tutor.user.email,
                        'department': tutor.user.department.name if tutor.user.department else 'N/A',
                        'class_count': len(tutor_classes),
                        'subjects': list(set(cls.subject for cls in tutor_classes))
                    })
            elif recipient_id.startswith('student_'):
                student_id = int(recipient_id.replace('student_', ''))
                student = Student.query.get(student_id)
                if student and student.email:
                    student_classes = []
                    for cls in classes:
                        if cls.primary_student_id == student.id:
                            student_classes.append(cls)
                        elif cls.students:
                            try:
                                import json
                                student_ids = json.loads(cls.students)
                                if student.id in student_ids:
                                    student_classes.append(cls)
                            except:
                                pass
                    
                    recipients_list.append({
                        'type': 'student',
                        'id': student.id,
                        'name': student.full_name,
                        'email': student.email,
                        'department': student.department.name if student.department else 'N/A',
                        'class_count': len(student_classes),
                        'subjects': list(set(cls.subject for cls in student_classes))
                    })
    
    return recipients_list

def get_recipients_breakdown(recipients_list):
    """Get breakdown of recipients by type"""
    breakdown = {'tutors': 0, 'students': 0, 'total': len(recipients_list)}
    for recipient in recipients_list:
        if recipient['type'] == 'tutor':
            breakdown['tutors'] += 1
        else:
            breakdown['students'] += 1
    return breakdown

def calculate_send_time(recipient_count):
    """Calculate estimated send time"""
    # Estimate ~2 emails per second
    estimated_seconds = max(1, recipient_count // 2)
    if estimated_seconds < 60:
        return f"{estimated_seconds}s"
    else:
        return f"{estimated_seconds // 60}m {estimated_seconds % 60}s"
    
def generate_enhanced_email_preview(classes, target_date, period, recipients_list, subject_style):
    """Generate enhanced email preview with detailed information"""
    period_names = {
        'single': target_date.strftime('%B %d, %Y'),
        'week': f"Week of {target_date.strftime('%B %d, %Y')}",
        'month': target_date.strftime('%B %Y'),
        'year': f"Year {target_date.year}"
    }
    period_name = period_names.get(period, target_date.strftime('%B %d, %Y'))
    
    recipients_breakdown = get_recipients_breakdown(recipients_list)
    
    # Generate sample subject
    if subject_style == 'standard':
        sample_subject = f"📋 Class Schedule for {period_name}"
    elif subject_style == 'personal':
        sample_subject = f"👋 Hi [Name]! Your classes for {period_name}"
    else:  # contextual
        sample_subject = f"🧠 Smart schedule update for {period_name}"
    
    preview_html = f"""
    <div style="font-family: Arial, sans-serif; max-width: 100%; line-height: 1.6;">
        <div style="background: linear-gradient(135deg, #10b981, #059669); color: white; padding: 1.5rem; border-radius: 8px; margin-bottom: 1.5rem;">
            <h3 style="margin: 0 0 0.5rem 0;">📧 Email Campaign Preview</h3>
            <p style="margin: 0; opacity: 0.9;">{period_name} • {len(classes)} Classes • {len(recipients_list)} Recipients</p>
        </div>
        
        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 1rem; margin-bottom: 1.5rem;">
            <div style="text-align: center; background: #f8f9fa; padding: 1rem; border-radius: 6px;">
                <div style="font-size: 1.5rem; font-weight: bold; color: #10b981;">{recipients_breakdown['total']}</div>
                <div style="font-size: 0.8rem; color: #666;">Total Recipients</div>
            </div>
            <div style="text-align: center; background: #f8f9fa; padding: 1rem; border-radius: 6px;">
                <div style="font-size: 1.5rem; font-weight: bold; color: #3b82f6;">{recipients_breakdown['tutors']}</div>
                <div style="font-size: 0.8rem; color: #666;">Tutors</div>
            </div>
            <div style="text-align: center; background: #f8f9fa; padding: 1rem; border-radius: 6px;">
                <div style="font-size: 1.5rem; font-weight: bold; color: #8b5cf6;">{recipients_breakdown['students']}</div>
                <div style="font-size: 0.8rem; color: #666;">Students</div>
            </div>
            <div style="text-align: center; background: #f8f9fa; padding: 1rem; border-radius: 6px;">
                <div style="font-size: 1.5rem; font-weight: bold; color: #F1A150;">{len(classes)}</div>
                <div style="font-size: 0.8rem; color: #666;">Classes</div>
            </div>
        </div>
        
        <div style="background: white; border: 1px solid #e5e7eb; border-radius: 8px; margin-bottom: 1.5rem; overflow: hidden;">
            <div style="background: #F1A150; color: white; padding: 1rem;">
                <h4 style="margin: 0;">📧 Sample Email</h4>
            </div>
            <div style="padding: 1.5rem;">
                <div style="margin-bottom: 1rem;">
                    <strong>Subject:</strong> <span style="color: #059669;">{sample_subject}</span>
                </div>
                <div style="border: 1px solid #e5e7eb; padding: 1rem; background: #f8f9fa; border-radius: 6px;">
                    <p style="margin: 0 0 1rem 0;"><strong>Dear [Recipient Name],</strong></p>
                    <p style="margin: 0 0 1rem 0;">Here's your schedule for {period_name}:</p>
                    <div style="background: white; padding: 1rem; border-radius: 4px; margin: 1rem 0;">
                        <strong>Sample Class Information:</strong><br>
    """
    
    # Add sample classes
    for i, cls in enumerate(classes[:3]):
        tutor_name = cls.tutor.user.full_name if cls.tutor and cls.tutor.user else 'No Tutor'
        preview_html += f"""
                        • {cls.subject} - {cls.scheduled_date.strftime('%m/%d')} at {cls.scheduled_time.strftime('%H:%M') if cls.scheduled_time else '00:00'} with {tutor_name}<br>
        """
    
    if len(classes) > 3:
        preview_html += f"                        ... and {len(classes) - 3} more classes<br>"
    
    preview_html += """
                    </div>
                    <p style="margin: 0;">Best regards,<br>The Academic Team</p>
                </div>
            </div>
        </div>
        
        <div style="background: #f8f9fa; border-radius: 8px; padding: 1rem;">
            <h4 style="margin: 0 0 1rem 0;">Recipients Sample (First 10)</h4>
            <div style="max-height: 200px; overflow-y: auto;">
    """
    
    for i, recipient in enumerate(recipients_list[:10]):
        preview_html += f"""
                <div style="display: flex; justify-content: space-between; align-items: center; padding: 0.5rem; background: white; margin-bottom: 0.5rem; border-radius: 4px; border: 1px solid #e5e7eb;">
                    <div>
                        <strong>{recipient['name']}</strong><br>
                        <small style="color: #666;">{recipient['email']} • {recipient['type'].title()}</small>
                    </div>
                    <div style="text-align: right;">
                        <small style="background: #{'3b82f6' if recipient['type'] == 'tutor' else '#8b5cf6'}; color: white; padding: 0.25rem 0.5rem; border-radius: 12px;">
                            {recipient['class_count']} classes
                        </small>
                    </div>
                </div>
        """
    
    if len(recipients_list) > 10:
        preview_html += f"""
                <div style="text-align: center; padding: 1rem; color: #666; font-style: italic;">
                    ... and {len(recipients_list) - 10} more recipients
                </div>
        """
    
    preview_html += """
            </div>
        </div>
    </div>
    """
    
    return preview_html


# ============ DASHBOARD REDIRECT ============

@bp.route('/dashboard') 
@login_required
@admin_required
def admin_dashboard():
    """Admin dashboard - redirect to main dashboard"""
    return redirect(url_for('dashboard.index'))

# ========== finance ========

@bp.route('/finance')
@login_required
@require_permission('finance_management')
def finance_dashboard():
    """Finance dashboard"""
    from datetime import datetime, date
    from sqlalchemy import func
    
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    try:
        # Get all active tutors
        tutors = Tutor.query.filter_by(status='active').all()
        
        # Calculate tutor salary data
        total_salary_expense = 0
        tutor_salaries = []
        
        for tutor in tutors:
            salary_calc = tutor.calculate_monthly_salary(current_month, current_year)
            outstanding = tutor.get_outstanding_salary()
            
            # Create calculation object that matches template expectations
            calculation_data = {
                'total_classes': salary_calc['total_classes'],
                'total_hours': salary_calc['total_classes'] * 1.0,  # Assuming 1 hour per class
                'net_salary': salary_calc['calculated_salary']
            }
            
            tutor_salaries.append({
                'tutor': tutor,
                'calculation': type('obj', (object,), calculation_data)  # Convert dict to object
            })
            
            total_salary_expense += salary_calc['calculated_salary']
        
        # Get all active students
        students = Student.query.filter_by(is_active=True).all()
        
        # Calculate fee defaulters
        total_outstanding = 0
        fee_defaulters = []
        
        for student in students:
            outstanding_amount = student.calculate_outstanding_fees()
            
            if outstanding_amount > 0:
                # Create outstanding object that matches template expectations
                outstanding_data = {
                    'outstanding_amount': outstanding_amount,
                    'overdue_amount': outstanding_amount  # Assuming all outstanding is overdue for now
                }
                
                fee_defaulters.append({
                    'student': student,
                    'outstanding': type('obj', (object,), outstanding_data)  # Convert dict to object
                })
                
                total_outstanding += outstanding_amount
        
        # Sort fee defaulters by outstanding amount (highest first)
        fee_defaulters.sort(key=lambda x: x['outstanding'].outstanding_amount, reverse=True)
        
        return render_template('admin/finance_dashboard.html',
                             month=current_month,
                             year=current_year,
                             total_salary_expense=total_salary_expense,
                             total_outstanding=total_outstanding,
                             tutor_salaries=tutor_salaries,
                             fee_defaulters=fee_defaulters)
                             
    except Exception as e:
        flash(f'Error loading finance dashboard: {str(e)}', 'error')
        return render_template('admin/finance_dashboard.html',
                             month=current_month,
                             year=current_year,
                             total_salary_expense=0,
                             total_outstanding=0,
                             tutor_salaries=[],
                             fee_defaulters=[])
    
@bp.route('/salary-generation')
@login_required
@require_permission('finance_management')
def salary_generation():
    """Salary generation page"""
    from datetime import datetime
    
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Get all active tutors
    tutors = Tutor.query.filter_by(status='active').all()
    
    tutor_salary_data = []
    for tutor in tutors:
        salary_calc = tutor.calculate_monthly_salary(current_month, current_year)
        
        # Convert dict to object-like access for template compatibility
        calculation_obj = type('obj', (object,), salary_calc)
        
        tutor_salary_data.append({
            'tutor': tutor,
            'calculation': calculation_obj,
            'outstanding': tutor.get_outstanding_salary()
        })
    
    return render_template('admin/salary_generation.html',
                         tutors=tutor_salary_data,
                         current_month=current_month,
                         current_year=current_year)
    
@bp.route('/tutors/<int:tutor_id>/salary')
@login_required
@admin_required
def tutor_salary_details(tutor_id):
    """View detailed salary information for a specific tutor"""
    from datetime import datetime
    from app.models.attendance import Attendance
    
    tutor = Tutor.query.get_or_404(tutor_id)
    
    # Get month and year from query parameters
    month = request.args.get('month', datetime.now().month, type=int)
    year = request.args.get('year', datetime.now().year, type=int)
    
    # Calculate salary for the specified period
    salary_calculation = tutor.calculate_monthly_salary(month, year)
    
    # Get attendance records for this period
    attendance_records = tutor._get_monthly_attendance_records(month, year)
    
    # Get salary payment history
    salary_history = tutor.get_salary_history()
    
    # Get outstanding amount
    outstanding_amount = tutor.get_outstanding_salary()
    
    # Calculate additional statistics
    total_late_minutes = sum(att.tutor_late_minutes or 0 for att in attendance_records)
    total_early_leaves = sum(att.tutor_early_leave_minutes or 0 for att in attendance_records)
    
    return render_template('admin/tutor_salary_details.html',
                         tutor=tutor,
                         month=month,
                         year=year,
                         salary_calculation=salary_calculation,
                         attendance_records=attendance_records,
                         salary_history=salary_history,
                         outstanding_amount=outstanding_amount,
                         total_late_minutes=total_late_minutes,
                         total_early_leaves=total_early_leaves)

@bp.route('/fee-collection')
@login_required
@require_permission('finance_management')
def fee_collection():
    """Fee collection page"""
    # Get students with outstanding fees
    students = Student.query.filter_by(is_active=True).all()
    
    students_with_fees = []
    for student in students:
        outstanding = student.calculate_outstanding_fees()
        if outstanding > 0:
            students_with_fees.append({
                'student': student,
                'outstanding': outstanding,
                'fee_structure': student.get_fee_structure()
            })
    
    # Sort by outstanding amount (highest first)
    students_with_fees.sort(key=lambda x: x['outstanding'], reverse=True)
    
    return render_template('admin/fee_collection.html',
                         students_with_fees=students_with_fees)

@bp.route('/api/v1/finance/dashboard')
@login_required
@admin_required
def finance_dashboard_api():
    """API endpoint for finance dashboard data"""
    from datetime import datetime
    
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Get summary data
    tutors = Tutor.query.filter_by(status='active').all()
    students = Student.query.filter_by(is_active=True).all()
    
    total_salary_expense = sum(
        tutor.calculate_monthly_salary(current_month, current_year)['calculated_salary'] 
        for tutor in tutors
    )
    
    total_outstanding_fees = sum(
        student.calculate_outstanding_fees() 
        for student in students
    )
    
    return jsonify({
        'month': current_month,
        'year': current_year,
        'total_salary_expense': total_salary_expense,
        'total_outstanding_fees': total_outstanding_fees,
        'total_tutors': len(tutors),
        'total_students': len(students)
    })

# Add these routes to your app/routes/admin.py file

@bp.route('/system-documents')
@login_required
@require_permission('system_documents')
def system_documents():
    """Manage system documents"""
    from app.models.system_document import SystemDocument
    
    documents = SystemDocument.query.order_by(SystemDocument.updated_at.desc()).all()
    document_types = SystemDocument.get_document_types()
    
    return render_template('admin/system_documents.html', 
                         documents=documents,
                         document_types=document_types)

@bp.route('/system-documents/upload', methods=['GET', 'POST'])
@login_required
@admin_required
def upload_system_document():
    """Upload system document"""
    from app.models.system_document import SystemDocument
    
    if request.method == 'POST':
        try:
            document_type = request.form.get('document_type')
            title = request.form.get('title')
            description = request.form.get('description', '')
            available_roles = request.form.getlist('available_roles')
            
            if 'document_file' not in request.files:
                flash('No file selected', 'error')
                return redirect(request.url)
            
            file = request.files['document_file']
            if file.filename == '':
                flash('No file selected', 'error')
                return redirect(request.url)
            
            # Check if document type already exists
            existing_doc = SystemDocument.query.filter_by(document_type=document_type).first()
            if existing_doc:
                flash('Document type already exists. Please update the existing document.', 'error')
                return redirect(request.url)
            
            # Save file
            filename = save_uploaded_file(file, 'system_documents')
            if not filename:
                flash('Error uploading file', 'error')
                return redirect(request.url)
            
            # Create document record
            doc = SystemDocument(
                document_type=document_type,
                title=title,
                description=description,
                filename=filename,
                file_path=f'system_documents/{filename}',
                file_size=len(file.read()),
                mime_type=file.content_type,
                uploaded_by=current_user.id
            )
            
            # Reset file pointer after reading size
            file.seek(0)
            
            doc.set_available_roles(available_roles)
            
            db.session.add(doc)
            db.session.commit()
            
            flash(f'System document "{title}" uploaded successfully!', 'success')
            return redirect(url_for('admin.system_documents'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error uploading document: {str(e)}', 'error')
    
    document_types = SystemDocument.get_document_types()
    roles = ['superadmin', 'admin', 'coordinator', 'tutor', 'student']
    
    return render_template('admin/upload_system_document.html',
                         document_types=document_types,
                         roles=roles)

@bp.route('/system-documents/<int:doc_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_system_document(doc_id):
    """Edit system document"""
    from app.models.system_document import SystemDocument
    
    doc = SystemDocument.query.get_or_404(doc_id)
    
    if request.method == 'POST':
        try:
            doc.title = request.form.get('title')
            doc.description = request.form.get('description', '')
            available_roles = request.form.getlist('available_roles')
            doc.set_available_roles(available_roles)
            
            # Handle file replacement
            if 'document_file' in request.files and request.files['document_file'].filename:
                file = request.files['document_file']
                
                # Delete old file
                old_file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], doc.file_path)
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)
                
                # Save new file
                filename = save_uploaded_file(file, 'system_documents')
                if filename:
                    doc.filename = filename
                    doc.file_path = f'system_documents/{filename}'
                    doc.file_size = len(file.read())
                    doc.mime_type = file.content_type
                    file.seek(0)
            
            doc.updated_at = datetime.now()
            db.session.commit()
            
            flash(f'Document "{doc.title}" updated successfully!', 'success')
            return redirect(url_for('admin.system_documents'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating document: {str(e)}', 'error')
    
    roles = ['superadmin', 'admin', 'coordinator', 'tutor', 'student']
    return render_template('admin/edit_system_document.html', doc=doc, roles=roles)

@bp.route('/system-documents/<int:doc_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_system_document(doc_id):
    """Delete system document"""
    from app.models.system_document import SystemDocument
    
    try:
        doc = SystemDocument.query.get_or_404(doc_id)
        
        # Delete physical file
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], doc.file_path)
        if os.path.exists(file_path):
            os.remove(file_path)
        
        db.session.delete(doc)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Document deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@bp.route('/system-documents/<int:doc_id>/toggle-status', methods=['POST'])
@login_required
@admin_required
def toggle_document_status(doc_id):
    """Toggle document active status"""
    from app.models.system_document import SystemDocument
    
    try:
        doc = SystemDocument.query.get_or_404(doc_id)
        doc.is_active = not doc.is_active
        db.session.commit()
        
        status = 'activated' if doc.is_active else 'deactivated'
        return jsonify({'success': True, 'message': f'Document {status} successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# Enhanced API routes to add to your existing app/routes/admin.py
# Add these routes after your existing API routes

@bp.route('/api/v1/tutors/smart-search')
@login_required
@admin_required
def api_smart_tutor_search():
    """Advanced tutor search with multiple criteria and scoring"""
    try:
        # Get search parameters
        student_id = request.args.get('student_id', type=int)
        subject = request.args.get('subject', '').strip().lower()
        min_test_score = request.args.get('min_test_score', type=float)
        min_rating = request.args.get('min_rating', type=float)
        experience_level = request.args.get('experience_level', '').strip()
        search_term = request.args.get('search_term', '').strip().lower()
        availability_day = request.args.get('availability_day', '').strip().lower()
        availability_time = request.args.get('availability_time', '').strip()
        
        # Get base query
        tutors = Tutor.query.filter_by(status='active').all()
        
        # Get student context if provided
        student = None
        if student_id:
            student = Student.query.get(student_id)
        
        compatible_tutors = []
        
        for tutor in tutors:
            # Must have availability
            if not tutor.get_availability():
                continue
            
            # Initialize tutor score
            tutor_score = {
                'tutor_id': tutor.id,
                'tutor_name': tutor.user.full_name if tutor.user else 'Unknown',
                'email': tutor.user.email if tutor.user else '',
                'test_score': tutor.test_score or 0,
                'test_grade': tutor.get_test_score_grade(),
                'rating': tutor.rating or 0,
                'total_classes': tutor.total_classes or 0,
                'completed_classes': tutor.completed_classes or 0,
                'qualification': tutor.qualification or '',
                'subjects': tutor.get_subjects(),
                'grades': tutor.get_grades(),
                'boards': tutor.get_boards(),
                'compatibility_score': 0,
                'match_reasons': [],
                'availability_status': 'available'
            }
            
            # Student-based compatibility scoring
            if student:
                # Grade compatibility (mandatory)
                tutor_grades = [str(g) for g in tutor.get_grades()]
                if tutor_grades and str(student.grade) not in tutor_grades:
                    continue
                tutor_score['compatibility_score'] += 20
                tutor_score['match_reasons'].append(f"Teaches Grade {student.grade}")
                
                # Board compatibility (mandatory)
                tutor_boards = [b.lower() for b in tutor.get_boards()]
                if tutor_boards and student.board.lower() not in tutor_boards:
                    continue
                tutor_score['compatibility_score'] += 15
                tutor_score['match_reasons'].append(f"Familiar with {student.board}")
                
                # Subject compatibility with student's enrolled subjects
                student_subjects = [s.lower() for s in student.get_subjects_enrolled()]
                tutor_subjects = [s.lower() for s in tutor.get_subjects()]
                
                if student_subjects and tutor_subjects:
                    common_subjects = set(student_subjects) & set(tutor_subjects)
                    partial_matches = [s for s in student_subjects for ts in tutor_subjects if s in ts or ts in s]
                    
                    if common_subjects or partial_matches:
                        tutor_score['compatibility_score'] += 25
                        matched_subjects = list(common_subjects) or partial_matches[:2]
                        tutor_score['match_reasons'].append(f"Teaches: {', '.join(matched_subjects)}")
            
            # Subject filter (if specified independently)
            if subject:
                tutor_subjects = [s.lower() for s in tutor.get_subjects()]
                subject_match = any(subject in ts or ts in subject for ts in tutor_subjects)
                if not subject_match:
                    continue
                tutor_score['compatibility_score'] += 20
                tutor_score['match_reasons'].append(f"Subject expertise: {subject}")
            
            # Test score filter and scoring
            if min_test_score and (not tutor.test_score or tutor.test_score < min_test_score):
                continue
            
            if tutor.test_score:
                if tutor.test_score >= 90:
                    tutor_score['compatibility_score'] += 15
                    tutor_score['match_reasons'].append("Excellent test performance (90+)")
                elif tutor.test_score >= 80:
                    tutor_score['compatibility_score'] += 10
                    tutor_score['match_reasons'].append("Strong test performance (80+)")
                elif tutor.test_score >= 70:
                    tutor_score['compatibility_score'] += 5
                    tutor_score['match_reasons'].append("Good test performance (70+)")
            
            # Rating filter and scoring
            if min_rating and (not tutor.rating or tutor.rating < min_rating):
                continue
            
            if tutor.rating:
                if tutor.rating >= 4.5:
                    tutor_score['compatibility_score'] += 10
                    tutor_score['match_reasons'].append("Highly rated (4.5+ stars)")
                elif tutor.rating >= 4.0:
                    tutor_score['compatibility_score'] += 7
                    tutor_score['match_reasons'].append("Well rated (4.0+ stars)")
                elif tutor.rating >= 3.5:
                    tutor_score['compatibility_score'] += 5
                    tutor_score['match_reasons'].append("Good rating (3.5+ stars)")
            
            # Experience level filter
            if experience_level:
                qualification_lower = tutor.qualification.lower() if tutor.qualification else ''
                if experience_level == 'master' and 'master' not in qualification_lower:
                    continue
                elif experience_level == 'phd' and not any(term in qualification_lower for term in ['phd', 'doctorate', 'ph.d']):
                    continue
                elif experience_level == 'expert' and not any(term in qualification_lower for term in ['expert', 'specialist', 'senior']):
                    continue
                
                tutor_score['compatibility_score'] += 8
                tutor_score['match_reasons'].append(f"Advanced qualification: {experience_level}")
            
            # Search term filter (name/qualification)
            if search_term:
                tutor_name = tutor.user.full_name.lower() if tutor.user else ''
                qualification_lower = tutor.qualification.lower() if tutor.qualification else ''
                
                if search_term not in tutor_name and search_term not in qualification_lower:
                    continue
                
                tutor_score['compatibility_score'] += 5
                tutor_score['match_reasons'].append("Matches search term")
            
            # Availability check for specific day/time
            if availability_day and availability_time:
                is_available = tutor.is_available_at(availability_day, availability_time)
                if not is_available:
                    tutor_score['availability_status'] = 'busy'
                    continue
                
                tutor_score['compatibility_score'] += 10
                tutor_score['match_reasons'].append(f"Available on {availability_day.title()} at {availability_time}")
            
            # Completion rate bonus
            if tutor.total_classes > 5:  # Only consider if tutor has meaningful experience
                completion_rate = tutor.get_completion_rate()
                if completion_rate >= 95:
                    tutor_score['compatibility_score'] += 8
                    tutor_score['match_reasons'].append("Excellent completion rate (95%+)")
                elif completion_rate >= 90:
                    tutor_score['compatibility_score'] += 5
                    tutor_score['match_reasons'].append("High completion rate (90%+)")
            
            # Add comprehensive tutor info
            tutor_score['completion_rate'] = tutor.get_completion_rate()
            tutor_score['overall_score'] = tutor.calculate_overall_score() if hasattr(tutor, 'calculate_overall_score') else tutor_score['compatibility_score']
            
            compatible_tutors.append(tutor_score)
        
        # Sort by compatibility score (highest first)
        compatible_tutors.sort(key=lambda x: x['compatibility_score'], reverse=True)
        
        # Add rank information
        for i, tutor in enumerate(compatible_tutors):
            tutor['rank'] = i + 1
            if i == 0:
                tutor['badge'] = 'Best Match'
            elif tutor['compatibility_score'] >= 80:
                tutor['badge'] = 'Excellent Match'
            elif tutor['compatibility_score'] >= 60:
                tutor['badge'] = 'Good Match'
            elif tutor['compatibility_score'] >= 40:
                tutor['badge'] = 'Fair Match'
            else:
                tutor['badge'] = 'Basic Match'
        
        return jsonify({
            'success': True,
            'tutors': compatible_tutors,
            'total_found': len(compatible_tutors),
            'search_criteria': {
                'student_id': student_id,
                'subject': subject,
                'min_test_score': min_test_score,
                'min_rating': min_rating,
                'experience_level': experience_level,
                'search_term': search_term
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Error in smart tutor search: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@bp.route('/api/v1/tutors/browse-all')
@login_required
@admin_required
def api_browse_all_tutors():
    """Get all available tutors with detailed information for browsing"""
    try:
        # Get query parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        sort_by = request.args.get('sort_by', 'rating')  # rating, test_score, name, experience
        order = request.args.get('order', 'desc')  # asc, desc
        
        # Get all active tutors with availability
        tutors_query = Tutor.query.filter_by(status='active')
        
        all_tutors = []
        for tutor in tutors_query.all():
            if not tutor.get_availability():
                continue
                
            tutor_info = {
                'id': tutor.id,
                'name': tutor.user.full_name if tutor.user else 'Unknown',
                'email': tutor.user.email if tutor.user else '',
                'qualification': tutor.qualification or '',
                'experience': tutor.experience or '',
                'subjects': tutor.get_subjects(),
                'grades': tutor.get_grades(),
                'boards': tutor.get_boards(),
                'test_score': tutor.test_score or 0,
                'test_grade': tutor.get_test_score_grade(),
                'rating': tutor.rating or 0,
                'total_classes': tutor.total_classes or 0,
                'completed_classes': tutor.completed_classes or 0,
                'completion_rate': tutor.get_completion_rate(),
                'salary_type': tutor.salary_type,
                'monthly_salary': tutor.monthly_salary,
                'hourly_rate': tutor.hourly_rate,
                'created_at': tutor.created_at.isoformat() if tutor.created_at else None,
                'last_class': tutor.last_class.isoformat() if tutor.last_class else None,
                'availability_summary': get_availability_summary(tutor.get_availability())
            }
            
            all_tutors.append(tutor_info)
        
        # Sort tutors
        if sort_by == 'rating':
            all_tutors.sort(key=lambda x: x['rating'], reverse=(order == 'desc'))
        elif sort_by == 'test_score':
            all_tutors.sort(key=lambda x: x['test_score'], reverse=(order == 'desc'))
        elif sort_by == 'name':
            all_tutors.sort(key=lambda x: x['name'], reverse=(order == 'desc'))
        elif sort_by == 'experience':
            all_tutors.sort(key=lambda x: x['total_classes'], reverse=(order == 'desc'))
        
        # Paginate
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_tutors = all_tutors[start_idx:end_idx]
        
        return jsonify({
            'success': True,
            'tutors': paginated_tutors,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': len(all_tutors),
                'pages': (len(all_tutors) + per_page - 1) // per_page,
                'has_next': end_idx < len(all_tutors),
                'has_prev': page > 1
            },
            'sort_info': {
                'sort_by': sort_by,
                'order': order
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Error browsing tutors: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

def get_availability_summary(availability_dict):
    """Generate a summary of tutor availability"""
    if not availability_dict:
        return "No availability set"
    
    available_days = []
    total_hours = 0
    
    for day, slots in availability_dict.items():
        if slots:
            available_days.append(day.title())
            # Calculate total hours for this day
            for slot in slots:
                try:
                    start_time = datetime.strptime(slot['start'], '%H:%M').time()
                    end_time = datetime.strptime(slot['end'], '%H:%M').time()
                    start_datetime = datetime.combine(date.today(), start_time)
                    end_datetime = datetime.combine(date.today(), end_time)
                    hours = (end_datetime - start_datetime).total_seconds() / 3600
                    total_hours += hours
                except:
                    continue
    
    if not available_days:
        return "No availability set"
    
    return f"{len(available_days)} days/week ({total_hours:.1f}h total)"

@bp.route('/api/v1/tutors/<int:tutor_id>/detailed-profile')
@login_required
@admin_required
def api_tutor_detailed_profile(tutor_id):
    """Get comprehensive tutor profile for detailed view"""
    try:
        tutor = Tutor.query.get_or_404(tutor_id)
        
        # Get recent classes
        recent_classes = Class.query.filter_by(tutor_id=tutor_id)\
                                   .order_by(Class.scheduled_date.desc())\
                                   .limit(10).all()
        
        # Get student feedback/ratings (if you have feedback model)
        # This would need to be implemented based on your feedback system
        
        detailed_profile = {
            'basic_info': {
                'id': tutor.id,
                'name': tutor.user.full_name if tutor.user else 'Unknown',
                'email': tutor.user.email if tutor.user else '',
                'phone': tutor.user.phone if tutor.user else '',
                'qualification': tutor.qualification,
                'experience': tutor.experience,
                'date_of_birth': tutor.date_of_birth.isoformat() if tutor.date_of_birth else None,
                'state': tutor.state,
                'joining_date': tutor.created_at.isoformat() if tutor.created_at else None
            },
            'academic_info': {
                'subjects': tutor.get_subjects(),
                'grades': tutor.get_grades(),
                'boards': tutor.get_boards(),
                'test_score': tutor.test_score,
                'test_grade': tutor.get_test_score_grade(),
                'test_date': tutor.test_date.isoformat() if tutor.test_date else None,
                'test_notes': tutor.test_notes
            },
            'performance_metrics': {
                'rating': tutor.rating,
                'total_classes': tutor.total_classes,
                'completed_classes': tutor.completed_classes,
                'completion_rate': tutor.get_completion_rate(),
                'overall_score': tutor.calculate_overall_score() if hasattr(tutor, 'calculate_overall_score') else 0
            },
            'availability': {
                'schedule': tutor.get_availability(),
                'summary': get_availability_summary(tutor.get_availability())
            },
            'compensation': {
                'salary_type': tutor.salary_type,
                'monthly_salary': tutor.monthly_salary,
                'hourly_rate': tutor.hourly_rate
            },
            'recent_activity': {
                'last_class': tutor.last_class.isoformat() if tutor.last_class else None,
                'recent_classes': [
                    {
                        'id': cls.id,
                        'subject': cls.subject,
                        'date': cls.scheduled_date.isoformat(),
                        'time': cls.scheduled_time.strftime('%H:%M'),
                        'status': cls.status,
                        'student_count': len(cls.get_students()) if hasattr(cls, 'get_students') else 0
                    }
                    for cls in recent_classes
                ],
                'status': tutor.status,
                'verification_status': tutor.verification_status
            }
        }
        
        return jsonify({
            'success': True,
            'tutor': detailed_profile
        })
        
    except Exception as e:
        current_app.logger.error(f"Error getting tutor detailed profile: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Add this helper function to improve the existing compatible tutors API
@bp.route('/api/v1/compatible-tutors-enhanced')
@login_required
@admin_required
def api_compatible_tutors_enhanced():
    """Enhanced version of existing compatible tutors API with scoring"""
    try:
        student_id = request.args.get('student_id', type=int)
        subject = request.args.get('subject', '').lower()
        grade = request.args.get('grade', '')
        board = request.args.get('board', '')
        include_scores = request.args.get('include_scores', 'true').lower() == 'true'
        
        # Use the smart search function with basic parameters
        search_result = api_smart_tutor_search()
        
        # If it's a JSON response (error case), return it
        if hasattr(search_result, 'get_json'):
            return search_result
        
        # Otherwise, call the smart search internally
        from flask import current_app
        with current_app.test_request_context(
            f'/admin/api/v1/tutors/smart-search?student_id={student_id}&subject={subject}'
        ):
            smart_results = api_smart_tutor_search()
            return smart_results
            
    except Exception as e:
        current_app.logger.error(f"Error in enhanced compatible tutors: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
        

@bp.route('/api/v1/tutors/intelligent-search')
@login_required
@admin_required
@monitor_search_performance
def api_intelligent_tutor_search():
    """Most advanced tutor search with ML-like matching"""
    try:
        # Get search parameters
        student_id = request.args.get('student_id', type=int)
        subject = request.args.get('subject', '').strip()
        preferences = request.args.get('preferences', '{}')
        
        # Parse preferences
        try:
            prefs = json.loads(preferences) if preferences else {}
        except:
            prefs = {}
        
        # Build filters from preferences
        filters = {}
        if prefs.get('min_test_score'):
            filters['min_test_score'] = float(prefs['min_test_score'])
        if prefs.get('min_rating'):
            filters['min_rating'] = float(prefs['min_rating'])
        if prefs.get('experience_level'):
            filters['experience_level'] = prefs['experience_level']
        
        # Use the matching engine
        matches = matching_engine.find_best_matches(
            student_id=student_id,
            subject=subject,
            filters=filters,
            limit=prefs.get('limit', 10)
        )
        
        # Enhanced response with detailed analytics
        response_data = {
            'success': True,
            'matches': matches,
            'search_metadata': {
                'total_found': len(matches),
                'search_criteria': {
                    'student_id': student_id,
                    'subject': subject,
                    'filters_applied': list(filters.keys()),
                    'preferences': prefs
                },
                'algorithm_version': '2.0',
                'timestamp': datetime.utcnow().isoformat()
            }
        }
        
        # Add analytics if matches found
        if matches:
            scores = [m['score_data']['total_score'] for m in matches]
            response_data['search_metadata']['score_analytics'] = {
                'highest_score': max(scores),
                'average_score': round(sum(scores) / len(scores), 1),
                'score_distribution': {
                    'excellent': len([s for s in scores if s >= 85]),
                    'good': len([s for s in scores if 70 <= s < 85]),
                    'fair': len([s for s in scores if 55 <= s < 70]),
                    'basic': len([s for s in scores if s < 55])
                }
            }
        
        return jsonify(response_data)
        
    except Exception as e:
        current_app.logger.error(f"Intelligent search error: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Search failed',
            'details': str(e) if current_app.debug else 'Internal server error'
        }), 500

@bp.route('/api/v1/tutors/availability-analysis')
@login_required
@admin_required
def api_tutor_availability_analysis():
    """Analyze tutor availability for a specific date/time"""
    try:
        tutor_id = request.args.get('tutor_id', type=int)
        date_str = request.args.get('date')
        time_str = request.args.get('time')
        duration = request.args.get('duration', 60, type=int)
        
        if not all([tutor_id, date_str, time_str]):
            return jsonify({
                'success': False,
                'error': 'Missing required parameters: tutor_id, date, time'
            }), 400
        
        # Parse date and time
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        target_time = datetime.strptime(time_str, '%H:%M').time()
        
        # Check availability
        availability_result = AvailabilityChecker.check_tutor_availability(
            tutor_id, target_date, target_time, duration
        )
        
        # Get alternative slots if not available
        alternative_slots = []
        if not availability_result['available']:
            # Check next 7 days for alternatives
            for i in range(1, 8):
                alt_date = target_date + timedelta(days=i)
                slots = AvailabilityChecker.get_available_slots(
                    tutor_id, alt_date, duration
                )
                if slots:
                    alternative_slots.extend([{
                        'date': alt_date.isoformat(),
                        'day_name': alt_date.strftime('%A'),
                        'slots': slots[:3]  # First 3 slots per day
                    }])
                    if len(alternative_slots) >= 5:  # Limit to 5 alternative days
                        break
        
        # Get tutor's weekly schedule summary
        tutor = Tutor.query.get(tutor_id)
        schedule_summary = {}
        if tutor:
            availability_dict = tutor.get_availability()
            for day, slots in availability_dict.items():
                if slots:
                    schedule_summary[day] = {
                        'slot_count': len(slots),
                        'first_slot': f"{slots[0]['start']}-{slots[0]['end']}" if slots else None,
                        'total_hours': sum([
                            (datetime.strptime(slot['end'], '%H:%M') - 
                             datetime.strptime(slot['start'], '%H:%M')).total_seconds() / 3600
                            for slot in slots
                        ])
                    }
        
        return jsonify({
            'success': True,
            'availability_check': {
                'requested_slot': {
                    'date': date_str,
                    'time': time_str,
                    'duration': duration
                },
                'result': availability_result,
                'alternative_slots': alternative_slots,
                'weekly_schedule': schedule_summary
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Availability analysis error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Analysis failed'
        }), 500

@bp.route('/api/v1/students/<int:student_id>/recommended-tutors')
@login_required
@admin_required
def api_student_recommended_tutors(student_id):
    """Get personalized tutor recommendations for a student"""
    try:
        student = Student.query.get_or_404(student_id)
        
        # Get student's learning preferences and history
        academic_profile = student.get_academic_profile()
        subjects_enrolled = student.get_subjects_enrolled()
        difficult_subjects = student.get_difficult_subjects()
        
        recommendations = []
        
        # Recommend tutors for each enrolled subject
        for subject in subjects_enrolled:
            subject_matches = matching_engine.find_best_matches(
                student_id=student_id,
                subject=subject,
                limit=3
            )
            
            # Add context about why this subject needs attention
            priority = 'high' if subject.lower() in [d.lower() for d in difficult_subjects] else 'normal'
            
            if subject_matches:
                recommendations.append({
                    'subject': subject,
                    'priority': priority,
                    'reason': f"Student enrolled in {subject}" + 
                             (f" (marked as difficult)" if priority == 'high' else ""),
                    'recommended_tutors': subject_matches[:2],  # Top 2 per subject
                    'alternative_tutors': subject_matches[2:3] if len(subject_matches) > 2 else []
                })
        
        # Recommend tutors for difficult subjects (if not already covered)
        for difficult_subject in difficult_subjects:
            if difficult_subject not in subjects_enrolled:
                matches = matching_engine.find_best_matches(
                    student_id=student_id,
                    subject=difficult_subject,
                    limit=2
                )
                
                if matches:
                    recommendations.append({
                        'subject': difficult_subject,
                        'priority': 'urgent',
                        'reason': f"Student needs help with {difficult_subject} (marked as difficult)",
                        'recommended_tutors': matches,
                        'alternative_tutors': []
                    })
        
        # Sort by priority
        priority_order = {'urgent': 0, 'high': 1, 'normal': 2}
        recommendations.sort(key=lambda x: priority_order.get(x['priority'], 3))
        
        return jsonify({
            'success': True,
            'student_info': {
                'id': student.id,
                'name': student.full_name,
                'grade': student.grade,
                'board': student.board,
                'subjects_enrolled': subjects_enrolled,
                'difficult_subjects': difficult_subjects
            },
            'recommendations': recommendations,
            'summary': {
                'total_subjects': len(recommendations),
                'urgent_subjects': len([r for r in recommendations if r['priority'] == 'urgent']),
                'high_priority_subjects': len([r for r in recommendations if r['priority'] == 'high']),
                'total_tutor_options': sum([len(r['recommended_tutors']) + len(r['alternative_tutors']) 
                                           for r in recommendations])
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Student recommendations error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to generate recommendations'
        }), 500

@bp.route('/api/v1/tutors/batch-availability-check')
@login_required
@admin_required
def api_batch_availability_check():
    """Check availability for multiple tutors at once"""
    try:
        # Get parameters
        tutor_ids = request.args.getlist('tutor_ids', type=int)
        date_str = request.args.get('date')
        time_str = request.args.get('time')
        duration = request.args.get('duration', 60, type=int)
        
        if not all([tutor_ids, date_str, time_str]):
            return jsonify({
                'success': False,
                'error': 'Missing required parameters'
            }), 400
        
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        target_time = datetime.strptime(time_str, '%H:%M').time()
        
        results = []
        
        for tutor_id in tutor_ids:
            tutor = Tutor.query.get(tutor_id)
            if not tutor:
                continue
                
            availability_result = AvailabilityChecker.check_tutor_availability(
                tutor_id, target_date, target_time, duration
            )
            
            results.append({
                'tutor_id': tutor_id,
                'tutor_name': tutor.user.full_name if tutor.user else 'Unknown',
                'availability': availability_result,
                'tutor_summary': {
                    'rating': tutor.rating or 0,
                    'test_score': tutor.test_score or 0,
                    'total_classes': tutor.total_classes or 0
                }
            })
        
        # Sort by availability and then by rating
        results.sort(key=lambda x: (
            not x['availability']['available'],  # Available tutors first
            -(x['tutor_summary']['rating'] or 0)  # Then by rating (descending)
        ))
        
        available_count = len([r for r in results if r['availability']['available']])
        
        return jsonify({
            'success': True,
            'batch_check': {
                'requested_slot': {
                    'date': date_str,
                    'time': time_str,
                    'duration': duration
                },
                'results': results,
                'summary': {
                    'total_checked': len(results),
                    'available_tutors': available_count,
                    'unavailable_tutors': len(results) - available_count
                }
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Batch availability check error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Batch check failed'
        }), 500

@bp.route('/api/v1/search/suggestions')
@login_required
@admin_required
def api_search_suggestions():
    """Get search suggestions for autocomplete"""
    try:
        query = request.args.get('query', '').strip().lower()
        search_type = request.args.get('type', 'all')  # all, subjects, tutors, students
        
        suggestions = {
            'subjects': [],
            'tutors': [],
            'students': [],
            'qualifications': []
        }
        
        if len(query) < 2:
            return jsonify({'success': True, 'suggestions': suggestions})
        
        # Subject suggestions
        if search_type in ['all', 'subjects']:
            # Get unique subjects from tutors
            tutors = Tutor.query.filter_by(status='active').all()
            all_subjects = set()
            for tutor in tutors:
                all_subjects.update([s.lower() for s in tutor.get_subjects()])
            
            # Process with SearchQueryProcessor
            processed_subjects = SearchQueryProcessor.process_subject_query(query)
            matching_subjects = [s for s in all_subjects if query in s]
            matching_subjects.extend(processed_subjects)
            
            suggestions['subjects'] = list(set(matching_subjects))[:5]
        
        # Tutor suggestions
        if search_type in ['all', 'tutors']:
            tutors = Tutor.query.join(User).filter(
                Tutor.status == 'active',
                User.full_name.ilike(f'%{query}%')
            ).limit(5).all()
            
            suggestions['tutors'] = [
                {
                    'id': t.id,
                    'name': t.user.full_name,
                    'subjects': t.get_subjects()[:2],
                    'rating': t.rating or 0
                }
                for t in tutors
            ]
        
        # Student suggestions
        if search_type in ['all', 'students']:
            students = Student.query.filter(
                Student.is_active == True,
                Student.full_name.ilike(f'%{query}%')
            ).limit(5).all()
            
            suggestions['students'] = [
                {
                    'id': s.id,
                    'name': s.full_name,
                    'grade': s.grade,
                    'board': s.board
                }
                for s in students
            ]
        
        # Qualification suggestions
        if search_type in ['all', 'qualifications']:
            tutors_with_qual = Tutor.query.filter(
                Tutor.status == 'active',
                Tutor.qualification.ilike(f'%{query}%')
            ).all()
            
            qualifications = set()
            for tutor in tutors_with_qual:
                if tutor.qualification and query in tutor.qualification.lower():
                    qualifications.add(tutor.qualification)
            
            suggestions['qualifications'] = list(qualifications)[:5]
        
        return jsonify({
            'success': True,
            'query': query,
            'suggestions': suggestions
        })
        
    except Exception as e:
        current_app.logger.error(f"Search suggestions error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to get suggestions'
        }), 500

# Analytics and reporting endpoints

@bp.route('/api/v1/analytics/tutor-matching-stats')
@login_required
@admin_required
def api_tutor_matching_analytics():
    """Get analytics about tutor matching performance"""
    try:
        # Date range
        days_back = request.args.get('days', 30, type=int)
        end_date = date.today()
        start_date = end_date - timedelta(days=days_back)
        
        # Get all active tutors
        tutors = Tutor.query.filter_by(status='active').all()
        
        analytics = {
            'overview': {
                'total_active_tutors': len(tutors),
                'tutors_with_availability': len([t for t in tutors if t.get_availability()]),
                'highly_rated_tutors': len([t for t in tutors if t.rating and t.rating >= 4.5]),
                'excellent_test_scores': len([t for t in tutors if t.test_score and t.test_score >= 90])
            },
            'availability_stats': {},
            'performance_distribution': {},
            'subject_coverage': {},
            'grade_coverage': {}
        }
        
        # Availability statistics
        total_hours_available = 0
        availability_by_day = {day: 0 for day in ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']}
        
        for tutor in tutors:
            availability = tutor.get_availability()
            if availability:
                for day, slots in availability.items():
                    if slots:
                        availability_by_day[day] += 1
                        for slot in slots:
                            try:
                                start = datetime.strptime(slot['start'], '%H:%M')
                                end = datetime.strptime(slot['end'], '%H:%M')
                                hours = (end - start).total_seconds() / 3600
                                total_hours_available += hours
                            except:
                                continue
        
        analytics['availability_stats'] = {
            'total_hours_per_week': round(total_hours_available, 1),
            'average_hours_per_tutor': round(total_hours_available / len(tutors), 1) if tutors else 0,
            'availability_by_day': availability_by_day,
            'busiest_day': max(availability_by_day.items(), key=lambda x: x[1])[0] if availability_by_day else None
        }
        
        # Performance distribution
        test_score_ranges = {'90+': 0, '80-89': 0, '70-79': 0, '60-69': 0, 'below_60': 0, 'not_tested': 0}
        rating_ranges = {'4.5+': 0, '4.0-4.4': 0, '3.5-3.9': 0, '3.0-3.4': 0, 'below_3': 0, 'no_rating': 0}
        
        for tutor in tutors:
            # Test score distribution
            if tutor.test_score is None:
                test_score_ranges['not_tested'] += 1
            elif tutor.test_score >= 90:
                test_score_ranges['90+'] += 1
            elif tutor.test_score >= 80:
                test_score_ranges['80-89'] += 1
            elif tutor.test_score >= 70:
                test_score_ranges['70-79'] += 1
            elif tutor.test_score >= 60:
                test_score_ranges['60-69'] += 1
            else:
                test_score_ranges['below_60'] += 1
            
            # Rating distribution
            if tutor.rating is None:
                rating_ranges['no_rating'] += 1
            elif tutor.rating >= 4.5:
                rating_ranges['4.5+'] += 1
            elif tutor.rating >= 4.0:
                rating_ranges['4.0-4.4'] += 1
            elif tutor.rating >= 3.5:
                rating_ranges['3.5-3.9'] += 1
            elif tutor.rating >= 3.0:
                rating_ranges['3.0-3.4'] += 1
            else:
                rating_ranges['below_3'] += 1
        
        analytics['performance_distribution'] = {
            'test_scores': test_score_ranges,
            'ratings': rating_ranges
        }
        
        # Subject and grade coverage
        all_subjects = {}
        all_grades = {}
        
        for tutor in tutors:
            for subject in tutor.get_subjects():
                all_subjects[subject] = all_subjects.get(subject, 0) + 1
            for grade in tutor.get_grades():
                all_grades[str(grade)] = all_grades.get(str(grade), 0) + 1
        
        analytics['subject_coverage'] = dict(sorted(all_subjects.items(), key=lambda x: x[1], reverse=True)[:10])
        analytics['grade_coverage'] = dict(sorted(all_grades.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 999))
        
        return jsonify({
            'success': True,
            'analytics': analytics,
            'date_range': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'days_analyzed': days_back
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Analytics error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Analytics generation failed'
        }), 500
        
        


# FIXED VERSION - Replace your existing course batch routes with this

@bp.route('/course-batches')
@login_required
@admin_required
def course_batches():
    """Course batch management page - groups classes by tutor with expandable batches"""
    from collections import defaultdict
    from datetime import datetime, timedelta, date
    import time
    import json
    import urllib.parse
    from sqlalchemy.orm import selectinload

    t0 = time.perf_counter()

    query = (
        Class.query
        .options(selectinload(Class.tutor))   
        .order_by(Class.scheduled_date.desc())
    )

    tutor_filter  = request.args.get('tutor', type=int)
    status_filter = request.args.get('status', type=str)
    month_filter  = request.args.get('month', type=str)

    if tutor_filter:
        query = query.filter(Class.tutor_id == tutor_filter)
    if status_filter:
        query = query.filter(Class.status == status_filter)
    if month_filter:
        try:
            month_year = datetime.strptime(month_filter, '%Y-%m')
            start = month_year.replace(day=1)
            next_month = (start + timedelta(days=32)).replace(day=1)
            query = query.filter(
                Class.scheduled_date >= start,
                Class.scheduled_date < next_month
            )
        except ValueError:
            pass

    classes = query.all()

    # Get all students efficiently
    student_ids = set()
    for cls in classes:
        try:
            ids = cls.get_students() or []
            student_ids.update(ids)
        except Exception as e:
            print(f"Error getting students for class {cls.id}: {str(e)}")
            continue

    students_by_id = {}
    if student_ids:
        try:
            students = Student.query.filter(Student.id.in_(student_ids)).all()
            students_by_id = {s.id: s for s in students}
        except Exception as e:
            print(f"Error fetching students: {str(e)}")

    # Group by tutor first, then create individual batches within tutor
    tutors_data = defaultdict(lambda: {
        'tutor': None,
        'batches': defaultdict(lambda: {
            'classes': [],
            'students': set(),
            'tutor': None,
            'subject': '',
            'date_range': {'start': None, 'end': None},
            'total_classes': 0,
            'completed_classes': 0,
            'scheduled_classes': 0,
            'cancelled_classes': 0,
            'active_students': set(),
            'completed_students': set()
        }),
        'total_classes': 0,
        'completed_classes': 0,
        'scheduled_classes': 0,
        'cancelled_classes': 0,
        'all_students': set(),
        'active_students': set(),
        'completed_students': set(),
        'subjects': set(),
        'date_range': {'start': None, 'end': None}
    })

    for cls in classes:
        if not cls.tutor or not cls.scheduled_date:
            continue
        
        # FIXED: Better subject encoding for batch key
        try:
            # Create a safe batch key
            safe_subject = cls.subject.replace(' ', '_').replace('/', '_').replace('&', 'and')
            batch_key = f"{safe_subject}_{cls.tutor_id}"
        except Exception:
            batch_key = f"Unknown_Subject_{cls.tutor_id}"
        
        tutor_data = tutors_data[cls.tutor_id]
        tutor_data['tutor'] = cls.tutor
        tutor_data['subjects'].add(cls.subject)
        tutor_data['total_classes'] += 1
        
        # Update tutor date range
        dr = tutor_data['date_range']
        dr['start'] = min(dr['start'], cls.scheduled_date) if dr['start'] else cls.scheduled_date
        dr['end'] = max(dr['end'], cls.scheduled_date) if dr['end'] else cls.scheduled_date

        # Count by status
        if cls.status == 'completed':
            tutor_data['completed_classes'] += 1
        elif cls.status == 'scheduled':
            tutor_data['scheduled_classes'] += 1
        elif cls.status == 'cancelled':
            tutor_data['cancelled_classes'] += 1

        # Individual batch data
        batch = tutor_data['batches'][batch_key]
        batch['classes'].append(cls)
        batch['tutor'] = cls.tutor
        batch['subject'] = cls.subject
        batch['total_classes'] += 1

        if cls.status == 'completed':
            batch['completed_classes'] += 1
        elif cls.status == 'scheduled':
            batch['scheduled_classes'] += 1
        elif cls.status == 'cancelled':
            batch['cancelled_classes'] += 1

        batch_dr = batch['date_range']
        batch_dr['start'] = min(batch_dr['start'], cls.scheduled_date) if batch_dr['start'] else cls.scheduled_date
        batch_dr['end'] = max(batch_dr['end'], cls.scheduled_date) if batch_dr['end'] else cls.scheduled_date

        # FIXED: Better student handling
        try:
            ids = cls.get_students() or []
            for sid in ids:
                if not isinstance(sid, int):
                    continue
                    
                tutor_data['all_students'].add(sid)
                batch['students'].add(sid)
                
                st = students_by_id.get(sid)
                if st:
                    try:
                        if hasattr(st, "is_course_active") and callable(st.is_course_active):
                            if st.is_course_active(cls.scheduled_date):
                                tutor_data['active_students'].add(sid)
                                batch['active_students'].add(sid)
                        elif getattr(st, 'enrollment_status', None) == 'active':
                            tutor_data['active_students'].add(sid)
                            batch['active_students'].add(sid)
                            
                        if getattr(st, 'enrollment_status', None) == 'completed':
                            tutor_data['completed_students'].add(sid)
                            batch['completed_students'].add(sid)
                    except Exception as e:
                        print(f"Error checking student {sid} status: {str(e)}")
        except Exception as e:
            print(f"Error processing students for class {cls.id}: {str(e)}")

    # Convert to list for tutors
    tutor_list = []
    for tutor_id, tutor_data in tutors_data.items():
        # Process individual batches
        batch_list = []
        for key, b in tutor_data['batches'].items():
            b['batch_id'] = key
            b['student_count'] = len(b['students'])
            b['active_student_count'] = len(b['active_students'])
            b['completed_student_count'] = len(b['completed_students'])
            
            # FIXED: Safe progress calculation
            if b['total_classes'] > 0:
                b['progress_percentage'] = round((b['completed_classes'] / b['total_classes']) * 100, 1)
            else:
                b['progress_percentage'] = 0

            # Get sample student objects
            ids_slice = list(b['students'])[:5]
            b['student_objects'] = []
            for sid in ids_slice:
                if sid in students_by_id:
                    b['student_objects'].append(students_by_id[sid])
            
            batch_list.append(b)
        
        # Sort batches by date (most recent first)
        batch_list.sort(key=lambda x: x['date_range']['start'] or date.min, reverse=True)
        
        # Tutor summary
        tutor_summary = {
            'tutor_id': tutor_id,
            'tutor': tutor_data['tutor'],
            'total_batches': len(batch_list),
            'total_classes': tutor_data['total_classes'],
            'completed_classes': tutor_data['completed_classes'],
            'scheduled_classes': tutor_data['scheduled_classes'],
            'cancelled_classes': tutor_data['cancelled_classes'],
            'total_students': len(tutor_data['all_students']),
            'active_students': len(tutor_data['active_students']),
            'completed_students': len(tutor_data['completed_students']),
            'subjects': list(tutor_data['subjects']),
            'date_range': tutor_data['date_range'],
            'batches': batch_list
        }
        
        # FIXED: Safe progress calculation
        if tutor_summary['total_classes'] > 0:
            tutor_summary['progress_percentage'] = round(
                (tutor_summary['completed_classes'] / tutor_summary['total_classes']) * 100, 1
            )
        else:
            tutor_summary['progress_percentage'] = 0
        
        # Get student objects for preview
        ids_slice = list(tutor_data['active_students'])[:5]
        tutor_summary['student_objects'] = []
        for sid in ids_slice:
            if sid in students_by_id:
                tutor_summary['student_objects'].append(students_by_id[sid])
        
        tutor_list.append(tutor_summary)

    # Sort tutors by name
    tutor_list.sort(key=lambda x: x['tutor'].user.full_name if x['tutor'] and x['tutor'].user else 'ZZZ')

    # Apply activity filter
    activate_param = request.args.get('activate')  
    if activate_param in ('0', '1'):
        want_active = (activate_param == '1')
        filtered = []
        for tutor in tutor_list:
            has_any = tutor['active_students'] > 0
            if has_any is want_active:
                filtered.append(tutor)
        tutor_list = filtered

    # Get filter options
    months = sorted(
        {cls.scheduled_date.strftime('%Y-%m') for cls in classes if cls.scheduled_date},
        reverse=True
    )
    tutors = Tutor.query.filter_by(status='active').all()
    
    # Simple pagination
    from math import ceil

    class SimplePagination:
        def __init__(self, page, per_page, total):
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = ceil(total / per_page) if per_page > 0 else 0
            self.has_prev = page > 1
            self.has_next = page < self.pages
            self.prev_num = page - 1 if self.has_prev else None
            self.next_num = page + 1 if self.has_next else None
            
        def iter_pages(self, left_edge=1, right_edge=1, left_current=1, right_current=2):
            max_shown = 10
            for num in range(1, min(self.pages, max_shown) + 1):
                yield num
            if self.pages > max_shown:
                yield None
                yield self.pages

    page = request.args.get('page', 1, type=int)
    per_page = 8
    total = len(tutor_list)

    start = (page - 1) * per_page
    end = start + per_page
    page_items = tutor_list[start:end]

    filtered_args = request.args.to_dict()
    filtered_args.pop('page', None)

    pagination = SimplePagination(page, per_page, total)

    print(f"Course batches loaded in {time.perf_counter() - t0:.3f}s")

    return render_template(
        'admin/course_batches.html',
        tutors_data=page_items,
        pagination=pagination,
        filtered_args=filtered_args,
        tutors=tutors,
        months=months,
        total_tutors=total
    )


@bp.route('/course-batches/<batch_id>')
@login_required
@admin_required
def course_batch_details(batch_id):
    """View detailed information about a specific course batch."""
    from datetime import datetime, date
    from sqlalchemy.orm import selectinload
    import traceback

    try:
        # FIXED: Better batch_id parsing
        parts = batch_id.split('_')
        if len(parts) < 2:
            raise ValueError(f"Invalid batch_id format: {batch_id}")

        # Last part is tutor_id, everything else is subject
        tutor_id = int(parts[-1])
        subject_parts = parts[:-1]
        
        # Reconstruct subject name (reverse the encoding)
        subject = '_'.join(subject_parts).replace('_', ' ').replace('and', '&')

        print(f"Parsed batch_id '{batch_id}' -> subject: '{subject}', tutor_id: {tutor_id}")

    except Exception as e:
        print(f"Error parsing batch_id '{batch_id}': {str(e)}")
        traceback.print_exc()
        flash('Invalid batch ID format', 'error')
        return redirect(url_for('admin.course_batches'))

    # FIXED: Query classes with better subject matching
    classes = (
        Class.query
        .options(selectinload(Class.tutor))
        .filter(
            Class.subject == subject,
            Class.tutor_id == tutor_id
        )
        .order_by(Class.scheduled_date, Class.scheduled_time)
        .all()
    )

    print(f"Found {len(classes)} classes for subject '{subject}' and tutor {tutor_id}")

    if not classes:
        # Debug: Try to find similar subjects
        debug_classes = (
            Class.query
            .filter(Class.tutor_id == tutor_id)
            .all()
        )
        
        if debug_classes:
            subjects_found = list(set(cls.subject for cls in debug_classes))
            print(f"Available subjects for tutor {tutor_id}: {subjects_found}")
            flash(f'No classes found for subject "{subject}". Available: {", ".join(subjects_found)}', 'warning')
        else:
            print(f"No classes found for tutor {tutor_id}")
            flash(f'No classes found for this tutor', 'error')
        
        return redirect(url_for('admin.course_batches'))

    # FIXED: Better student collection
    all_student_ids = set()
    for c in classes:
        try:
            ids = c.get_students() or []
            # Ensure all IDs are integers
            valid_ids = [int(sid) for sid in ids if isinstance(sid, (int, str)) and str(sid).isdigit()]
            all_student_ids.update(valid_ids)
        except Exception as e:
            print(f"Error getting students for class {c.id}: {str(e)}")

    students_by_id = {}
    students = []
    if all_student_ids:
        try:
            students = Student.query.filter(Student.id.in_(all_student_ids)).all()
            students_by_id = {s.id: s for s in students}
            print(f"Found {len(students)} students for batch")
        except Exception as e:
            print(f"Error fetching students: {str(e)}")

    # Calculate statistics
    total_classes = len(classes)
    completed_classes = sum(1 for c in classes if c.status == 'completed')
    scheduled_classes = sum(1 for c in classes if c.status == 'scheduled')
    cancelled_classes = sum(1 for c in classes if c.status == 'cancelled')

    total_students = len(students)
    active_students = sum(1 for s in students if getattr(s, 'enrollment_status', '') == 'active')
    completed_students = sum(1 for s in students if getattr(s, 'enrollment_status', '') == 'completed')

    stats = {
        'total_classes': total_classes,
        'completed_classes': completed_classes,
        'scheduled_classes': scheduled_classes,
        'cancelled_classes': cancelled_classes,
        'total_students': total_students,
        'active_students': active_students,
        'completed_students': completed_students,
        'progress_percentage': round((completed_classes / total_classes) * 100, 1) if total_classes > 0 else 0
    }

    # Get tutor info
    tutor = classes[0].tutor if classes else Tutor.query.get(tutor_id)
    
    # Get available tutors for change functionality
    available_tutors = Tutor.query.filter_by(status='active').all()
    available_tutors = [t for t in available_tutors if hasattr(t, 'get_availability') and t.get_availability()]

    print(f"Batch details loaded successfully")

    return render_template(
        'admin/course_batch_details.html',
        classes=classes,
        students=students,
        tutor=tutor,
        available_tutors=available_tutors,
        batch_id=batch_id,
        subject=subject,
        stats=stats
    )

    
# ============ BATCH MANAGEMENT API ROUTES ============

@bp.route('/api/batch/<batch_id>/change-tutor', methods=['POST'])
@login_required
@admin_required
def api_batch_change_tutor(batch_id):
    """Change tutor for all classes in a batch"""
    try:
        new_tutor_id = request.json.get('new_tutor_id')
        if not new_tutor_id:
            return jsonify({'error': 'New tutor ID is required'}), 400

        # Parse batch_id to get class filters - UPDATED for new format
        parts = batch_id.split('_')
        if len(parts) < 2:
            return jsonify({'error': 'Invalid batch ID format'}), 400
            
        raw_tutor_id = parts[-1]
        raw_subject = '_'.join(parts[:-1])
        
        subject = urllib.parse.unquote(raw_subject)
        current_tutor_id = int(raw_tutor_id)

        # Get new tutor and verify availability
        new_tutor = Tutor.query.get_or_404(new_tutor_id)
        if not new_tutor.get_availability():
            return jsonify({'error': 'Selected tutor has not set their availability'}), 400

        # Get all classes in the batch - REMOVED date filtering
        classes = Class.query.filter(
            Class.subject == subject,
            Class.tutor_id == current_tutor_id,
            Class.status.in_(['scheduled'])  # Only change scheduled classes
        ).all()

        if not classes:
            return jsonify({'error': 'No scheduled classes found in this batch'}), 404

        # Check availability for each class
        conflicts = []
        successful_changes = 0
        
        for cls in classes:
            day_of_week = cls.scheduled_date.strftime('%A').lower()
            time_str = cls.scheduled_time.strftime('%H:%M')
            
            # Check if new tutor is available
            if not new_tutor.is_available_at(day_of_week, time_str):
                conflicts.append({
                    'class_id': cls.id,
                    'date': cls.scheduled_date.strftime('%Y-%m-%d'),
                    'time': time_str,
                    'reason': 'Tutor not available'
                })
                continue
            
            # Check for scheduling conflicts
            existing_class = Class.query.filter_by(
                tutor_id=new_tutor_id,
                scheduled_date=cls.scheduled_date,
                scheduled_time=cls.scheduled_time,
                status='scheduled'
            ).first()
            
            if existing_class:
                conflicts.append({
                    'class_id': cls.id,
                    'date': cls.scheduled_date.strftime('%Y-%m-%d'),
                    'time': time_str,
                    'reason': 'Tutor already has a class at this time'
                })
                continue
            
            # Change tutor
            cls.tutor_id = new_tutor_id
            cls.updated_at = datetime.utcnow()
            successful_changes += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Successfully changed tutor for {successful_changes} classes',
            'successful_changes': successful_changes,
            'conflicts': conflicts,
            'total_classes': len(classes)
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error changing tutor: {str(e)}'}), 500

@bp.route('/api/batch/<batch_id>/delete-classes', methods=['POST'])
@login_required
@admin_required
def api_batch_delete_classes(batch_id):
    """Delete selected classes from a batch"""
    try:
        class_ids = request.json.get('class_ids', [])
        if not class_ids:
            return jsonify({'error': 'No classes selected for deletion'}), 400

        # Verify classes belong to the batch
        classes = Class.query.filter(Class.id.in_(class_ids)).all()
        
        if not classes:
            return jsonify({'error': 'No valid classes found'}), 404

        deleted_count = 0
        for cls in classes:
            # Only delete scheduled classes
            if cls.status == 'scheduled':
                cls.status = 'cancelled'
                cls.cancellation_reason = 'Deleted via batch management'
                cls.cancelled_at = datetime.utcnow()
                cls.cancelled_by = current_user.id
                deleted_count += 1

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Successfully cancelled {deleted_count} classes',
            'deleted_count': deleted_count
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error deleting classes: {str(e)}'}), 500

@bp.route('/api/tutor/<int:tutor_id>/batch-availability', methods=['GET'])
@login_required
@admin_required
def api_tutor_batch_availability(tutor_id):
    """Check tutor availability for a batch of classes"""
    try:
        class_ids = request.args.getlist('class_ids')
        if not class_ids:
            return jsonify({'error': 'No class IDs provided'}), 400

        tutor = Tutor.query.get_or_404(tutor_id)
        availability = tutor.get_availability()
        
        if not availability:
            return jsonify({
                'success': False,
                'has_availability': False,
                'message': 'Tutor has not set their availability'
            })

        # Get classes and check availability
        classes = Class.query.filter(Class.id.in_(class_ids)).all()
        availability_results = []
        
        for cls in classes:
            day_of_week = cls.scheduled_date.strftime('%A').lower()
            time_str = cls.scheduled_time.strftime('%H:%M')
            
            is_available = tutor.is_available_at(day_of_week, time_str)
            
            # Check for conflicts
            existing_class = Class.query.filter_by(
                tutor_id=tutor_id,
                scheduled_date=cls.scheduled_date,
                scheduled_time=cls.scheduled_time,
                status='scheduled'
            ).first()
            
            availability_results.append({
                'class_id': cls.id,
                'date': cls.scheduled_date.strftime('%Y-%m-%d'),
                'time': time_str,
                'day': day_of_week.title(),
                'available': is_available and not existing_class,
                'reason': 'Available' if is_available and not existing_class 
                         else 'Not available' if not is_available 
                         else 'Scheduling conflict'
            })

        available_count = sum(1 for r in availability_results if r['available'])
        
        return jsonify({
            'success': True,
            'has_availability': bool(availability),
            'tutor_name': tutor.user.full_name,
            'total_classes': len(classes),
            'available_classes': available_count,
            'conflicts': len(classes) - available_count,
            'details': availability_results
        })

    except Exception as e:
        return jsonify({'error': f'Error checking availability: {str(e)}'}), 500
    
    
@bp.route('/api/v1/check-class-conflict')
@login_required
@admin_required
def api_check_class_conflict():
    """Enhanced conflict checking that includes reschedule exclusions"""
    try:
        tutor_id = request.args.get('tutor_id', type=int)
        date_str = request.args.get('date')
        time_str = request.args.get('time')
        duration = request.args.get('duration', type=int)
        exclude_class = request.args.get('exclude_class', type=int)  # For reschedule checks
        
        if not all([tutor_id, date_str, time_str, duration]):
            return jsonify({
                'success': False,
                'error': 'Missing required parameters'
            }), 400
        
        # Parse date and time
        from datetime import datetime, date, time
        schedule_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        schedule_time = datetime.strptime(time_str, '%H:%M').time()
        
        # Check for time conflicts
        conflict_exists, conflicting_class = Class.check_time_conflict(
            tutor_id, schedule_date, schedule_time, duration, exclude_class
        )
        
        tutor = Tutor.query.get(tutor_id)
        if not tutor:
            return jsonify({
                'success': False,
                'error': 'Tutor not found'
            }), 404
        
        # Check tutor availability
        day_name = schedule_date.strftime('%A').lower()
        availability_check = True
        
        if hasattr(tutor, 'is_available_at'):
            availability_check = tutor.is_available_at(day_name, time_str)
        
        can_schedule = not conflict_exists and availability_check
        
        message = "Available"
        if conflict_exists:
            message = f"Tutor has another class at {conflicting_class.scheduled_time.strftime('%H:%M')}"
        elif not availability_check:
            message = f"Tutor is not available on {day_name.title()} at {time_str}"
        
        return jsonify({
            'success': True,
            'can_schedule': can_schedule,
            'has_conflict': conflict_exists,
            'tutor_available': availability_check,
            'message': message,
            'conflicting_class_id': conflicting_class.id if conflicting_class else None
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
        

@bp.route('/api/v1/classes/<int:class_id>/reschedule', methods=['POST'])
@login_required
@admin_required
def api_reschedule_class(class_id):
    """API endpoint for rescheduling classes"""
    try:
        class_item = Class.query.get_or_404(class_id)
        
        # Check department access for coordinators
        if current_user.role == 'coordinator':
            if class_item.tutor.user.department_id != current_user.department_id:
                return jsonify({'success': False, 'error': 'Access denied'}), 403
        
        data = request.get_json()
        new_date_str = data.get('scheduled_date')
        new_time_str = data.get('scheduled_time')
        reason = data.get('reschedule_reason', 'Rescheduled by admin')
        
        if not new_date_str or not new_time_str:
            return jsonify({'success': False, 'error': 'Date and time are required'}), 400
        
        # Parse date and time
        from datetime import datetime
        new_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
        new_time = datetime.strptime(new_time_str, '%H:%M').time()
        
        # Check for conflicts
        conflict_exists, conflicting_class = Class.check_time_conflict(
            class_item.tutor_id, new_date, new_time, class_item.duration, class_id
        )
        
        if conflict_exists:
            return jsonify({
                'success': False,
                'error': f'Time conflict with existing class at {conflicting_class.scheduled_time}'
            })
        
        # Update class
        old_date = class_item.scheduled_date
        old_time = class_item.scheduled_time
        
        class_item.scheduled_date = new_date
        class_item.scheduled_time = new_time
        class_item.calculate_end_time()
        class_item.updated_at = datetime.utcnow()
        
        # Add admin note
        note = f"Rescheduled by {current_user.full_name} from {old_date} {old_time} to {new_date} {new_time}. Reason: {reason}"
        if class_item.admin_notes:
            class_item.admin_notes += f"\n{note}"
        else:
            class_item.admin_notes = note
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Class rescheduled successfully',
            'new_date': new_date.strftime('%Y-%m-%d'),
            'new_time': new_time.strftime('%H:%M')
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@bp.route('/api/v1/classes/<int:class_id>/cancel', methods=['POST'])
@login_required
@admin_required
def api_cancel_class(class_id):
    """API endpoint for cancelling classes"""
    try:
        class_item = Class.query.get_or_404(class_id)
        
        # Check department access for coordinators
        if current_user.role == 'coordinator':
            if class_item.tutor.user.department_id != current_user.department_id:
                return jsonify({'success': False, 'error': 'Access denied'}), 403
        
        data = request.get_json() or {}
        reason = data.get('reason', 'Cancelled by admin')
        
        # Cancel the class
        class_item.cancel_class(reason)
        
        return jsonify({
            'success': True,
            'message': 'Class cancelled successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ADD THIS ROUTE for dashboard statistics
@bp.route('/api/v1/dashboard/reschedule-stats')
@login_required
@require_permission('schedule_management')
def api_reschedule_stats():
    """Get reschedule request statistics for dashboard"""
    try:
        from app.models.reschedule_request import RescheduleRequest
        
        # Base query
        query = RescheduleRequest.query
        
        # Filter by department for coordinators
        if current_user.role == 'coordinator':
            from app.models.tutor import Tutor
            from app.models.user import User
            query = query.join(Class).join(Tutor).join(User).filter(
                User.department_id == current_user.department_id
            )
        
        # Get counts
        total_requests = query.count()
        pending_requests = query.filter_by(status='pending').count()
        approved_requests = query.filter_by(status='approved').count()
        rejected_requests = query.filter_by(status='rejected').count()
        
        # Get requests with conflicts
        conflict_requests = query.filter_by(has_conflicts=True, status='pending').count()
        
        return jsonify({
            'success': True,
            'stats': {
                'total_requests': total_requests,
                'pending_requests': pending_requests,
                'approved_requests': approved_requests,
                'rejected_requests': rejected_requests,
                'conflict_requests': conflict_requests
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
    
# ADD THESE ROUTES TO app/routes/admin.py

@bp.route('/api/v1/tutors/active')
@login_required
@admin_required
def api_get_active_tutors():
    """Get active tutors for bulk edit and other operations"""
    try:
        tutors = Tutor.query.join(Tutor.user).filter(
            Tutor.status == 'active',
            Tutor.user.has(is_active=True)
        ).all()
        
        tutor_data = []
        for tutor in tutors:
            tutor_data.append({
                'id': tutor.id,
                'name': tutor.user.full_name,
                'subjects': tutor.get_subjects(),
                'grades': tutor.get_grades(),
                'status': tutor.status,
                'department': tutor.user.department.name if tutor.user.department else None
            })
        
        return jsonify({
            'success': True,
            'tutors': tutor_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@bp.route('/classes/<int:class_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_class(class_id):
    """Edit single class"""
    from app.forms.class_forms import EditClassForm
    import json
    
    class_obj = Class.query.get_or_404(class_id)
    
    # Check if class can be edited
    if not class_obj.is_editable():
        flash('This class cannot be edited. Classes that are completed, ongoing, or start within 1 hour cannot be modified.', 'error')
        return redirect(url_for('admin.class_details', class_id=class_id))
    
    # Check department access for coordinators
    if current_user.role == 'coordinator':
        if class_obj.tutor and class_obj.tutor.user.department_id != current_user.department_id:
            flash('Access denied. You can only edit classes from your department.', 'error')
            return redirect(url_for('admin.classes'))
    
    form = EditClassForm(class_obj=class_obj, department_id=current_user.department_id if current_user.role == 'coordinator' else None)
    
    if form.validate_on_submit():
        try:
            # Store original values for change tracking
            original_tutor_id = class_obj.tutor_id
            original_date = class_obj.scheduled_date
            original_time = class_obj.scheduled_time
            
            # Update basic information
            class_obj.subject = form.subject.data
            class_obj.class_type = form.class_type.data
            class_obj.grade = form.grade.data
            class_obj.board = form.board.data
            
            # Update scheduling
            class_obj.scheduled_date = form.scheduled_date.data
            class_obj.scheduled_time = form.scheduled_time.data
            class_obj.duration = form.duration.data
            
            # Calculate end time
            from datetime import datetime, timedelta
            start_datetime = datetime.combine(class_obj.scheduled_date, class_obj.scheduled_time)
            end_datetime = start_datetime + timedelta(minutes=class_obj.duration)
            class_obj.end_time = end_datetime.time()
            
            # Update tutor assignment
            class_obj.tutor_id = form.tutor_id.data
            
            # Update student assignments
            if form.class_type.data == 'one_on_one':
                class_obj.primary_student_id = form.primary_student_id.data
                class_obj.students = None  # Clear group students
            elif form.class_type.data == 'group':
                if form.students.data:
                    class_obj.students = json.dumps(form.students.data)
                if form.primary_student_id.data and form.primary_student_id.data != 0:
                    class_obj.primary_student_id = form.primary_student_id.data
                else:
                    class_obj.primary_student_id = None
            else:  # demo
                class_obj.primary_student_id = form.primary_student_id.data
                class_obj.students = None
            
            # Update platform and links
            class_obj.platform = form.platform.data
            class_obj.meeting_link = form.meeting_link.data
            class_obj.meeting_id = form.meeting_id.data
            class_obj.meeting_password = form.meeting_password.data
            
            # Update notes
            class_obj.class_notes = form.class_notes.data
            
            # Update timestamp
            class_obj.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            # Create change notification message
            changes = []
            if original_tutor_id != class_obj.tutor_id:
                old_tutor = Tutor.query.get(original_tutor_id).user.full_name if original_tutor_id else 'None'
                new_tutor = Tutor.query.get(class_obj.tutor_id).user.full_name
                changes.append(f"Tutor: {old_tutor} → {new_tutor}")
            
            if original_date != class_obj.scheduled_date or original_time != class_obj.scheduled_time:
                old_schedule = f"{original_date} at {original_time}"
                new_schedule = f"{class_obj.scheduled_date} at {class_obj.scheduled_time}"
                changes.append(f"Schedule: {old_schedule} → {new_schedule}")
            
            if changes:
                flash(f'Class updated successfully! Changes: {", ".join(changes)}', 'success')
            else:
                flash('Class updated successfully!', 'success')
            
            return redirect(url_for('admin.class_details', class_id=class_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating class: {str(e)}', 'error')
    
    return render_template('admin/edit_class.html', form=form, class_obj=class_obj)


@bp.route('/classes/bulk-edit', methods=['GET', 'POST'])
@login_required
@admin_required
def bulk_edit_classes():
    """Bulk edit multiple classes"""
    from app.forms.class_forms import BulkEditClassForm
    import json
    from datetime import datetime, timedelta
    
    form = BulkEditClassForm(department_id=current_user.department_id if current_user.role == 'coordinator' else None)
    
    if form.validate_on_submit():
        try:
            # Parse selected classes
            class_ids = json.loads(form.selected_classes.data)
            classes = Class.query.filter(Class.id.in_(class_ids)).all()
            
            # Filter editable classes
            editable_classes = [cls for cls in classes if cls.is_editable()]
            skipped_count = len(classes) - len(editable_classes)
            
            if not editable_classes:
                flash('No classes could be edited. All selected classes are either completed, ongoing, or start within 1 hour.', 'warning')
                return redirect(url_for('admin.classes'))
            
            updated_count = 0
            error_count = 0
            
            for class_obj in editable_classes:
                try:
                    changes_made = False
                    
                    # Update tutor if specified
                    if form.tutor_id.data and form.tutor_id.data != 0:
                        # Check for conflicts before updating
                        conflict_exists, _ = Class.check_time_conflict(
                            form.tutor_id.data,
                            class_obj.scheduled_date,
                            class_obj.scheduled_time,
                            class_obj.duration,
                            class_obj.id
                        )
                        
                        if not conflict_exists:
                            class_obj.tutor_id = form.tutor_id.data
                            changes_made = True
                    
                    # Update platform if specified
                    if form.platform.data:
                        class_obj.platform = form.platform.data
                        changes_made = True
                    
                    # Apply time adjustments
                    if form.time_adjustment.data:
                        adjustment_map = {
                            'add_15': 15,
                            'subtract_15': -15,
                            'add_30': 30,
                            'subtract_30': -30
                        }
                        
                        if form.time_adjustment.data in adjustment_map:
                            minutes_delta = adjustment_map[form.time_adjustment.data]
                            current_datetime = datetime.combine(class_obj.scheduled_date, class_obj.scheduled_time)
                            new_datetime = current_datetime + timedelta(minutes=minutes_delta)
                            
                            # Ensure new time is not in the past
                            if new_datetime.date() >= datetime.now().date():
                                class_obj.scheduled_time = new_datetime.time()
                                # Update end time
                                end_datetime = new_datetime + timedelta(minutes=class_obj.duration)
                                class_obj.end_time = end_datetime.time()
                                changes_made = True
                    
                    # Update duration if specified
                    if form.new_duration.data:
                        class_obj.duration = form.new_duration.data
                        # Recalculate end time
                        start_datetime = datetime.combine(class_obj.scheduled_date, class_obj.scheduled_time)
                        end_datetime = start_datetime + timedelta(minutes=class_obj.duration)
                        class_obj.end_time = end_datetime.time()
                        changes_made = True
                    
                    # Update meeting link if specified
                    if form.meeting_link.data:
                        class_obj.meeting_link = form.meeting_link.data
                        changes_made = True
                    
                    # Append bulk notes if specified
                    if form.bulk_notes.data:
                        if class_obj.class_notes:
                            class_obj.class_notes += f"\n\n{form.bulk_notes.data}"
                        else:
                            class_obj.class_notes = form.bulk_notes.data
                        changes_made = True
                    
                    if changes_made:
                        class_obj.updated_at = datetime.utcnow()
                        updated_count += 1
                
                except Exception as e:
                    error_count += 1
                    print(f"Error updating class {class_obj.id}: {str(e)}")
                    continue
            
            db.session.commit()
            
            # Create success message
            message = f'Bulk edit completed! Updated {updated_count} classes.'
            if skipped_count > 0:
                message += f' {skipped_count} classes were skipped (not editable).'
            if error_count > 0:
                message += f' {error_count} classes had errors.'
            
            flash(message, 'success' if error_count == 0 else 'warning')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error during bulk edit: {str(e)}', 'error')
    
    return redirect(url_for('admin.classes'))


@bp.route('/api/v1/classes/<int:class_id>/editable')
@login_required
@admin_required
def api_check_class_editable(class_id):
    """Check if a class can be edited via API"""
    try:
        class_obj = Class.query.get_or_404(class_id)
        
        return jsonify({
            'success': True,
            'editable': class_obj.is_editable(),
            'reason': None if class_obj.is_editable() else 'Class is completed, ongoing, or starts within 1 hour'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/api/v1/classes/<int:class_id>/quick-edit', methods=['POST'])
@login_required
@admin_required
def api_quick_edit_class(class_id):
    """Quick edit class via API (for simple changes)"""
    try:
        class_obj = Class.query.get_or_404(class_id)
        
        if not class_obj.is_editable():
            return jsonify({
                'success': False,
                'error': 'Class cannot be edited'
            }), 400
        
        data = request.get_json()
        
        # Handle quick edits
        if 'tutor_id' in data:
            # Check for conflicts
            conflict_exists, conflicting_class = Class.check_time_conflict(
                data['tutor_id'],
                class_obj.scheduled_date,
                class_obj.scheduled_time,
                class_obj.duration,
                class_obj.id
            )
            
            if conflict_exists:
                return jsonify({
                    'success': False,
                    'error': f'Tutor has a conflicting class: {conflicting_class.subject}'
                }), 400
            
            class_obj.tutor_id = data['tutor_id']
        
        if 'platform' in data:
            class_obj.platform = data['platform']
        
        if 'meeting_link' in data:
            class_obj.meeting_link = data['meeting_link']
        
        if 'class_notes' in data:
            class_obj.class_notes = data['class_notes']
        
        class_obj.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Class updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/classes/<int:class_id>/duplicate')
@login_required
@admin_required
def duplicate_class(class_id):
    """Create a duplicate of an existing class"""
    try:
        original_class = Class.query.get_or_404(class_id)
        
        # Create new class with same data
        new_class = Class(
            subject=original_class.subject,
            class_type=original_class.class_type,
            grade=original_class.grade,
            board=original_class.board,
            scheduled_date=original_class.scheduled_date,
            scheduled_time=original_class.scheduled_time,
            duration=original_class.duration,
            tutor_id=original_class.tutor_id,
            primary_student_id=original_class.primary_student_id,
            students=original_class.students,
            platform=original_class.platform,
            meeting_link=original_class.meeting_link,
            meeting_id=original_class.meeting_id,
            meeting_password=original_class.meeting_password,
            class_notes=f"Duplicated from class #{original_class.id}",
            status='scheduled',
            created_by=current_user.id
        )
        
        # Calculate end time
        from datetime import datetime, timedelta
        start_datetime = datetime.combine(new_class.scheduled_date, new_class.scheduled_time)
        end_datetime = start_datetime + timedelta(minutes=new_class.duration)
        new_class.end_time = end_datetime.time()
        
        db.session.add(new_class)
        db.session.commit()
        
        flash(f'Class duplicated successfully! New class ID: {new_class.id}', 'success')
        return redirect(url_for('admin.edit_class', class_id=new_class.id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error duplicating class: {str(e)}', 'error')
        return redirect(url_for('admin.class_details', class_id=class_id))


@bp.route('/classes/<int:class_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_class(class_id):
    """Delete a class"""
    try:
        class_obj = Class.query.get_or_404(class_id)
        
        if not class_obj.is_deletable():
            return jsonify({
                'success': False,
                'error': 'Class cannot be deleted. Completed classes or classes that have started cannot be removed.'
            }), 400
        
        # Check permissions
        if current_user.role == 'coordinator':
            if class_obj.tutor and class_obj.tutor.user.department_id != current_user.department_id:
                return jsonify({
                    'success': False,
                    'error': 'Access denied'
                }), 403
        
        # Delete related attendance records first
        Attendance.query.filter_by(class_id=class_id).delete()
        
        # Delete the class
        db.session.delete(class_obj)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Class deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
        
        
### ADDITIONAL ROUTES FOR ALLOCATION HELPERS

@bp.route('/allocation-dashboard')
@login_required
@require_permission('tutor_management')
def allocation_dashboard():
    """Smart Allocation Dashboard - Main View"""
    
    # Get analytics for overview
    analytics = allocation_helper.get_allocation_analytics()
    
    # Get recent unallocated students (last 20)
    unallocated_students = allocation_helper.get_unallocated_students()[:20]
    
    # Get available tutors (top 15 by rating)
    available_tutors = allocation_helper.get_available_tutors()[:15]
    
    # Get departments for filtering
    departments = Department.query.filter_by(is_active=True).all()
    
    # Get unique subjects and grades for filters
    all_subjects = set()
    all_grades = set()
    all_boards = set()
    
    students = Student.query.filter_by(is_active=True).all()
    for student in students:
        try:
            subjects = student.get_subjects_enrolled()
            all_subjects.update(subjects)
            all_grades.add(student.grade)
            all_boards.add(student.board)
        except:
            continue
    
    filter_options = {
        'subjects': sorted(list(all_subjects)),
        'grades': sorted(list(all_grades), key=lambda x: int(x) if x.isdigit() else 999),
        'boards': sorted(list(all_boards))
    }
    
    return render_template('admin/allocation_dashboard.html',
                         analytics=analytics,
                         unallocated_students=unallocated_students,
                         available_tutors=available_tutors,
                         departments=departments,
                         filter_options=filter_options)


@bp.route('/api/allocation/unallocated-students')
@login_required
@admin_required
def api_unallocated_students():
    """API: Get unallocated students with filters"""
    try:
        # Get filters from request
        filters = {}
        if request.args.get('grade'):
            filters['grade'] = request.args.get('grade')
        if request.args.get('board'):
            filters['board'] = request.args.get('board')
        if request.args.get('subject'):
            filters['subject'] = request.args.get('subject')
        if request.args.get('department_id'):
            filters['department_id'] = int(request.args.get('department_id'))
        if request.args.get('days_unallocated'):
            filters['days_unallocated'] = int(request.args.get('days_unallocated'))
        
        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        # Get unallocated students
        all_unallocated = allocation_helper.get_unallocated_students(filters)
        
        # Apply pagination
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paginated_students = all_unallocated[start_idx:end_idx]
        
        return jsonify({
            'success': True,
            'students': paginated_students,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': len(all_unallocated),
                'pages': (len(all_unallocated) + per_page - 1) // per_page,
                'has_next': end_idx < len(all_unallocated),
                'has_prev': page > 1
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Unallocated students API error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to fetch unallocated students'
        }), 500


@bp.route('/api/allocation/available-tutors')
@login_required
@admin_required
def api_available_tutors():
    """API: Get available tutors with filters"""
    try:
        # Get filters from request
        filters = {}
        if request.args.get('subject'):
            filters['subject'] = request.args.get('subject')
        if request.args.get('grade'):
            filters['grade'] = request.args.get('grade')
        if request.args.get('board'):
            filters['board'] = request.args.get('board')
        if request.args.get('min_rating'):
            filters['min_rating'] = float(request.args.get('min_rating'))
        if request.args.get('min_test_score'):
            filters['min_test_score'] = float(request.args.get('min_test_score'))
        
        available_tutors = allocation_helper.get_available_tutors(filters)
        
        return jsonify({
            'success': True,
            'tutors': available_tutors
        })
        
    except Exception as e:
        current_app.logger.error(f"Available tutors API error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to fetch available tutors'
        }), 500


@bp.route('/api/allocation/smart-match/<int:student_id>')
@login_required
@admin_required
def api_smart_match(student_id):
    """API: Get smart tutor matches for specific student"""
    try:
        limit = request.args.get('limit', 3, type=int)
        matches = allocation_helper.get_smart_matches(student_id, limit)
        
        # Get student info
        student = Student.query.get_or_404(student_id)
        student_info = {
            'id': student.id,
            'full_name': student.full_name,
            'grade': student.grade,
            'board': student.board,
            'subjects_enrolled': student.get_subjects_enrolled()
        }
        
        return jsonify({
            'success': True,
            'student': student_info,
            'matches': matches
        })
        
    except Exception as e:
        current_app.logger.error(f"Smart match API error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to generate matches'
        }), 500


@bp.route('/api/allocation/quick-assign', methods=['POST'])
@login_required
@admin_required
def api_quick_assign():
    """API: Quick assign student to tutor"""
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        tutor_id = data.get('tutor_id')
        subject = data.get('subject', 'General')
        
        if not all([student_id, tutor_id]):
            return jsonify({
                'success': False,
                'error': 'Student ID and Tutor ID are required'
            }), 400
        
        # Verify student and tutor exist
        student = Student.query.get_or_404(student_id)
        tutor = Tutor.query.get_or_404(tutor_id)
        
        # Check if tutor has availability
        if not tutor.get_availability():
            return jsonify({
                'success': False,
                'error': 'Selected tutor has not set their availability'
            }), 400
        
        # Check if student is already allocated
        existing_class = Class.query.filter(
            or_(
                Class.primary_student_id == student_id,
                Class.students.like(f'%{student_id}%')
            ),
            Class.status.in_(['scheduled', 'ongoing'])
        ).first()
        
        if existing_class:
            return jsonify({
                'success': False,
                'error': 'Student is already allocated to a class'
            }), 400
        
        # Create new class assignment
        # Note: You'll need to set scheduled_date and scheduled_time based on availability
        # This is a simplified version - you may want to enhance this
        new_class = Class(
            subject=subject,
            class_type='one_on_one',
            tutor_id=tutor_id,
            primary_student_id=student_id,
            grade=student.grade,
            board=student.board,
            status='scheduled',
            duration=60,  # Default 60 minutes
            created_by=current_user.id,
            created_at=datetime.utcnow()
        )
        
        db.session.add(new_class)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{student.full_name} assigned to {tutor.user.full_name}',
            'class_id': new_class.id
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Quick assign error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to assign student'
        }), 500


@bp.route('/api/allocation/bulk-auto-assign', methods=['POST'])
@login_required
@admin_required
def api_bulk_auto_assign():
    """API: Bulk auto-assign students to tutors"""
    try:
        data = request.get_json()
        filters = data.get('filters', {})
        dry_run = data.get('dry_run', True)
        
        result = allocation_helper.bulk_auto_assign(filters, dry_run)
        
        return jsonify(result)
        
    except Exception as e:
        current_app.logger.error(f"Bulk auto assign error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Bulk assignment failed'
        }), 500


@bp.route('/api/allocation/analytics')
@login_required
@admin_required
def api_allocation_analytics():
    """API: Get allocation analytics data"""
    try:
        analytics = allocation_helper.get_allocation_analytics()
        return jsonify({
            'success': True,
            'analytics': analytics
        })
        
    except Exception as e:
        current_app.logger.error(f"Analytics API error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Failed to fetch analytics'
        }), 500